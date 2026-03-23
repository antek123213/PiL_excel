from pathlib import Path

import pytest
from openpyxl import load_workbook


WORKSPACE_ROOT = Path(__file__).resolve().parent.parent
GENERATED_FILE = WORKSPACE_ROOT / "output" / "PL_TVM_2026-02.xlsx"

REQUIRED_HEADERS = [
    "Nr automatu",
    "Obrót brutto (zł)",
    "Liczba transakcji",
    "Prowizja (zł)",
]


@pytest.fixture(scope="module")
def worksheet():
    if not GENERATED_FILE.exists():
        pytest.skip(f"Brak wygenerowanego pliku: {GENERATED_FILE}")
    wb = load_workbook(GENERATED_FILE)
    return wb.active


def _header_index(worksheet):
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if isinstance(h, str)}


def test_generated_headers_contain_required_columns(worksheet):
    header_idx = _header_index(worksheet)
    missing = [h for h in REQUIRED_HEADERS if h not in header_idx]
    assert not missing, f"Brak wymaganych kolumn: {missing}"


def test_generated_has_data_rows(worksheet):
    assert worksheet.max_row >= 2, "Plik nie zawiera danych (brak wierszy po nagłówku)"


def test_generated_column_value_types_are_consistent(worksheet):
    header_idx = _header_index(worksheet)
    for row in range(2, worksheet.max_row + 1):
        device_id = worksheet.cell(row=row, column=header_idx["Nr automatu"]).value
        gross = worksheet.cell(row=row, column=header_idx["Obrót brutto (zł)"]).value
        tx_count = worksheet.cell(row=row, column=header_idx["Liczba transakcji"]).value
        commission = worksheet.cell(row=row, column=header_idx["Prowizja (zł)"]).value

        location = worksheet.cell(row=row, column=header_idx["Lokalizacja"]).value if "Lokalizacja" in header_idx else None
        status = worksheet.cell(row=row, column=header_idx["Status"]).value if "Status" in header_idx else ""

        assert isinstance(device_id, int), f"Wiersz {row}: Nr automatu powinien być int"
        if status is not None:
            assert isinstance(status, str), f"Wiersz {row}: Status powinien być tekstem"
        assert location is None or isinstance(location, str), f"Wiersz {row}: Lokalizacja powinna być tekstem lub pusta"

        assert isinstance(gross, (int, float)), f"Wiersz {row}: Obrót brutto powinien być liczbą"
        assert isinstance(tx_count, int), f"Wiersz {row}: Liczba transakcji powinna być int"

        # Prowizja może być pusta dla urządzeń bez danych prowizji.
        assert commission in (None, "") or isinstance(commission, (int, float)), (
            f"Wiersz {row}: Prowizja powinna być liczbą lub pusta"
        )


def test_generated_no_shifted_values_between_turnover_and_commission(worksheet):
    header_idx = _header_index(worksheet)
    # Jeżeli liczba transakcji jest > 0, obrót brutto powinien być nieujemny i liczbowy.
    # Test zabezpiecza przed przesunięciem wartości między kolumnami 4-6.
    for row in range(2, worksheet.max_row + 1):
        gross = worksheet.cell(row=row, column=header_idx["Obrót brutto (zł)"]).value
        tx_count = worksheet.cell(row=row, column=header_idx["Liczba transakcji"]).value
        commission = worksheet.cell(row=row, column=header_idx["Prowizja (zł)"]).value

        if isinstance(tx_count, int) and tx_count > 0:
            assert isinstance(gross, (int, float)), f"Wiersz {row}: Obrót brutto nie jest liczbą"
            assert gross >= 0, f"Wiersz {row}: Obrót brutto powinien być >= 0"

        if isinstance(commission, (int, float)):
            assert abs(commission) < 10_000_000, f"Wiersz {row}: Prowizja wygląda na przesuniętą/niepoprawną"
