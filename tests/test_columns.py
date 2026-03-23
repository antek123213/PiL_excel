"""
Test sprawdza czy export_to_excel_PL tworzy poprawne nagłówki:
    1-INFO: Nr automatu, Lokalizacja, Status
    2-OBRÓT: Obrót brutto (zł), Liczba transakcji
    3-PROWIZJA: Prowizja (zł)
    5/7/8: Koszty (zł), Suma roczna, Uwagi
"""
import sys
import os
import pytest
from openpyxl import load_workbook

# Dodaj folder nadrzędny do ścieżki
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from export_automaty import export_to_excel_PL


MONTH = "2026-02"
FILENAME = f"TEST_PL_{MONTH}.xlsx"

SAMPLE_DICT = {
    1101: {"status": "unchanged", "data": {"value": "TVM1", "description": "Automat 1101", "groupid": "G1"}},
    1102: {"status": "new",       "data": {"value": "TVM2", "description": "Automat 1102", "groupid": "G1"}},
}

SAMPLE_REVENUE = {
    1101: {
        "by_carrier": {
            "IC": {"obrot_brutto_zl": 1000.00, "liczba_transakcji": 30},
            "PR": {"obrot_brutto_zl": 234.56, "liczba_transakcji": 12},
        }
    },
}


@pytest.fixture(scope="module")
def excel_path(tmp_path_factory):
    """Tworzy plik Excel w katalogu tymczasowym."""
    tmp = tmp_path_factory.mktemp("output")
    
    # Patch OUTPUT_DIR w module
    import export_automaty
    original_dir = export_automaty.OUTPUT_DIR
    export_automaty.OUTPUT_DIR = tmp

    export_to_excel_PL(SAMPLE_DICT, SAMPLE_REVENUE, MONTH, FILENAME)

    export_automaty.OUTPUT_DIR = original_dir
    return tmp / FILENAME


@pytest.fixture(scope="module")
def headers(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    return [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]


# ── 1-INFO ─────────────────────────────────────────────────────────────────────
def test_col1_nr_automatu(headers):
    assert headers[0] == "Nr automatu", f"Kolumna 1: oczekiwano 'Nr automatu', got '{headers[0]}'"

def test_col2_lokalizacja(headers):
    assert headers[1] == "Lokalizacja", f"Kolumna 2: oczekiwano 'Lokalizacja', got '{headers[1]}'"

def test_col3_status(headers):
    assert headers[2] == "Status", f"Kolumna 3: oczekiwano 'Status', got '{headers[2]}'"

def test_col4_przewoznik(headers):
    assert headers[3] == "Przewoźnik", f"Kolumna 4: oczekiwano 'Przewoźnik', got '{headers[3]}'"

# ── 2-OBRÓT ────────────────────────────────────────────────────────────────────
def test_col4_obrot(headers):
    assert headers[4] == "Obrót brutto (zł)", f"Kolumna 5: oczekiwano 'Obrót brutto (zł)', got '{headers[4]}'"

def test_col5_liczba_transakcji(headers):
    assert headers[5] == "Liczba transakcji", f"Kolumna 6: oczekiwano 'Liczba transakcji', got '{headers[5]}'"


# ── 3-PROWIZJA ─────────────────────────────────────────────────────────────────
def test_col6_prowizja(headers):
    assert headers[6] == "Prowizja (zł)", f"Kolumna 7: oczekiwano 'Prowizja (zł)', got '{headers[6]}'"


# ── 5,7,8 ─────────────────────────────────────────────────────────────────────
def test_col7_koszty(headers):
    assert headers[7] == "Koszty (zł)", f"Kolumna 8: oczekiwano 'Koszty (zł)', got '{headers[7]}'"

def test_col8_suma_roczna(headers):
    assert headers[8] == "Suma roczna", f"Kolumna 9: oczekiwano 'Suma roczna', got '{headers[8]}'"

def test_col9_uwagi(headers):
    assert headers[9] == "Uwagi", f"Kolumna 10: oczekiwano 'Uwagi', got '{headers[9]}'"

def test_removed_columns_are_not_present(headers):
    assert "Value" not in headers
    assert "Description" not in headers
    assert "GroupID" not in headers
    assert "Rodzaj transakcji" not in headers
    assert "Wynik netto (zł)" not in headers

def test_headers_count_after_column_removal(headers):
    assert len(headers) == 10


# ── Dane w wierszach ────────────────────────────────────────────────────────────
def test_data_device_id(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    device_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert 1101 in device_ids
    assert 1102 in device_ids

def test_data_obrot(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    found_ic = False
    found_pr = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == 1101:
            carrier = ws.cell(row=row, column=4).value
            if carrier == "IC":
                assert ws.cell(row=row, column=5).value == pytest.approx(1000.00)
                found_ic = True
            if carrier == "PR":
                assert ws.cell(row=row, column=5).value == pytest.approx(234.56)
                found_pr = True

    assert found_ic, "Nie znaleziono wiersza dla 1101 + IC"
    assert found_pr, "Nie znaleziono wiersza dla 1101 + PR"
