import psycopg2
from psycopg2 import sql
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import json
import argparse
import time
from pathlib import Path
import re
import calendar
import unicodedata
import os
import sys
import hashlib

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


def _silent_print(*args, **kwargs):
    return None


if os.environ.get('PIL_SILENT', '').strip() == '1':
    print = _silent_print

# === KONFIGURACJA ===
DB_CONFIG = {
    'host': '192.168.101.20',
    'port': 5432,
    'database': 'postgres',  # Zmienione - będzie automatycznie wykrywana właściwa baza
    'user': 'sprzedaz',
    'password': 'tVregNm5',
    'sslmode': 'prefer'
}

# Domyślny schemat źródłowy danych operacyjnych.
SCHEMA='ARP'
SNAPSHOT_DIR = Path(__file__).parent / 'snapshots'
MONTH_PAYLOAD_CACHE_DIR = SNAPSHOT_DIR / 'cache'
OUTPUT_DIR = Path(__file__).parent / 'output'
DEFAULT_COSTS_FILE = Path(__file__).parent / 'Koszty' / 'P&L_koszty_ROP.xlsx'
DEFAULT_RENT_FILE = Path(__file__).parent / 'Koszty' / 'Najem powierzchni 2025-2026.xlsx'
DEFAULT_AMORTYZACJA_2025_FILE = Path(__file__).parent / 'Koszty' / 'amortyzacja_25.xlsx'
DEFAULT_AMORTYZACJA_2026_FILE = Path(__file__).parent / 'Koszty' / 'amortyzacja_26.xlsx'
DEFAULT_AMORTYZACJA_FILE = DEFAULT_AMORTYZACJA_2026_FILE
DEFAULT_PROWIZJE_FILE = Path(__file__).parent / 'Prowizje_AB.xlsx'
DEFAULT_LISTA_AUTOMATOW_FILE = Path(__file__).parent / 'lista automatów.xlsx'
DEFAULT_RELOKACJE_FILE = Path(__file__).parent / 'Relokacje AB.txt'
DEFAULT_SERWIS_FILE = Path(__file__).parent / 'Koszty' / 'serwis_2026.xlsx'
DEFAULT_IT_CARD_SWITCH_FILE = Path(__file__).parent / 'Koszty' / 'IT CARD.xlsx'
DEFAULT_SERWIS_SHEET_INDEX = 4
DEFAULT_AMORTYZACJA_SHEET = None
LOCATION_CACHE_STRATEGY_VERSION = '2026-04-19-relokacje-magazyn-rop-cleanup'
MONTH_PAYLOAD_CACHE_VERSION = '2026-04-22-contract-start-filter-v1'
DEVICE_ID_RANGES = (
    (1101, 1141),
    (1201, 1275),
    (1276, 1278),
    (1286, 1299),
)

CARRIERS = ['ARP', 'IC', 'KD', 'KML', 'KS', 'KW', 'LKA', 'PR', 'SKM']
DISALLOWED_CARRIERS = {'KM'}
TRANSACTIONS_SOURCE_CONFIG = {
    'table': 'transactions',
    'amount_col': 'fin_nalezn',
    'date_col': 'fin_data_sp',
    'device_col': 'tvm_tvm_id',
    'payment_method_col': 'fin_spos_opl',
}
TRANSACTIONS_TICKET_COUNT_COL = 'fin_l_bil'
TRANSACTIONS_TABLE_OVERRIDES = {
    'IC': 'transactionsns',
}
TRANSACTIONS_DEVICE_COL_OVERRIDES = {
    # User-required precedence: tvm_aitomatnum first, then tvm_automatnum.
    'IC': ['tvm_aitomatnum', 'tvm_automatnum'],
}
TRANSACTIONS_AMOUNT_COL_OVERRIDES = {
}
TRANSACTIONS_PAYMENT_METHOD_COL_OVERRIDES = {
}
TRANSACTIONS_TICKET_COUNT_COL_OVERRIDES = {
}
TRANSACTIONS_PAYMENT_METHOD_COL_CANDIDATES = [
    'fin_spos_opl',
    'spos_opl',
    'payment_method',
    'payment_method_code',
]
PAYMENT_METHOD_CODE_TO_KEY = {
    '1': 'gotowka',
    '2': 'karta',
    '6': 'blik',
}
PAYMENT_METHOD_KEYS = ('gotowka', 'karta', 'blik')
PAYMENT_METHOD_LABELS = {
    'gotowka': 'Gotówka',
    'karta': 'Karta',
    'blik': 'BLIK',
}
INTERCHANGE_RATE_BY_CARRIER = {
    'KD': 0.0154,
    'KML': 0.0135,
    'LKA': 0.0154,
    'ARP': 0.0154,
    'PR': 0.0154,
}
DEFAULT_INTERCHANGE_RATE = 0.0135
IT_CARD_RATE_BY_CARRIER = INTERCHANGE_RATE_BY_CARRIER
DEFAULT_IT_CARD_RATE = DEFAULT_INTERCHANGE_RATE
TVM_COST_KEYS = (
    'czynsz',
    'prad',
    'elavon',
    'poczta_polska',
    'amortyzacja',
    'utrzymanie_oprogramowania',
    'papier',
    'transmisja_danych',
    'serwis',
    'it_card',
    'ubezpieczenie',
)
TVM_COST_LABELS = {
    'czynsz': 'czynsz',
    'prad': 'prad',
    'elavon': 'elavon',
    'poczta_polska': 'poczta polska',
    'amortyzacja': 'amortyzacja',
    'utrzymanie_oprogramowania': 'utrzymanie oprogramowania',
    'papier': 'papier',
    'transmisja_danych': 'transmisja danych',
    'serwis': 'serwis',
    'it_card': 'IT card',
    'ubezpieczenie': 'ubezpieczenie',
}
OTHER_COST_KEYS = ('non_tvm', 'project_variable_costs', 'oh')
OTHER_COST_LABELS = {
    'non_tvm': 'NON TVM',
    'project_variable_costs': 'Project Variable Costs',
    'oh': 'OH',
}
ROP_EXCLUDED_DEVICE_IDS = set()
INFOLINIA_MONTHLY_COST = 2650.0
INFOLINIA_START_MONTH = '2026-03'
ADDITIONAL_PROFIT_RULES = {
    ('2026-02', 1122): 4500.0,
}
TRANSACTIONS_DEVICE_COL_CANDIDATES = [
    'fin_nr_urz',
    'nr_urz',
    'tvm_id',
    'fin_tvm_id',
    'device_id',
    'id_automatu',
    'automat_id',
    'ms_deviceid',
]
TRANSACTIONS_CARRIER_COL_CANDIDATES = [
    'provider',
    'przewoznik',
    'carrier',
    'operator',
    'source',
    'payment_operator',
    'fin_przewoznik',
]

# Konfiguracja źródła prowizji - można nadpisać parametrami CLI
COMMISSION_SOURCE_CONFIG = {
    'schema': 'public',
    'table': 'provision',
    'device_col': 'tvm_id_prov',
    'amount_col': 'provision_prov',
    'date_col': 'valid_from',
    'date_to_col': 'valid_to',
    'provider_col': None,
    'provider_values': None
}


def _end_of_month_date(month_str):
    """
    Zwraca ostatni dzień miesiąca jako datetime (00:00:00).
    """
    year, month = map(int, month_str.split('-'))
    last_day = calendar.monthrange(year, month)[1]
    return datetime(year, month, last_day)


def build_output_filename(month_str, naming_mode='default'):
    """
    Buduje nazwę pliku wyjściowego.
    - default: PL_TVM_YYYY-MM.xlsm
    - monitor-style: PROVISIONYYYYMMDDHHMMSS.xlsm
    """
    if naming_mode == 'monitor-style':
        ts = datetime.now().strftime('%Y%m%d%H%M%S')
        return f"PROVISION{ts}.xlsm"
    return f"PL_TVM_{month_str}.xlsm"


# === OBSŁUGA VBA I KONWERSJI NA .XLSM ===

def _load_vba_code_from_file(vba_file_path):
    """
    Ładuje kod VBA z pliku .vba.txt
    """
    candidates = [Path(vba_file_path)]
    # Some files in the repo may have an extra .txt suffix (e.g. .vba.txt.txt)
    p = Path(vba_file_path)
    if p.suffix == '.txt':
        candidates.append(p.with_name(p.name + '.txt'))
    # Also try removing a trailing .txt (in case user renamed differently)
    if p.name.endswith('.txt.txt'):
        candidates.append(p.with_name(p.name.replace('.txt.txt', '.txt')))

    for candidate in candidates:
        try:
            with open(str(candidate), 'r', encoding='utf-8') as f:
                print(f"✓ Załadowano VBA z: {candidate}")
                return f.read()
        except FileNotFoundError:
            continue

    print(f"⚠️ Plik VBA nie znaleziony: {vba_file_path}")
    return ""


def _get_vba_modules_dict():
    """
    Przygotowuje słownik modułów VBA do załadowania.
    Ładuje kod z plików .vba.txt ze ścieżki scripts/vba/
    """
    vba_dir = Path(__file__).parent / 'scripts' / 'vba'
    
    vba_wyszukiwarka_file = vba_dir / 'Wyszukiwarka_Arkusz.vba.txt'
    vba_helpers_file = vba_dir / 'WyszukiwarkaHelpers.vba.txt'
    
    vba_wyszukiwarka_code = _load_vba_code_from_file(vba_wyszukiwarka_file)
    vba_helpers_code = _load_vba_code_from_file(vba_helpers_file)
    
    return {
        'WyszukiwarkaCode': {
            'code': vba_wyszukiwarka_code,
            'type': 'WorksheetModule',
            'target_sheet': 'Wyszukiwarka'
        },
        'WyszukiwarkaHelpers': {
            'code': vba_helpers_code,
            'type': 'StandardModule'
        }
    }


def _convert_xlsx_to_xlsm_with_vba(xlsx_path, xlsm_path, vba_modules_dict):
    """
    Konwertuje .xlsx → .xlsm i wstawia kod VBA za pomocą COM (pywin32).
    
    vba_modules_dict: {
        'module_name': {
            'code': 'VBA code string',
            'type': 'StandardModule' | 'WorksheetModule',
            'target_sheet': 'Wyszukiwarka'  # dla WorksheetModule
        }
    }
    """
    try:
        import win32com.client as win32
    except ImportError:
        print("⚠️ pywin32 nie zainstalowany. Pomijam VBA. Zainstaluj: pip install pywin32")
        import shutil
        shutil.copy(str(xlsx_path), str(xlsm_path))
        os.remove(str(xlsx_path))
        return
    
    excel = None
    try:
        excel = win32.Dispatch("Excel.Application")
        try:
            excel.Visible = False
        except Exception as e:
            print(f"⚠️ Nie udało się ustawić Visible=False: {e}")
        try:
            excel.DisplayAlerts = False
        except Exception as e:
            print(f"⚠️ Nie udało się ustawić DisplayAlerts=False: {e}")
        
        wb = excel.Workbooks.Open(str(xlsx_path))

        try:
            _ = wb.VBProject.VBComponents.Count
        except Exception as e:
            trust_hint = "Programistyczny dostęp do projektu w języku Visual Basic nie jest zaufany"
            if trust_hint in str(e):
                print(
                    "⚠️ Excel blokuje dostęp do VBProject. Aby wstrzyknąć VBA, włącz w Excelu: "
                    "Plik -> Opcje -> Centrum zaufania -> Ustawienia Centrum zaufania -> Ustawienia makr -> "
                    "'Zaufaj dostępowi do modelu obiektowego projektu VBA'."
                )
                if os.path.exists(str(xlsm_path)):
                    os.remove(str(xlsm_path))
                wb.SaveAs(str(xlsm_path), FileFormat=52)
                wb.Close()
                print(f"✓ Zapisano plik .xlsm bez wstrzykniętego VBA: {xlsm_path}")
                os.remove(str(xlsx_path))
                return
            raise
        
        for module_name, module_info in vba_modules_dict.items():
            vba_code = module_info['code']
            if not vba_code:
                print(f"⚠️ Brak kodu dla modułu '{module_name}', pomijam")
                continue
            
            module_type = module_info.get('type', 'StandardModule')
            
            if module_type == 'WorksheetModule':
                sheet_name = module_info.get('target_sheet', 'Wyszukiwarka')
                try:
                    ws = wb.Sheets(sheet_name)
                    ws.Activate()
                    wb.VBProject.VBComponents(ws.CodeName).CodeModule.AddFromString(vba_code)
                    print(f"✓ Dodano kod do arkusza '{sheet_name}'")
                except Exception as e:
                    if 'Programistyczny dostęp do projektu w języku Visual Basic nie jest zaufany' in str(e):
                        print(
                            "⚠️ Nie można dodać kodu do arkusza, bo Excel blokuje dostęp do VBProject. "
                            "Plik zostanie zapisany bez VBA."
                        )
                        break
                    print(f"⚠️ Nie udało się dodać do '{sheet_name}': {e}")
            else:
                try:
                    vb_component = wb.VBProject.VBComponents.Add(1)
                    vb_component.Name = module_name
                    vb_component.CodeModule.AddFromString(vba_code)
                    print(f"✓ Dodano moduł '{module_name}'")
                except Exception as e:
                    if 'Programistyczny dostęp do projektu w języku Visual Basic nie jest zaufany' in str(e):
                        print(
                            "⚠️ Nie można dodać modułu VBA, bo Excel blokuje dostęp do VBProject. "
                            "Plik zostanie zapisany bez VBA."
                        )
                        break
                    print(f"⚠️ Nie udało się dodać '{module_name}': {e}")
        
        # Ensure no workbook with the same name is open in this Excel instance
        try:
            target_name = Path(xlsm_path).name
            for w in list(excel.Workbooks):
                try:
                    if w.Name == target_name and w.FullName != str(xlsx_path):
                        w.Close(SaveChanges=False)
                except Exception:
                    pass
        except Exception:
            pass

        # Remove existing file on disk if possible
        if os.path.exists(str(xlsm_path)):
            try:
                os.remove(str(xlsm_path))
            except PermissionError:
                print(f"⚠️ Błąd: Zamknij plik {xlsm_path} przed ponownym generowaniem!")
                wb.Close(SaveChanges=False)
                return

        # Try multiple save strategies to avoid COM SaveAs failures
        saved = False
        try:
            wb.SaveAs(str(xlsm_path), FileFormat=52)
            saved = True
        except Exception as e_save:
            print(f"⚠️ SaveAs nie powiodło się: {e_save}. Spróbuję SaveCopyAs...")
            try:
                wb.SaveCopyAs(str(xlsm_path))
                saved = True
            except Exception as e_copy:
                print(f"⚠️ SaveCopyAs nie powiodło się: {e_copy}. Spróbuję ponownie z widocznym Excelem...")
                try:
                    excel.DisplayAlerts = True
                    excel.Visible = True
                    wb.SaveAs(str(xlsm_path), FileFormat=52)
                    saved = True
                except Exception as e2:
                    print(f"❌ Wszystkie metody zapisu nie powiodły się: {e2}")
                    import shutil
                    try:
                        shutil.copy(str(xlsx_path), str(xlsm_path))
                        saved = True
                        print("⚠️ Uwaga: zapisano kopię bez gwarancji wstrzyknięcia VBA (shutil.copy).")
                    except Exception as e3:
                        print(f"❌ Fallback (shutil.copy) też nie powiódł się: {e3}")

        # Close workbook and cleanup
        try:
            wb.Close(SaveChanges=False)
        except Exception:
            try:
                wb.Close()
            except Exception:
                pass

        if saved:
            print(f"✓ Plik .xlsm z VBA: {xlsm_path}")
            try:
                os.remove(str(xlsx_path))
            except Exception:
                pass
        else:
            raise Exception("Nie udało się zapisać pliku .xlsm; wszystkie metody zawiodły.")
        
    except Exception as e:
        print(f"❌ Błąd konwersji: {e}")
        raise
    finally:
        if excel:
            try:
                quit_method = getattr(excel, 'Quit', None)
                if callable(quit_method):
                    quit_method()
            except Exception:
                pass


def _build_device_filter_sql(column_sql):
    """
    Buduje warunek SQL filtrujący urządzenia wg ustalonych zakresów ID.
    """
    clauses = []
    for start_id, end_id in DEVICE_ID_RANGES:
        clauses.append(
            sql.SQL("({column} BETWEEN {start_id} AND {end_id})").format(
                column=column_sql,
                start_id=sql.Literal(start_id),
                end_id=sql.Literal(end_id),
            )
        )
    return sql.SQL("(") + sql.SQL(" OR ").join(clauses) + sql.SQL(")")


def _build_device_filter_preview_expr(column_name):
    """
    Buduje tekstowy warunek SQL zakresow urzadzen do podgladu/debugu.
    """
    parts = [
        f"({column_name}::bigint BETWEEN {start_id} AND {end_id})"
        for start_id, end_id in DEVICE_ID_RANGES
    ]
    return "(" + " OR ".join(parts) + ")"


def extract_city_name(description):
    """
    Wyciąga nazwę miasta z pola description, usuwając numery automatów.
    Przykłady:
      'WROC-GL_1204'  -> 'WROC-GL'
      '1214_RZEPIN'   -> 'RZEPIN'
      'KRAK_GN_1263'  -> 'KRAK-GN'
      '2100779032'    -> ''  (czysto numeryczne)
      '1203: Opis'    -> ''  (placeholder)
    """
    if not description:
        return ''
    normalized = str(description).strip()
    if 'opis' in normalized.lower():
        return ''

    parts = normalized.split('_')
    letter_parts = [p for p in parts if re.search(r'[A-Za-z]', p)]
    result = '-'.join(letter_parts)
    if result in ('Opis', 'opis', ''):
        return ''
    return result


def _normalize_tvm_location_text(value):
    """
    Normalizuje pełny tekst lokalizacji z pól TVM.
    """
    if value is None:
        return ''
    normalized = str(value).strip()
    if not normalized:
        return ''
    if 'opis' in normalized.lower():
        return ''
    return normalized


def _resolve_service_file_for_month(month_str, service_file=None):
    """
    Wybiera plik serwisu dla danego miesiąca.

    Priorytet:
    1. jawny plik z CLI
    2. plik miesięczny w Koszty/Serwis AB_MM.YYYY.xlsx
    3. brak override'ów, jeśli nie ma pasującego pliku
    """
    if service_file is not None:
        service_path = Path(service_file)
        return service_path if service_path.exists() else None

    dt = datetime.strptime(month_str, '%Y-%m')
    year = dt.year
    candidates = [
        Path(__file__).parent / 'Koszty' / 'serwis_2026.xlsx',
        Path(__file__).parent / f'serwis_{year}.xlsx',
        Path(__file__).parent / 'Koszty' / f'serwis_{year}.xlsx',
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def _resolve_lista_automatow_file_for_month(month_str, lista_file=None):
    """
    Wybiera plik listy automatow dla danego miesiaca.

    Priorytet:
    1. jawny plik z CLI
    2. pliki domyslne w katalogu projektu/Koszty
    """
    if lista_file is not None:
        lista_path = Path(lista_file)
        if lista_path.exists():
            return lista_path

    candidates = [
        Path(__file__).parent / 'lista_automat\u00f3w.xlsx',
        Path(__file__).parent / 'lista_automatow.xlsx',
        Path(__file__).parent / 'lista automatów.xlsx',
        Path(__file__).parent / 'lista automatow.xlsx',
        Path(__file__).parent / 'lista automat\u00f3w.xlsx',
        Path(__file__).parent / 'Koszty' / 'lista_automat\u00f3w.xlsx',
        Path(__file__).parent / 'Koszty' / 'lista_automatow.xlsx',
        Path(__file__).parent / 'Koszty' / 'lista automatów.xlsx',
        Path(__file__).parent / 'Koszty' / 'lista automatow.xlsx',
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate

    return None


def _load_lista_automatow_locations_from_xlsx(lista_file=DEFAULT_LISTA_AUTOMATOW_FILE, month_str=None):
    """
    Wczytuje lokalizacje/stacje per automat z pliku lista automatow.
    Preferuje kolumne "Nazwa stacji", fallback: "Adres".
    """
    result = {}

    if lista_file is None:
        return result

    lista_path = Path(lista_file)
    if not lista_path.exists():
        print(f"WARN Lista automatow nie istnieje: {lista_path}")
        return result

    try:
        wb = load_workbook(lista_path, data_only=True)
    except Exception as e:
        print(f"WARN Lista automatow: nie mozna otworzyc pliku {lista_path}: {e}")
        return result

    sheet_name = _resolve_sheet_name_by_month(wb, month_str or '2025-01')
    if sheet_name is None:
        try:
            wb.close()
        except Exception:
            pass
        return result

    ws = wb[sheet_name]

    header_row = None
    device_col = None
    station_col = None
    address_col = None

    scan_rows = min(ws.max_row, 60)
    for r in range(1, scan_rows + 1):
        row_map = {}
        for c in range(1, min(ws.max_column, 80) + 1):
            key = _normalize_header_text(ws.cell(row=r, column=c).value)
            if key:
                row_map[key] = c

        if not row_map:
            continue

        detected_device_col = _find_header_col(
            row_map,
            lambda key: key in {'NRTVM', 'NRAUTOMATU', 'NUMERAUTOMATU', 'TVM', 'IDAUTOMATU'}
        )
        if detected_device_col is None:
            continue

        detected_station_col = _find_header_col(
            row_map,
            lambda key: ('NAZWASTACJI' in key) or (key in {'STACJA', 'NAZWAPRZYSTANKU'})
        )
        detected_address_col = _find_header_col(
            row_map,
            lambda key: ('ADRES' in key) or ('LOKALIZACJA' in key)
        )

        if detected_station_col is None and detected_address_col is None:
            continue

        header_row = r
        device_col = detected_device_col
        station_col = detected_station_col
        address_col = detected_address_col
        break

    if header_row is None or device_col is None:
        print(f"WARN Lista automatow: nie znaleziono naglowka Nr TVM ({lista_path.name})")
        try:
            wb.close()
        except Exception:
            pass
        return result

    for r in range(header_row + 1, ws.max_row + 1):
        device_id = _parse_device_id(ws.cell(row=r, column=device_col).value)
        if device_id is None:
            continue

        station_value = ''
        address_value = ''
        if station_col is not None:
            station_value = _normalize_tvm_location_text(ws.cell(row=r, column=station_col).value)
        if address_col is not None:
            address_value = _normalize_tvm_location_text(ws.cell(row=r, column=address_col).value)

        location_value = station_value or address_value
        if location_value:
            result[int(device_id)] = location_value

    try:
        wb.close()
    except Exception:
        pass

    print(f"Wczytano lokalizacje z listy automatow: {len(result)}")
    return result


def _load_lista_automatow_models_from_xlsx(lista_file=DEFAULT_LISTA_AUTOMATOW_FILE):
    """
    Wczytuje model automatu z pliku lista automatow, arkusz "1".
    Oczekiwane kolumny: "Nr TVM" oraz "Model".
    """
    result = {}

    if lista_file is None:
        return result

    lista_path = Path(lista_file)
    if not lista_path.exists():
        print(f"WARN Lista automatow nie istnieje: {lista_path}")
        return result

    try:
        wb = load_workbook(lista_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"WARN Lista automatow: nie mozna otworzyc pliku {lista_path}: {e}")
        return result

    try:
        sheet_name = next((name for name in wb.sheetnames if str(name).strip() == '1'), None)
        if sheet_name is None:
            print(f"WARN Lista automatow: nie znaleziono arkusza '1' ({lista_path.name})")
            return result

        ws = wb[sheet_name]

        header_row = None
        device_col = None
        model_col = None

        scan_rows = min(ws.max_row, 60)
        for r in range(1, scan_rows + 1):
            row_map = {}
            for c in range(1, min(ws.max_column, 80) + 1):
                key = _normalize_header_text(ws.cell(row=r, column=c).value)
                if key:
                    row_map[key] = c

            if not row_map:
                continue

            detected_device_col = _find_header_col(
                row_map,
                lambda key: key in {'NRTVM', 'NRAUTOMATU', 'NUMERAUTOMATU', 'TVM', 'IDAUTOMATU'}
            )
            detected_model_col = _find_header_col(row_map, lambda key: key == 'MODEL')

            if detected_device_col is not None and detected_model_col is not None:
                header_row = r
                device_col = detected_device_col
                model_col = detected_model_col
                break

        if header_row is None or device_col is None or model_col is None:
            print(f"WARN Lista automatow: nie znaleziono kolumn Nr TVM/Model ({lista_path.name})")
            return result

        for r in range(header_row + 1, ws.max_row + 1):
            device_id = _parse_device_id(ws.cell(row=r, column=device_col).value)
            if device_id is None:
                continue

            model_value = ws.cell(row=r, column=model_col).value
            if _has_nonempty_value(model_value):
                result[int(device_id)] = str(model_value).strip()

        print(f"Wczytano modele z listy automatow: {len(result)}")
        return result
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _load_automat_type_map_from_lista(month_str, lista_file=None, included_device_ids=None):
    """
    Zwraca mapę typu/modelu automatu z lista_automatow dla raportu.
    """
    resolved_lista_file = _resolve_lista_automatow_file_for_month(month_str, lista_file)
    model_map = _load_lista_automatow_models_from_xlsx(lista_file=resolved_lista_file)
    if included_device_ids is None:
        return model_map

    included_ids = {int(device_id) for device_id in (included_device_ids or [])}
    return {
        int(device_id): model
        for device_id, model in model_map.items()
        if int(device_id) in included_ids
    }


# === FUNKCJE POMOCNICZE: DIAGNOSTYKA BAZY DANYCH ===

def list_available_databases(host, port, user, password):
    """
    Listuje dostępne bazy danych na serwerze PostgreSQL.
    """
    try:
        conn = psycopg2.connect(
            host=host,
            port=port,
            database='postgres',  # Domyślna baza systemowa
            user=user,
            password=password,
            sslmode=DB_CONFIG.get('sslmode', 'prefer'),
            connect_timeout=10
        )
        cursor = conn.cursor()
        cursor.execute("SELECT datname FROM pg_database WHERE datistemplate = false ORDER BY datname;")
        databases = [row[0] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return databases
    except psycopg2.Error as e:
        print(f"❌ Nie można pobrać listy baz danych: {e}")
        return []


def database_has_required_tables(database_name, required_tables=None):
    """
    Sprawdza czy baza zawiera wymagane tabele raportowe.
    """
    if required_tables is None:
        required_tables = [
            (SCHEMA, 'dictionary'),
            (SCHEMA, 'moneystats'),
        ]

    try:
        conn = psycopg2.connect(
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port'],
            database=database_name,
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            sslmode=DB_CONFIG.get('sslmode', 'prefer'),
            connect_timeout=5
        )
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT table_schema, table_name
            FROM information_schema.tables
            WHERE table_type = 'BASE TABLE'
              AND table_schema = %s
              AND table_name = ANY(%s)
            """,
            (SCHEMA, [tbl for _, tbl in required_tables]),
        )
        existing = {(row[0], row[1]) for row in cursor.fetchall()}
        cursor.close()
        conn.close()
        return all(t in existing for t in required_tables)
    except psycopg2.Error:
        return False


def connect_to_db(database_name=None, max_retries=3, retry_delay=5):
    """
    Łączy się z bazą danych z mechanizmem retry.
    Jeśli database_name nie podane, automatycznie wykrywa bazę 'monitor'.
    """
    # Sprawdź dostępne bazy danych
    if database_name is None:
        print("\n🔍 Sprawdzanie dostępnych baz danych...")
        databases = list_available_databases(
            DB_CONFIG['host'],
            DB_CONFIG['port'],
            DB_CONFIG['user'],
            DB_CONFIG['password']
        )
        
        if databases:
            print(f"✓ Dostępne bazy danych: {', '.join(databases)}")
            
            # Szukaj bazy danych z nazwą "monitor"
            monitor_dbs = [db for db in databases if 'monitor' in db.lower()]
            if monitor_dbs:
                preferred_monitor = None
                for candidate_db in monitor_dbs:
                    if database_has_required_tables(candidate_db):
                        preferred_monitor = candidate_db
                        break

                if preferred_monitor:
                    database_name = preferred_monitor
                    print(f"✓ Wybrano bazę danych: {database_name} (z wymaganymi tabelami {SCHEMA})")
                else:
                    print(f"⚠ Baza monitor nie zawiera kompletu {SCHEMA}.dictionary + {SCHEMA}.moneystats")
                    print("  Szukanie alternatywnej bazy z wymaganymi tabelami...")
                    preferred_any = None
                    for candidate_db in databases:
                        if candidate_db in ['postgres', 'template0', 'template1']:
                            continue
                        if database_has_required_tables(candidate_db):
                            preferred_any = candidate_db
                            break

                    if preferred_any:
                        database_name = preferred_any
                        print(f"✓ Wybrano alternatywną bazę danych: {database_name}")
                    else:
                        database_name = monitor_dbs[0]
                        print(f"⚠ Używam bazy monitor bez pełnego zestawu tabel: {database_name}")
            else:
                print(f"⚠ Brak bazy 'monitor'. Dostępne: {', '.join(databases)}")
                # Spróbuj pierwszej bazy z listy (pomijając postgres/template)
                non_system = [db for db in databases if db not in ['postgres', 'template0', 'template1']]
                if non_system:
                    database_name = non_system[0]
                    print(f"⚠ Używam pierwszej dostępnej bazy: {database_name}")
                else:
                    print("❌ Brak dostępnych baz danych użytkownika")
                    return None
        else:
            print("⚠ Nie można pobrać listy baz danych")

            fallback_candidates = []
            for candidate in [
                'Monitor',
                'monitor',
                DB_CONFIG.get('database'),
                'postgres',
            ]:
                if not candidate:
                    continue
                if candidate not in fallback_candidates:
                    fallback_candidates.append(candidate)

            preferred_fallback = None
            for candidate_db in fallback_candidates:
                if database_has_required_tables(candidate_db):
                    preferred_fallback = candidate_db
                    break

            if preferred_fallback:
                database_name = preferred_fallback
                print(f"✓ Fallback: wybrano bazę danych: {database_name} (z wymaganymi tabelami {SCHEMA})")
            elif fallback_candidates:
                database_name = fallback_candidates[0]
                print(f"⚠ Fallback: próba połączenia z bazą: {database_name}")
            else:
                print("❌ Brak kandydatów do fallback połączenia DB")
                return None
    
    # Połącz się z wybraną bazą
    for attempt in range(1, max_retries + 1):
        try:
            conn = psycopg2.connect(
                host=DB_CONFIG['host'],
                port=DB_CONFIG['port'],
                database=database_name,
                user=DB_CONFIG['user'],
                password=DB_CONFIG['password'],
                sslmode=DB_CONFIG.get('sslmode', 'prefer'),
                connect_timeout=10
            )
            if attempt > 1:
                print(f"✓ Połączono z bazą '{database_name}' (próba {attempt}/{max_retries})")
            return conn
        except psycopg2.OperationalError as e:
            if attempt < max_retries:
                print(f"⚠ Próba {attempt}/{max_retries} nie powiodła się: {e}")
                print(f"  Ponowna próba za {retry_delay} sekund...")
                time.sleep(retry_delay)
            else:
                print(f"❌ Nie udało się połączyć z bazą '{database_name}' po {max_retries} próbach")
                raise
    return None


# === 1-INFO: SŁOWNIK AUTOMATÓW (ARP.dictionary) ===

def get_dictionary_snapshot(conn, schema=SCHEMA):
    """
    Pobiera słownik automatów z <schema>.dictionary.
    Zwraca: dict {device_id: {'description': description}}
    Filtruje tylko rzeczywiste ID automatów (numeric) z zakresów:
    1101-1141, 1201-1275, 1276-1278, 1286-1299.
    """
    cursor = conn.cursor()
    query = sql.SQL(
        """
                SELECT id, description
                FROM {schema}.dictionary
                WHERE id::text ~ '^[0-9]+$'
                    AND {device_filter}
                ORDER BY id
                """
        ).format(
            schema=sql.Identifier(schema),
            device_filter=_build_device_filter_sql(sql.SQL("id::bigint")),
        )
    cursor.execute(query)
    rows = cursor.fetchall()
    cursor.close()
    
    snapshot = {}
    for row in rows:
        device_id, description = row
        snapshot[int(device_id)] = {
            'description': description,
        }
    
    return snapshot


def _extract_automat_type_suffix(groupid_value):
    """
    Wyciąga suffix typu automatu z pola groupid, np. AUTABB -> BB.
    """
    normalized = str(groupid_value or '').strip().upper()
    if not normalized.startswith('AUTA'):
        return ''
    return normalized[4:]


def get_automat_type_by_device_from_av(
    conn,
    device_ids,
    schema_name='AV',
    table_name='dictionary',
    value_col='value',
    groupid_col='groupid',
):
    """
    Pobiera typ automatu z AV.dictionary po kolumnie `value` i zwraca mapę:
    {device_id: 'BB' | ...}.
    """
    normalized_ids = sorted({int(device_id) for device_id in device_ids})
    if not normalized_ids:
        return {}

    columns = _get_table_columns(conn, schema_name, table_name)
    if not columns:
        print(f"⚠ Brak tabeli typu automatu: {schema_name}.{table_name}")
        return {}

    lowered = {c.lower(): c for c in columns}
    resolved_value_col = lowered.get(str(value_col).lower())
    resolved_groupid_col = lowered.get(str(groupid_col).lower())
    if resolved_value_col is None or resolved_groupid_col is None:
        print(
            f"⚠ Brak kolumn typu automatu w {schema_name}.{table_name}: "
            f"{value_col}, {groupid_col}"
        )
        return {}

    search_values = [str(device_id) for device_id in normalized_ids]
    cursor = conn.cursor()
    query = sql.SQL(
        """
        SELECT {value_col}, {groupid_col}
        FROM {schema}.{table}
        WHERE {value_col}::text = ANY(%s)
        """
    ).format(
        value_col=sql.Identifier(resolved_value_col),
        groupid_col=sql.Identifier(resolved_groupid_col),
        schema=sql.Identifier(schema_name),
        table=sql.Identifier(table_name),
    )
    cursor.execute(query, (search_values,))
    rows = cursor.fetchall()
    cursor.close()

    type_map = {}
    for raw_value, groupid in rows:
        device_id = _parse_device_id(raw_value)
        if device_id is None:
            continue
        type_map[device_id] = _extract_automat_type_suffix(groupid)

    return type_map


def _resolve_dictionary_location_source(conn, schema_name, table_name='dictionary'):
    """
    Rozpoznaje kolumny słownika dla lokalizacji (id, tvm_tvmlocation1, tvm_tvmlocation2, description).
    Zwraca dict lub None, gdy tabela/kolumny nie są dostępne.
    """
    columns = _get_table_columns(conn, schema_name, table_name)
    if not columns:
        return None

    lowered_map = {c.lower(): c for c in columns}
    device_col = lowered_map.get('id')
    tvm_location1_col = lowered_map.get('tvm_tvmlocation1')
    tvm_location2_col = lowered_map.get('tvm_tvmlocation2')
    description_col = lowered_map.get('description')

    if device_col is None:
        return None
    if tvm_location1_col is None and tvm_location2_col is None and description_col is None:
        return None

    return {
        'schema': schema_name,
        'table': table_name,
        'device_col': device_col,
        'tvm_location1_col': tvm_location1_col,
        'tvm_location2_col': tvm_location2_col,
        'description_col': description_col,
    }


def _pick_location_value(tvm_location1, tvm_location2, description):
    """
    Wybiera lokalizację z priorytetem:
    tvm_tvmlocation1 -> tvm_tvmlocation2 -> description.

    Dla pól TVM zachowuje pełny adres.
    Description jest używane tylko, gdy oba pola TVM są puste.
    Zwraca pusty string, gdy brak sensownej wartości.
    """
    normalized_1 = _normalize_tvm_location_text(tvm_location1)
    normalized_2 = _normalize_tvm_location_text(tvm_location2)

    if normalized_1 and normalized_2:
        return f"{normalized_1}, {normalized_2}"
    if normalized_1:
        return normalized_1
    if normalized_2:
        return normalized_2

    return extract_city_name(description)


def get_locations_by_carrier(
    conn,
    device_ids,
    carriers=None,
    dictionary_table='dictionary',
):
    """
    Uzupełnia lokalizacje automatów sekwencyjnie po przewoźnikach.
    Priorytet pól: tvm_tvmlocation1 -> tvm_tvmlocation2 -> description.
    Dla automatu, który już ma lokalizację, kolejne schematy są pomijane.

    Zwraca dict {device_id: 'lokalizacja'}.
    """
    carriers = carriers or CARRIERS
    missing_device_ids = {int(device_id) for device_id in device_ids}
    location_by_device = {}

    for carrier in carriers:
        if not missing_device_ids:
            break

        carrier_code = _normalize_carrier_code(carrier)
        source = _resolve_dictionary_location_source(
            conn,
            schema_name=carrier_code,
            table_name=dictionary_table,
        )
        if source is None:
            print(
                f"⚠ Pomijam lokalizacje dla {carrier} (schemat {carrier_code}): "
                f"brak źródła {carrier_code}.{dictionary_table} lub wymaganych kolumn"
            )
            continue

        selected_columns = [
            sql.Identifier(source['device_col']),
            sql.Identifier(source['tvm_location1_col']) if source['tvm_location1_col'] else sql.SQL('NULL'),
            sql.Identifier(source['tvm_location2_col']) if source['tvm_location2_col'] else sql.SQL('NULL'),
            sql.Identifier(source['description_col']) if source['description_col'] else sql.SQL('NULL'),
        ]

        query = sql.SQL(
            """
                        SELECT {device_col}, {tvm_location1_col}, {tvm_location2_col}, {description_col}
            FROM {schema}.{table}
            WHERE {device_col}::text ~ '^[0-9]+$'
              AND {device_filter}
            ORDER BY {device_col}
            """
        ).format(
            device_col=selected_columns[0],
                        tvm_location1_col=selected_columns[1],
                        tvm_location2_col=selected_columns[2],
                        description_col=selected_columns[3],
            schema=sql.Identifier(source['schema']),
            table=sql.Identifier(source['table']),
            device_filter=_build_device_filter_sql(
                sql.SQL("{device_col}::bigint").format(device_col=selected_columns[0])
            ),
        )

        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()

        filled_for_carrier = 0
        for row in rows:
            device_id, tvm_location1, tvm_location2, description = row
            normalized_device_id = int(device_id)
            if normalized_device_id not in missing_device_ids:
                continue

            picked_location = _pick_location_value(tvm_location1, tvm_location2, description)
            if not picked_location:
                continue

            location_by_device[normalized_device_id] = picked_location
            missing_device_ids.remove(normalized_device_id)
            filled_for_carrier += 1

        print(
            f"✓ Lokalizacje: {carrier_code} -> +{filled_for_carrier}, "
            f"braki po kroku: {len(missing_device_ids)}"
        )

    return location_by_device


def load_previous_snapshot(month_str):
    """
    Ładuje poprzedni snapshot dictionary (miesiąc wcześniej).
    month_str: 'YYYY-MM'
    Zwraca: dict lub None
    Konwertuje klucze stringowe z JSON z powrotem na integery.
    """
    # Oblicz poprzedni miesiąc
    year, month = map(int, month_str.split('-'))
    prev_date = datetime(year, month, 1) - relativedelta(months=1)
    prev_month_str = prev_date.strftime('%Y-%m')
    
    snapshot_file = SNAPSHOT_DIR / f'dictionary_{prev_month_str}.json'
    if snapshot_file.exists():
        with open(snapshot_file, 'r', encoding='utf-8') as f:
            loaded = json.load(f)
            # Konwertuj string keys na integer keys
            return {int(k): v for k, v in loaded.items()}
    return None


def save_snapshot(snapshot, month_str):
    """
    Zapisuje snapshot dictionary do pliku JSON.
    """
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_file = SNAPSHOT_DIR / f'dictionary_{month_str}.json'
    
    with open(snapshot_file, 'w', encoding='utf-8') as f:
        json.dump(snapshot, f, ensure_ascii=False, indent=2)
    
    print(f"✓ Snapshot zapisany: {snapshot_file}")


def _build_month_payload_cache_key(month_str, args, dictionary_comparison):
    """
    Buduje stabilny klucz cache payloadu miesiąca na podstawie wejścia raportu.
    """
    cache_signature = {
        'location_strategy_version': LOCATION_CACHE_STRATEGY_VERSION,
        'month_payload_cache_version': MONTH_PAYLOAD_CACHE_VERSION,
        'month_str': month_str,
        'obrot_source_mode': args.obrot_source_mode,
        'obrot_carriers': list(args.obrot_carriers or []),
        'obrot_transactions_table': args.obrot_transactions_table,
        'obrot_transactions_amount_col': args.obrot_transactions_amount_col,
        'obrot_transactions_date_col': args.obrot_transactions_date_col,
        'obrot_transactions_device_col': args.obrot_transactions_device_col,
        'obrot_transactions_payment_method_col': args.obrot_transactions_payment_method_col,
        'obrot_actioncodes': list(args.obrot_actioncodes or []),
        'obrot_mctype': args.obrot_mctype,
        'obrot_no_device_range_filter': bool(args.obrot_no_device_range_filter),
        'prowizja_file': str(args.prowizja_file),
        'lista_automatow_file': str(args.lista_automatow_file),
        'relokacje_file': str(args.relokacje_file),
        'serwis_file': str(args.serwis_file),
        'serwis_sheet_index': int(args.serwis_sheet_index),
        'it_card_file': str(args.it_card_file),
        'dictionary_devices': sorted(int(k) for k in dictionary_comparison.keys()),
    }
    signature_raw = json.dumps(cache_signature, sort_keys=True, ensure_ascii=True)
    return hashlib.sha1(signature_raw.encode('utf-8')).hexdigest()


def _month_payload_cache_path(month_str, args, dictionary_comparison):
    """
    Zwraca ścieżkę pliku cache payloadu miesiąca.
    """
    cache_key = _build_month_payload_cache_key(month_str, args, dictionary_comparison)
    return MONTH_PAYLOAD_CACHE_DIR / f'month_payload_{cache_key}.json'


def _encode_commission_rules_for_cache(commission_rules):
    """
    Konwertuje mapę reguł prowizji z kluczem tuple na format JSON.
    """
    def _normalize_rule_for_cache(rule):
        if not isinstance(rule, dict):
            return None

        valid_from = _as_date(rule.get('valid_from'))
        valid_to = _as_date(rule.get('valid_to'))
        return {
            'valid_from': valid_from.isoformat() if valid_from else None,
            'valid_to': valid_to.isoformat() if valid_to else None,
            'fixed_amount': float(rule.get('fixed_amount', 0.0) or 0.0),
            'percent': float(rule.get('percent', 0.0) or 0.0),
        }

    rows = []
    for (device_id, carrier_code), rules in sorted(
        (commission_rules or {}).items(),
        key=lambda item: (int(item[0][0]), str(item[0][1])),
    ):
        normalized_rules = []
        for rule in rules or []:
            normalized_rule = _normalize_rule_for_cache(rule)
            if normalized_rule is not None:
                normalized_rules.append(normalized_rule)

        rows.append({
            'device_id': int(device_id),
            'carrier': str(carrier_code),
            'rules': normalized_rules,
        })
    return rows


def _decode_commission_rules_from_cache(rows):
    """
    Odtwarza mapę reguł prowizji z formatu JSON do klucza tuple.
    """
    def _parse_rule_from_cache(rule):
        if not isinstance(rule, dict):
            return None

        raw_valid_from = rule.get('valid_from')
        valid_from = _as_date(raw_valid_from)
        if valid_from is None and isinstance(raw_valid_from, str):
            try:
                valid_from = datetime.strptime(raw_valid_from, '%Y-%m-%d').date()
            except ValueError:
                valid_from = None

        raw_valid_to = rule.get('valid_to')
        valid_to = _as_date(raw_valid_to)
        if valid_to is None and isinstance(raw_valid_to, str):
            try:
                valid_to = datetime.strptime(raw_valid_to, '%Y-%m-%d').date()
            except ValueError:
                valid_to = None

        return {
            'valid_from': valid_from,
            'valid_to': valid_to,
            'fixed_amount': float(rule.get('fixed_amount', 0.0) or 0.0),
            'percent': float(rule.get('percent', 0.0) or 0.0),
        }

    commission_rules = {}
    for row in rows or []:
        device_id = _parse_device_id(row.get('device_id'))
        carrier_code = _normalize_carrier_code(row.get('carrier'))
        rules = row.get('rules')
        if device_id is None or not carrier_code or not isinstance(rules, list):
            continue

        parsed_rules = []
        for rule in rules:
            parsed_rule = _parse_rule_from_cache(rule)
            if parsed_rule is not None:
                parsed_rules.append(parsed_rule)

        commission_rules[(int(device_id), carrier_code)] = parsed_rules
    return commission_rules


def _serialize_month_payload_for_cache(payload):
    """
    Przygotowuje payload miesiąca do zapisu JSON.
    """
    revenue_data = {
        str(int(device_id)): entry
        for device_id, entry in (payload.get('revenue_data') or {}).items()
    }
    location_by_device = {
        str(int(device_id)): location
        for device_id, location in (payload.get('location_by_device') or {}).items()
    }
    automat_type_by_device = {
        str(int(device_id)): automat_type
        for device_id, automat_type in (payload.get('automat_type_by_device') or {}).items()
    }
    service_cost_by_device = {
        str(int(device_id)): float(value or 0.0)
        for device_id, value in (payload.get('service_cost_by_device') or {}).items()
    }
    it_card_switch_dates = {}
    for device_id, switch_date in (payload.get('it_card_switch_dates') or {}).items():
        normalized_switch_date = _as_date(switch_date)
        it_card_switch_dates[str(int(device_id))] = (
            normalized_switch_date.isoformat() if normalized_switch_date else None
        )
    relocation_active_device_ids = sorted(
        int(device_id)
        for device_id in (payload.get('relocation_active_device_ids') or set())
    )
    warehouse_device_ids = sorted(
        int(device_id)
        for device_id in (payload.get('warehouse_device_ids') or set())
    )
    included_device_ids = sorted(
        int(device_id)
        for device_id in (payload.get('included_device_ids') or [])
    )
    return {
        'revenue_data': revenue_data,
        'commission_rules': _encode_commission_rules_for_cache(payload.get('commission_rules') or {}),
        'location_by_device': location_by_device,
        'automat_type_by_device': automat_type_by_device,
        'service_cost_by_device': service_cost_by_device,
        'it_card_switch_dates': it_card_switch_dates,
        'relocation_active_device_ids': relocation_active_device_ids,
        'warehouse_device_ids': warehouse_device_ids,
        'included_device_ids': included_device_ids,
    }


def _deserialize_month_payload_from_cache(raw_payload):
    """
    Odtwarza payload miesiąca z pliku cache do formatu roboczego.
    """
    revenue_data = {
        int(device_id): entry
        for device_id, entry in (raw_payload.get('revenue_data') or {}).items()
    }
    location_by_device = {
        int(device_id): location
        for device_id, location in (raw_payload.get('location_by_device') or {}).items()
    }
    automat_type_by_device = {
        int(device_id): automat_type
        for device_id, automat_type in (raw_payload.get('automat_type_by_device') or {}).items()
    }
    service_cost_by_device = {
        int(device_id): float(value or 0.0)
        for device_id, value in (raw_payload.get('service_cost_by_device') or {}).items()
    }

    it_card_switch_dates = {}
    for device_id, raw_date in (raw_payload.get('it_card_switch_dates') or {}).items():
        parsed_date = _as_date(raw_date)
        if parsed_date is None and isinstance(raw_date, str):
            try:
                parsed_date = datetime.strptime(raw_date, '%Y-%m-%d').date()
            except ValueError:
                parsed_date = None
        if parsed_date is not None:
            it_card_switch_dates[int(device_id)] = parsed_date

    relocation_active_device_ids = {
        int(device_id)
        for device_id in (raw_payload.get('relocation_active_device_ids') or [])
    }
    warehouse_device_ids = {
        int(device_id)
        for device_id in (raw_payload.get('warehouse_device_ids') or [])
    }
    included_device_ids = [
        int(device_id)
        for device_id in (raw_payload.get('included_device_ids') or [])
    ]

    commission_rules = _decode_commission_rules_from_cache(raw_payload.get('commission_rules') or [])
    return {
        'revenue_data': revenue_data,
        'commission_rules': commission_rules,
        'location_by_device': location_by_device,
        'automat_type_by_device': automat_type_by_device,
        'service_cost_by_device': service_cost_by_device,
        'it_card_switch_dates': it_card_switch_dates,
        'relocation_active_device_ids': relocation_active_device_ids,
        'warehouse_device_ids': warehouse_device_ids,
        'included_device_ids': included_device_ids,
    }


def load_month_payload_cache(month_str, args, dictionary_comparison):
    """
    Próbuje wczytać payload miesiąca z cache.
    Zwraca: payload dict lub None.
    """
    cache_file = _month_payload_cache_path(month_str, args, dictionary_comparison)
    if not cache_file.exists():
        return None

    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            raw_payload = json.load(f)
    except Exception as e:
        print(f"WARN CACHE READ FAILED ({month_str}): {cache_file} ({e})")
        return None

    try:
        payload = _deserialize_month_payload_from_cache(raw_payload)
    except Exception as e:
        print(f"WARN CACHE PARSE FAILED ({month_str}): {cache_file} ({e})")
        return None

    required_keys = {
        'revenue_data',
        'commission_rules',
        'location_by_device',
        'automat_type_by_device',
        'service_cost_by_device',
        'it_card_switch_dates',
        'relocation_active_device_ids',
        'warehouse_device_ids',
        'included_device_ids',
    }
    if not required_keys.issubset(set(payload.keys())):
        print(f"WARN CACHE INVALID STRUCTURE ({month_str}): {cache_file}")
        return None

    payload['automat_type_by_device'] = _load_automat_type_map_from_lista(
        month_str,
        getattr(args, 'lista_automatow_file', DEFAULT_LISTA_AUTOMATOW_FILE),
        payload.get('included_device_ids') or [],
    )

    print(f"CACHE HIT {month_str}: {cache_file}")
    return payload


def save_month_payload_cache(month_str, args, dictionary_comparison, payload):
    """
    Zapisuje payload miesiąca do cache.
    """
    MONTH_PAYLOAD_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_file = _month_payload_cache_path(month_str, args, dictionary_comparison)
    serializable_payload = _serialize_month_payload_for_cache(payload)

    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(serializable_payload, f, ensure_ascii=False, indent=2)

    print(f"CACHE SAVE {month_str}: {cache_file}")


def compare_snapshots(current, previous):
    """
    Porównuje bieżący snapshot z poprzednim.
    Zwraca: dict {device_id: {status: 'new'|'changed'|'unchanged', data: {...}}}
    """
    result = {}
    
    for device_id, data in current.items():
        if device_id not in previous:
            result[device_id] = {'status': 'new', 'data': data}
        elif data != previous[device_id]:
            result[device_id] = {'status': 'changed', 'data': data}
        else:
            result[device_id] = {'status': 'unchanged', 'data': data}
    
    # Urządzenia usunięte (były w poprzednim, nie ma w bieżącym)
    for device_id in previous:
        if device_id not in current:
            result[device_id] = {'status': 'deleted', 'data': previous[device_id]}
    
    return result


# === 2-OBRÓT: DANE Z MONEYSTATS ZA MIESIĄC ===

def get_month_range(month_str):
    """
    Zwraca zakres dat dla miesiąca raportowego.
    month_str: 'YYYY-MM'
    """
    year, month = map(int, month_str.split('-'))
    start_date = datetime(year, month, 1)
    end_date = start_date + relativedelta(months=1)
    return start_date, end_date


def get_month_range_closed(month_str):
    """
    Zwraca domknięty zakres dat dla miesiąca raportowego:
    od 1 dnia 00:00:00 do ostatniego dnia 23:59:59.
    """
    start_date, _ = get_month_range(month_str)
    year, month = map(int, month_str.split('-'))
    last_day = calendar.monthrange(year, month)[1]
    end_date = datetime(year, month, last_day, 23, 59, 59)
    return start_date, end_date


def _iter_months(start_month_str, end_month_str):
    """
    Generuje kolejne miesiące w formacie YYYY-MM (włącznie).
    """
    current = datetime.strptime(start_month_str, '%Y-%m')
    end = datetime.strptime(end_month_str, '%Y-%m')
    while current <= end:
        yield current.strftime('%Y-%m')
        current += relativedelta(months=1)


def _build_commission_report_months():
    """
    Zwraca docelowy zakres miesięcy zestawienia prowizji:
    od 2025-01 do poprzedniego miesiąca względem daty uruchomienia.
    """
    prev_month = (datetime.now().replace(day=1) - relativedelta(months=1)).strftime('%Y-%m')
    return list(_iter_months('2025-01', prev_month))


def _forecast_source_month_for_2026(month_str):
    """
    Dla miesięcy 2026-04..2026-12 zwraca analogiczny miesiąc z 2025.
    Dla pozostałych miesięcy zwraca None.
    """
    year, month = map(int, month_str.split('-'))
    if year == 2026 and 4 <= month <= 12:
        return f"2025-{month:02d}"
    return None


def _pick_first_matching_column(columns, candidates):
    """
    Zwraca pierwszą kolumnę pasującą do listy kandydatów.
    Porównanie jest case-insensitive.
    """
    lowered = {c.lower(): c for c in columns}
    for candidate in candidates:
        if candidate.lower() in lowered:
            return lowered[candidate.lower()]
    return None


def discover_commission_source(conn, schema=SCHEMA):
    """
    Próbuje automatycznie znaleźć tabelę i kolumny prowizji w schema ARP.
    Zwraca dict ze źródłem lub None.
    """
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT table_name, column_name
        FROM information_schema.columns
        WHERE table_schema = %s
        ORDER BY table_name, ordinal_position
        """,
        (schema,)
    )
    rows = cursor.fetchall()
    cursor.close()

    if not rows:
        return None

    tables = {}
    for table_name, column_name in rows:
        tables.setdefault(table_name, []).append(column_name)

    device_candidates = [
        'device_id', 'deviceid', 'ms_deviceid', 'id_automatu', 'nr_automatu', 'automat_id'
    ]
    amount_candidates = [
        'commission', 'prowizja', 'prowizja_zl', 'fee', 'interchange',
        'elavon_fee', 'amount', 'kwota', 'kwota_prowizji'
    ]
    date_candidates = [
        'create_date', 'createdate', 'ms_creadate', 'data', 'date',
        'transaction_date', 'settlement_date'
    ]
    provider_candidates = [
        'provider', 'operator', 'aggregator', 'source', 'payment_operator'
    ]
    table_name_hints = ['prowiz', 'commission', 'interchange', 'elavon', 'summary', 'settlement']

    best_source = None
    best_score = -1

    for table_name, columns in tables.items():
        device_col = _pick_first_matching_column(columns, device_candidates)
        amount_col = _pick_first_matching_column(columns, amount_candidates)
        date_col = _pick_first_matching_column(columns, date_candidates)
        provider_col = _pick_first_matching_column(columns, provider_candidates)

        if not all([device_col, amount_col, date_col]):
            continue

        score = 0
        lowered_table = table_name.lower()
        score += sum(2 for token in table_name_hints if token in lowered_table)
        score += 3  # komplet kluczowych kolumn
        if provider_col:
            score += 1

        if score > best_score:
            best_score = score
            best_source = {
                'schema': schema,
                'table': table_name,
                'device_col': device_col,
                'amount_col': amount_col,
                'date_col': date_col,
                'provider_col': provider_col,
                'provider_values': None
            }

    return best_source


def get_monthly_commission(conn, month_str, source_config=None, strict_month_window=True):
    """
    Pobiera prowizję per automat za dany miesiąc.
    Zwraca: dict {device_id: {'prowizja_zl': float, 'liczba_rekordow': int}}
    """
    start_date, end_date = get_month_range(month_str)

    if source_config is None:
        source_config = COMMISSION_SOURCE_CONFIG

    source = dict(source_config)
    has_explicit_mapping = all([
        source.get('table'),
        source.get('device_col'),
        source.get('amount_col'),
        source.get('date_col')
    ])

    if not has_explicit_mapping:
        discovered = discover_commission_source(conn, schema=source.get('schema', SCHEMA))
        if discovered is None:
            print("⚠ Nie znaleziono źródła prowizji - kolumna 3-PROWIZJA pozostanie pusta")
            return {}
        source.update(discovered)
        print(
            "✓ Auto-detekcja prowizji: "
            f"{source['schema']}.{source['table']} "
            f"(device={source['device_col']}, amount={source['amount_col']}, date={source['date_col']})"
        )

    provider_col = source.get('provider_col')
    provider_values = source.get('provider_values')
    date_to_col = source.get('date_to_col')

    params = [start_date, end_date]
    time_filter_sql = sql.SQL("{date_col} >= %s AND {date_col} < %s").format(
        date_col=sql.Identifier(source['date_col'])
    )

    # Tryb ścisły: rekord musi mieścić się w danym miesiącu (np. 2026-02-01 .. 2026-02-28).
    if strict_month_window and date_to_col:
        end_of_month = _end_of_month_date(month_str)
        params = [start_date, end_of_month]
        time_filter_sql = sql.SQL(
            "{date_from} >= %s AND {date_to} <= %s"
        ).format(
            date_from=sql.Identifier(source['date_col']),
            date_to=sql.Identifier(date_to_col)
        )
    elif date_to_col:
        params = [end_date, start_date]
        time_filter_sql = sql.SQL(
            "{date_from} < %s AND ({date_to} IS NULL OR {date_to} >= %s)"
        ).format(
            date_from=sql.Identifier(source['date_col']),
            date_to=sql.Identifier(date_to_col)
        )

    provider_filter_sql = sql.SQL('')
    if provider_col and provider_values:
        placeholders = sql.SQL(', ').join([sql.Placeholder()] * len(provider_values))
        provider_filter_sql = sql.SQL(" AND UPPER(COALESCE({provider_col}::text, '')) IN ({vals})").format(
            provider_col=sql.Identifier(provider_col),
            vals=placeholders
        )
        params.extend([v.upper() for v in provider_values])

    query = sql.SQL(
        """
        SELECT
            {device_col} AS device_id,
            SUM(COALESCE({amount_col}, 0)) AS prowizja_zl,
            COUNT(*) AS liczba_rekordow
        FROM {schema}.{table}
        WHERE {time_filter}
          AND {device_filter}
          {provider_filter}
        GROUP BY {device_col}
        ORDER BY device_id
        """
    ).format(
        schema=sql.Identifier(source.get('schema', SCHEMA)),
        table=sql.Identifier(source['table']),
        device_col=sql.Identifier(source['device_col']),
        amount_col=sql.Identifier(source['amount_col']),
        time_filter=time_filter_sql,
        device_filter=_build_device_filter_sql(sql.Identifier(source['device_col'])),
        provider_filter=provider_filter_sql
    )

    cursor = conn.cursor()
    try:
        cursor.execute(query, tuple(params))
        rows = cursor.fetchall()
    except psycopg2.Error as e:
        print(f"⚠ Błąd pobierania prowizji: {e}")
        rows = []
    finally:
        cursor.close()

    commission_data = {}
    for row in rows:
        device_id, prowizja, liczba_rekordow = row
        commission_data[int(device_id)] = {
            'prowizja_zl': float(prowizja or 0),
            'liczba_rekordow': int(liczba_rekordow)
        }

    return commission_data


def load_reference_provision_from_xls(xls_path):
    """
    Wczytuje plik referencyjny .xls (single sheet) i zwraca mapę prowizji per TVM_ID.
    Wymaga biblioteki pandas + xlrd. Jeśli brak bibliotek, zwraca None.
    """
    try:
        import pandas as pd
    except ImportError:
        print("⚠ Brak pandas - pomijam walidację pliku .xls")
        return None

    try:
        df = pd.read_excel(xls_path, sheet_name=0, engine='xlrd')
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku .xls: {e}")
        return None

    normalized = {str(col).strip().upper(): col for col in df.columns}
    tvm_col = None
    amount_col = None

    for candidate in ['TVM_ID', 'TVMID', 'NR AUTOMATU', 'NR_AUTOMATU', 'DEVICE_ID']:
        if candidate in normalized:
            tvm_col = normalized[candidate]
            break

    for candidate in ['PROVISION_AMOUNT', 'PROWIZJA', 'PROWIZJA_ZL', 'KWOTA', 'AMOUNT']:
        if candidate in normalized:
            amount_col = normalized[candidate]
            break

    if tvm_col is None or amount_col is None:
        print("⚠ Nie znaleziono kolumn TVM_ID / PROVISION_AMOUNT w pliku referencyjnym")
        return None

    ref = {}
    for _, row in df[[tvm_col, amount_col]].dropna(subset=[tvm_col]).iterrows():
        try:
            device_id = int(row[tvm_col])
        except Exception:
            continue
        amount = row[amount_col]
        amount = 0.0 if amount is None else float(amount)
        ref[device_id] = amount

    return ref


def compare_commission_with_reference(commission_data, reference_map, tolerance=0.01):
    """
    Porównuje prowizję z SQL vs. plik referencyjny.
    """
    if reference_map is None:
        return None

    sql_map = {int(k): float(v.get('prowizja_zl', 0.0)) for k, v in commission_data.items()}
    sql_ids = set(sql_map.keys())
    ref_ids = set(reference_map.keys())

    missing_in_sql = sorted(ref_ids - sql_ids)
    extra_in_sql = sorted(sql_ids - ref_ids)
    mismatches = []

    for device_id in sorted(sql_ids & ref_ids):
        diff = abs(sql_map[device_id] - reference_map[device_id])
        if diff > tolerance:
            mismatches.append({
                'device_id': device_id,
                'sql_amount': sql_map[device_id],
                'xls_amount': reference_map[device_id],
                'difference': diff,
            })

    return {
        'missing_in_sql': missing_in_sql,
        'extra_in_sql': extra_in_sql,
        'mismatches': mismatches,
    }


def validate_provision_sql_vs_xls(conn, month_str, commission_source, xls_path, strict_month_window=True):
    """
    Waliduje prowizję: SQL (per TVM_ID) vs plik referencyjny PROVISION .xls (single sheet).
    Zwraca tuple: (comparison_dict_or_none, exit_code)
    """
    commission_data = get_monthly_commission(
        conn,
        month_str,
        source_config=commission_source,
        strict_month_window=strict_month_window,
    )
    reference_map = load_reference_provision_from_xls(xls_path)
    comparison = compare_commission_with_reference(commission_data, reference_map)

    if comparison is None:
        print("❌ Nie udało się porównać SQL vs XLS (sprawdź plik i biblioteki: pandas, xlrd)")
        return None, 2

    missing_count = len(comparison['missing_in_sql'])
    extra_count = len(comparison['extra_in_sql'])
    mismatch_count = len(comparison['mismatches'])

    print(f"✓ SQL rekordów (TVM_ID): {len(commission_data)}")
    reference_count = len(reference_map) if reference_map is not None else 0
    print(f"✓ XLS rekordów (TVM_ID): {reference_count}")
    print(f"  Brakujące w SQL: {missing_count}")
    print(f"  Nadmiarowe w SQL: {extra_count}")
    print(f"  Różnice kwot: {mismatch_count}")

    if missing_count:
        print(f"  Przykład brakujących TVM_ID w SQL: {comparison['missing_in_sql'][:10]}")
    if extra_count:
        print(f"  Przykład nadmiarowych TVM_ID w SQL: {comparison['extra_in_sql'][:10]}")
    if mismatch_count:
        print("  Przykłady różnic kwotowych:")
        for item in comparison['mismatches'][:10]:
            print(
                f"    TVM_ID={item['device_id']} SQL={item['sql_amount']:.2f} "
                f"XLS={item['xls_amount']:.2f} DIFF={item['difference']:.2f}"
            )

    exit_code = 0 if (missing_count == 0 and extra_count == 0 and mismatch_count == 0) else 1
    return comparison, exit_code


def _has_nonempty_value(value):
    """
    Sprawdza, czy wartość jest niepusta z perspektywy odczytu z Excela.
    """
    if value is None:
        return False
    return bool(str(value).strip())


def _as_float(value, as_percent=False):
    """
    Konwertuje wartość z Excela na float.

    Obsługuje m.in. polski zapis dziesiętny, separatory tysięcy,
    symbole walut i zapis procentowy.
    Dla as_percent=True zwraca ułamek (np. 2% -> 0.02, 2 -> 0.02).
    """
    if value is None or isinstance(value, bool):
        return 0.0

    if isinstance(value, (int, float)):
        numeric = float(value)
        if not as_percent:
            return numeric
        if abs(numeric) <= 1.0:
            return numeric
        return numeric / 100.0

    text = str(value).strip()
    if not text:
        return 0.0

    text = (
        text
        .replace('\xa0', ' ')
        .replace('\u202f', ' ')
        .replace('\u2009', ' ')
        .replace('−', '-')
        .replace('–', '-')
        .replace('—', '-')
    )

    is_parentheses_negative = text.startswith('(') and text.endswith(')')
    if is_parentheses_negative:
        text = text[1:-1].strip()

    has_percent = '%' in text
    text = re.sub(r'(?i)\b(PLN|ZL|ZŁ|EUR|USD)\b', '', text)
    text = text.replace('%', '')
    text = text.replace(' ', '')
    text = text.replace("'", '')

    text = re.sub(r'[^0-9,\.\-]', '', text)
    if not text or text in ('-', '--'):
        return 0.0

    if text.startswith('-'):
        text = '-' + text[1:].replace('-', '')
    else:
        text = text.replace('-', '')

    if ',' in text and '.' in text:
        if text.rfind(',') > text.rfind('.'):
            text = text.replace('.', '')
            text = text.replace(',', '.')
        else:
            text = text.replace(',', '')
    elif ',' in text:
        if text.count(',') == 1:
            text = text.replace(',', '.')
        else:
            parts = text.split(',')
            text = ''.join(parts[:-1]) + '.' + parts[-1]
    elif text.count('.') > 1:
        parts = text.split('.')
        text = ''.join(parts[:-1]) + '.' + parts[-1]

    try:
        numeric = float(text)
    except (TypeError, ValueError):
        return 0.0

    if is_parentheses_negative and numeric > 0:
        numeric = -numeric

    if not as_percent:
        return numeric

    if has_percent:
        return numeric / 100.0
    if abs(numeric) <= 1.0:
        return numeric
    return numeric / 100.0


def _matches_month(value, month_str):
    """
    Sprawdza czy wartość daty należy do miesiąca raportowego YYYY-MM.
    """
    if value is None:
        return False
    if hasattr(value, 'year') and hasattr(value, 'month'):
        return f"{value.year:04d}-{value.month:02d}" == month_str
    return False


def _extract_monthly_value_in_section(ws, section_label, month_str, month_col=2, value_col=3):
    """
    Odczytuje miesięczną wartość z bloku sekcji oznaczonej etykietą w kolumnie B.
    """
    section_row = None
    normalized_label = str(section_label or '').strip().upper()
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(r, month_col).value
        if isinstance(cell_val, str) and str(cell_val).strip().upper() == normalized_label:
            section_row = r
            break

    if section_row is None:
        return 0.0

    value = 0.0
    for r in range(section_row + 1, ws.max_row + 1):
        marker = ws.cell(r, month_col).value
        if isinstance(marker, str) and marker.strip():
            break
        if _matches_month(marker, month_str):
            value += _as_float(ws.cell(r, value_col).value)
    return value


def _find_section_row(ws, section_label, label_col=2):
    """
    Zwraca numer wiersza sekcji opisanej etykietą w zadanej kolumnie.
    """
    normalized_label = str(section_label or '').strip().upper()
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(r, label_col).value
        if isinstance(cell_val, str) and str(cell_val).strip().upper() == normalized_label:
            return r
    return None


def _extract_monthly_value_from_columns(ws, month_str, month_col, value_col, start_row=1, end_row=None):
    """
    Sumuje wartości dla miesiąca w zadanych kolumnach (układ tabelaryczny bez sekcji).
    """
    if end_row is None:
        end_row = ws.max_row

    total = 0.0
    for r in range(max(1, int(start_row)), min(ws.max_row, int(end_row)) + 1):
        month_cell = ws.cell(r, month_col).value
        if _matches_month(month_cell, month_str):
            total += _as_float(ws.cell(r, value_col).value)
    return total


def _get_section_end_row(ws, section_row, section_label_col=2):
    """
    Szuka końca sekcji: pierwszy kolejny wiersz z nową etykietą sekcji.
    """
    if section_row is None:
        return ws.max_row

    for r in range(section_row + 1, ws.max_row + 1):
        marker = ws.cell(r, section_label_col).value
        if isinstance(marker, str) and marker.strip():
            return r - 1
    return ws.max_row


def _extract_monthly_value_from_header_table(ws, month_str, header_label, month_col, value_col, header_col=None):
    """
    Odczytuje wartość dla miesiąca z tabeli identyfikowanej nagłówkiem (np. Non TVMs/Project).
    Zakłada, że dane tabeli są w kolejnych wierszach pod nagłówkiem i kończą się przy pierwszej
    pustej komórce miesiąca po rozpoczęciu bloku dat.
    """
    header_row = _find_section_row(ws, header_label, label_col=header_col or value_col)
    if header_row is None:
        return 0.0

    total = 0.0
    seen_date_rows = False
    for r in range(header_row + 1, ws.max_row + 1):
        month_cell = ws.cell(r, month_col).value

        if _matches_month(month_cell, month_str):
            total += _as_float(ws.cell(r, value_col).value)
            seen_date_rows = True
            continue

        if hasattr(month_cell, 'year') and hasattr(month_cell, 'month'):
            seen_date_rows = True
            continue

        if month_cell is None and seen_date_rows:
            break

        if isinstance(month_cell, str) and month_cell.strip():
            break

    return total


def _extract_yearly_value_in_section(ws, section_label, year, month_col=2, value_col=3):
    """
    Sumuje wartości z sekcji oznaczonej `section_label` dla podanego roku.
    Szuka etykiety sekcji w kolumnie `month_col` i sumuje wiersze, których komórka daty
    ma rok równy `year`.
    """
    section_row = _find_section_row(ws, section_label, label_col=month_col)
    if section_row is None:
        return 0.0

    end_row = _get_section_end_row(ws, section_row, section_label_col=month_col)
    total = 0.0
    for r in range(section_row + 1, end_row + 1):
        marker = ws.cell(r, month_col).value
        if hasattr(marker, 'year') and hasattr(marker, 'month') and getattr(marker, 'year', None) == int(year):
            total += _as_float(ws.cell(r, value_col).value)
    return total


def _resolve_annual_section_cost_for_month(wb, month_str, section_label, month_col=2, value_col=3):
    """
    Dla zadanego miesiąca `month_str` zwraca roczną sumę kosztu z sekcji `section_label`.
    Najpierw próbuje zebrać sumę z roku bieżącego (na podstawie arkusza dla `month_str`).
    Jeśli suma bieżącego roku == 0, próbuje znaleźć dane z poprzedniego roku (przeglądając arkusze
    poprzedniego roku) i zwraca ich sumę.

    Zwraca krotkę: (year_total: float, year: int|None, mode: 'current'|'previous'|'none')
    """
    try:
        year = int(month_str.split('-')[0])
        month = int(month_str.split('-')[1])
    except Exception:
        return 0.0, None, 'none'

    # Bieżący rok: sprawdź arkusz dla month_str
    sheet_name = _resolve_sheet_name_by_month(wb, month_str)
    if sheet_name:
        ws = wb[sheet_name]
        current_total = _extract_yearly_value_in_section(ws, section_label, year, month_col=month_col, value_col=value_col)
        if current_total and float(current_total) != 0.0:
            return float(current_total), year, 'current'

    # Fallback: poprzedni rok - przeszukaj arkusze poprzedniego roku
    prev_year = year - 1
    for m in range(1, 13):
        candidate_month = f"{prev_year:04d}-{m:02d}"
        cand_sheet = _resolve_sheet_name_by_month(wb, candidate_month)
        if not cand_sheet:
            continue
        ws_prev = wb[cand_sheet]
        prev_total = _extract_yearly_value_in_section(ws_prev, section_label, prev_year, month_col=month_col, value_col=value_col)
        if prev_total and float(prev_total) != 0.0:
            return float(prev_total), prev_year, 'previous'

    return 0.0, None, 'none'


def _normalize_header_text(value):
    """
    Normalizuje tekst nagłówka do porównań (usuwa spacje, kropki i znaki specjalne).
    """
    if value is None:
        return ''
    return re.sub(r'[^A-Z0-9]+', '', str(value).strip().upper())


def _normalize_text_for_match(value):
    """
    Normalizuje tekst do porównań fuzzy (lokalizacje, nazwy) usuwając separatory.
    """
    if value is None:
        return ''
    return re.sub(r'[^A-Z0-9]+', '', str(value).strip().upper())


def _build_month_year_labels(month_str):
    """
    Zwraca etykiety miesiąca w formatach użytecznych do wyszukiwania nagłówków.
    """
    dt = datetime.strptime(month_str, '%Y-%m')
    return dt.strftime('%m/%Y'), dt.strftime('%m%Y')


def _parse_device_id(value):
    """
    Konwertuje numer automatu do int; zwraca None dla wartości niejednoznacznych.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None

    text = str(value).strip().replace(',', '.')
    if not text:
        return None
    if re.fullmatch(r'\d+(\.0+)?', text):
        return int(float(text))
    return None


def _as_date(value):
    """
    Konwertuje wartość do date; zwraca None gdy brak poprawnej daty.
    Rozszerzono o obsługę dat wklejonych w Excelu jako czysty tekst.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if hasattr(value, 'date'):
        return value.date()
        
    # --- NOWY FRAGMENT: Obsługa dat w postaci tekstowej (string) ---
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
            
        # Lista najczęstszych formatów dat wpisywanych w Excelu
        formats = (
            '%Y-%m-%d', '%Y-%m-%d %H:%M:%S',
            '%d.%m.%Y', '%d-%m-%Y', 
            '%d/%m/%Y', '%m/%d/%Y'
        )
        for fmt in formats:
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
                
    return None

def _month_date_bounds(month_str):
    """
    Zwraca zakres dat miesiąca raportowego jako date (start, end).
    """
    start_dt, end_dt = get_month_range_closed(month_str)
    return start_dt.date(), end_dt.date()


def _parse_relokacje_date_token(value):
    """
    Parsuje datę z pliku relokacji (np. 08.04.2024, 6/18/2025).
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    formats = ('%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y')
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _load_relocations_for_month(month_str, relokacje_file=DEFAULT_RELOKACJE_FILE, device_ids=None):
    """
    Wczytuje relokacje z pliku tekstowego i zwraca aktywne wpisy dla miesiąca.
    Zwraca tuple:
      - location_by_device: {device_id: location}
      - active_device_ids: set(device_id)
    """
    relokacje_path = Path(relokacje_file)
    if not relokacje_path.exists():
        print(f"WARN Brak pliku relokacji: {relokacje_path}")
        return {}, set()

    target_ids = {int(device_id) for device_id in (device_ids or [])}
    month_start, month_end = _month_date_bounds(month_str)
    parsed_rows = {}

    with open(relokacje_path, 'r', encoding='utf-8', errors='replace') as f:
        for idx, raw_line in enumerate(f, start=1):
            line = str(raw_line or '').strip()
            if not line:
                continue

            cleaned = re.sub(r'\([^)]*\)', '', line).strip()
            match = re.match(r'^\s*(\d+)\s*-\s*(.+?)\s*$', cleaned)
            if not match:
                continue

            device_id = int(match.group(1))
            if target_ids and device_id not in target_ids:
                continue

            rest = match.group(2).strip()
            marker_match = re.search(r'\b(od|do)\b', rest, flags=re.IGNORECASE)
            location = rest
            if marker_match:
                location = rest[:marker_match.start()]
            location = location.strip(' ,.;:-')
            if not location:
                print(f"WARN Relokacje: brak lokalizacji w linii {idx}: {line}")
                continue

            valid_from = None
            valid_to = None
            for token, token_date in re.findall(r'\b(od|do)\b\s*([0-9]{1,2}[./-][0-9]{1,2}[./-][0-9]{2,4})', rest, flags=re.IGNORECASE):
                parsed_date = _parse_relokacje_date_token(token_date)
                if parsed_date is None:
                    continue
                if token.lower() == 'od':
                    valid_from = parsed_date
                elif token.lower() == 'do':
                    valid_to = parsed_date

            if valid_from and valid_to and valid_from > valid_to:
                print(f"WARN Relokacje: pomijam wpis z niepoprawnym zakresem dat (linia {idx}): {line}")
                continue

            parsed_rows.setdefault(device_id, []).append({
                'location': location,
                'valid_from': valid_from,
                'valid_to': valid_to,
            })

    active_locations = {}
    active_device_ids = set()
    for device_id, rows in parsed_rows.items():
        matched_rows = []
        for row in rows:
            valid_from = row.get('valid_from') or date.min
            valid_to = row.get('valid_to') or date.max
            if valid_from <= month_end and valid_to >= month_start:
                matched_rows.append(row)

        if not matched_rows:
            continue

        matched_rows.sort(
            key=lambda row: (
                row.get('valid_from') or date.min,
                row.get('valid_to') or date.max,
            )
        )
        best_row = matched_rows[-1]
        active_locations[device_id] = best_row['location']
        active_device_ids.add(device_id)

    print(f"Wczytano aktywne relokacje dla {len(active_device_ids)} automatow ({month_str})")
    return active_locations, active_device_ids


def _month_str_to_tuple(month_str):
    """
    Konwertuje miesiac YYYY-MM na tuple (year, month) do porownan.
    """
    year, month = map(int, month_str.split('-'))
    return year, month


def _month_gte(month_str, threshold_month_str):
    """
    Sprawdza czy miesiac YYYY-MM jest >= od wskazanego progu.
    """
    return _month_str_to_tuple(month_str) >= _month_str_to_tuple(threshold_month_str)


def _get_additional_profit_value(month_str, device_id):
    """
    Zwraca dodatkowy zysk dla automatu i miesiaca wg recznych regul.
    """
    return float(ADDITIONAL_PROFIT_RULES.get((month_str, int(device_id)), 0.0) or 0.0)


def _summarize_commission_contract_starts(contract_rows, report_date):
    """
    Agreguje najwczesniejsza date startu umowy per (automat, przewoznik).

    To jest odpowiednik Pandas:
    df.groupby(['numer automatu', 'przewoznik'])['waznosc od'].min()

    Po tej agregacji na poziomie automatu sprawdzamy, czy istnieje przynajmniej
    jedna umowa, ktora zaczela sie nie pozniej niz data raportu.
    """
    report_date = _as_date(report_date)
    first_valid_from_by_key = {}

    for row in contract_rows or []:
        if not isinstance(row, dict):
            continue

        device_id = _parse_device_id(row.get('device_id'))
        carrier_code = _normalize_carrier_code(row.get('carrier'))
        valid_from = _as_date(row.get('valid_from'))
        if device_id is None or not carrier_code or valid_from is None:
            continue

        key = (int(device_id), carrier_code)
        current_min = first_valid_from_by_key.get(key)
        if current_min is None or valid_from < current_min:
            first_valid_from_by_key[key] = valid_from

    summary_by_device = {}
    for (device_id, carrier_code), first_valid_from in first_valid_from_by_key.items():
        device_entry = summary_by_device.setdefault(
            int(device_id),
            {
                'first_valid_from': None,
                'first_valid_from_by_carrier': {},
                'has_started_contract': False,
                'has_only_future_contracts': False,
            },
        )
        device_entry['first_valid_from_by_carrier'][carrier_code] = first_valid_from

        current_first = device_entry.get('first_valid_from')
        if current_first is None or first_valid_from < current_first:
            device_entry['first_valid_from'] = first_valid_from

    for device_entry in summary_by_device.values():
        carrier_dates = list(device_entry.get('first_valid_from_by_carrier', {}).values())
        has_started_contract = any(valid_from <= report_date for valid_from in carrier_dates)
        device_entry['has_started_contract'] = has_started_contract
        device_entry['has_only_future_contracts'] = bool(carrier_dates) and not has_started_contract

    return summary_by_device


def _sum_revenue_for_device(revenue_entry):
    """
    Zwraca laczny obrot automatu niezaleznie od formatu revenue_data.
    """
    if not isinstance(revenue_entry, dict):
        return 0.0

    by_carrier = revenue_entry.get('by_carrier')
    if isinstance(by_carrier, dict):
        return sum(
            float(carrier_entry.get('obrot_brutto_zl', 0.0) or 0.0)
            for carrier_entry in by_carrier.values()
            if isinstance(carrier_entry, dict)
        )

    return float(revenue_entry.get('obrot_brutto_zl', 0.0) or 0.0)


def _filter_device_ids_for_report_month(device_ids, revenue_data, contract_summary_by_device, month_str):
    """
    Filtruje automaty do P&L.

    Automat zostaje w raporcie tylko wtedy, gdy:
    - ma dodatni obrot w miesiacu,
    - i ma przynajmniej jedna umowe rozpoczeta nie pozniej niz koniec miesiaca.

    Walidacja jest wykonywana dla calej listy kandydatow, wiec obejmuje rowniez
    urzadzenia, ktore w poprzedniej logice trafialy do Uwagi='do sprawdzenia'.
    """
    _month_start, report_date = _month_date_bounds(month_str)
    included_ids = []
    excluded_details = {}

    for device_id in sorted({int(device_id) for device_id in (device_ids or [])}):
        revenue_total = _sum_revenue_for_device((revenue_data or {}).get(device_id))
        contract_summary = (contract_summary_by_device or {}).get(device_id) or {}
        has_started_contract = bool(contract_summary.get('has_started_contract'))
        first_valid_from_by_carrier = contract_summary.get('first_valid_from_by_carrier', {}) or {}

        reasons = []
        if revenue_total <= 0.0:
            reasons.append('zero_turnover')

        if not has_started_contract:
            if first_valid_from_by_carrier:
                reasons.append('future_contracts_only')
            else:
                reasons.append('missing_contracts')

        if reasons:
            excluded_details[device_id] = {
                'reasons': reasons,
                'revenue_total': float(revenue_total),
                'report_date': report_date,
                'first_valid_from_by_carrier': dict(first_valid_from_by_carrier),
            }
            continue

        included_ids.append(device_id)

    return included_ids, excluded_details


def _log_device_filter_summary(month_str, included_ids, excluded_details):
    """
    Loguje podsumowanie filtra automatow.
    """
    excluded_details = excluded_details or {}
    zero_turnover_count = sum(
        1 for details in excluded_details.values()
        if 'zero_turnover' in details.get('reasons', [])
    )
    future_only_count = sum(
        1 for details in excluded_details.values()
        if 'future_contracts_only' in details.get('reasons', [])
    )
    missing_contract_count = sum(
        1 for details in excluded_details.values()
        if 'missing_contracts' in details.get('reasons', [])
    )

    print(
        f"✓ Filtr automatow {month_str}: pozostawiono {len(included_ids)}, "
        f"usunieto {len(excluded_details)} "
        f"(zero obrotu={zero_turnover_count}, przyszle umowy={future_only_count}, "
        f"brak umowy={missing_contract_count})"
    )

    future_samples = []
    for device_id, details in excluded_details.items():
        if 'future_contracts_only' not in details.get('reasons', []):
            continue
        carrier_dates = details.get('first_valid_from_by_carrier', {}) or {}
        if not carrier_dates:
            continue
        sample = ', '.join(
            f"{carrier}={valid_from.isoformat()}"
            for carrier, valid_from in sorted(carrier_dates.items())
        )
        future_samples.append(f"{device_id}: {sample}")
        if len(future_samples) >= 5:
            break
    if future_samples:
        print(f"  Przykłady automatów z wyłącznie przyszłymi umowami: {future_samples}")


def _build_rop_eligible_device_ids(device_ids, commission_rules, month_str):
    """
    Wyznacza automaty liczace sie do podzialu kosztow ROP.

    Warunki:
    - automat jest na liscie analizowanych urzadzen,
    - automat nie jest na liscie wykluczen,
    - dla automatu istnieje przynajmniej jedna regula prowizji aktywna
      w miesiacu raportowym.
    """
    month_start, month_end = _month_date_bounds(month_str)
    candidate_ids = {int(device_id) for device_id in (device_ids or [])}
    eligible_ids = set()

    for (device_id, _carrier_code), rules in (commission_rules or {}).items():
        device_id_int = int(device_id)
        if device_id_int not in candidate_ids:
            continue
        if device_id_int in ROP_EXCLUDED_DEVICE_IDS:
            continue

        picked_rules = _pick_commission_rules_for_month(rules or [], month_start, month_end)
        if picked_rules:
            eligible_ids.add(device_id_int)

    return eligible_ids


def _cash_total_from_revenue_entry(revenue_entry):
    """
    Sumuje wartosc gotowki z wpisu revenue_data.
    """
    if not isinstance(revenue_entry, dict):
        return 0.0

    by_carrier = revenue_entry.get('by_carrier')
    if isinstance(by_carrier, dict):
        source_entries = by_carrier.values()
    else:
        source_entries = [revenue_entry]

    total = 0.0
    for entry in source_entries:
        if not isinstance(entry, dict):
            continue
        by_payment_method = entry.get('by_payment_method')
        if isinstance(by_payment_method, dict):
            total += float(by_payment_method.get('gotowka', 0.0) or 0.0)
        else:
            total += float(entry.get('gotowka', 0.0) or 0.0)
    return total


def _build_cash_total_by_device(revenue_data, device_ids, warehouse_device_ids=None):
    """
    Buduje mapę Gotowka Suma per automat; magazyn zawsze ma 0.
    """
    warehouse_ids = {int(device_id) for device_id in (warehouse_device_ids or set())}
    result = {}
    for device_id in sorted({int(device_id) for device_id in (device_ids or [])}):
        if device_id in warehouse_ids:
            result[device_id] = 0.0
            continue
        result[device_id] = _cash_total_from_revenue_entry((revenue_data or {}).get(device_id))
    return result


def _days_in_month(month_str):
    """
    Zwraca liczbę dni w miesiącu raportowym.
    """
    year, month = map(int, month_str.split('-'))
    return calendar.monthrange(year, month)[1]


def _it_card_active_ratio_for_month(switch_date, month_str):
    """
    Zwraca udział dni miesiąca, w których automat był już na IT CARD.
    switch_date oznacza pierwszy dzień działania IT CARD.
    """
    if switch_date is None:
        return 0.0

    month_start, month_end = _month_date_bounds(month_str)
    total_days = _days_in_month(month_str)

    if switch_date <= month_start:
        return 1.0
    if switch_date > month_end:
        return 0.0

    it_card_days = (month_end - switch_date).days + 1
    return max(0.0, min(1.0, float(it_card_days) / float(total_days)))


def _load_it_card_switch_dates(it_card_file=DEFAULT_IT_CARD_SWITCH_FILE):
    """
    Wczytuje daty przejścia ELAVON -> IT CARD per automat.
    Oczekiwane kolumny: numer automatu + data przejścia.
    """
    result = {}
    switch_path = Path(it_card_file)
    if not switch_path.exists():
        print(f"⚠ Brak pliku IT CARD: {switch_path}")
        return result

    try:
        from openpyxl import load_workbook
        wb = load_workbook(switch_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku IT CARD: {e}")
        return result

    device_tokens = {'NRTVM', 'NRAUTOMATU', 'NUMERAUTOMATU', 'TVM', 'AUTOMAT'}
    date_tokens = {'DATA', 'DATAPRZEJSCIA', 'PRZEJSCIE', 'ODKIEDY', 'ITCARDOD', 'ITCARD'}

    try:
        for ws in wb.worksheets:
            header_row = None
            device_col = None
            date_col = None

            scan_to = min(ws.max_row, 40)
            for r in range(1, scan_to + 1):
                local_device_col = None
                local_date_col = None
                for c in range(1, ws.max_column + 1):
                    token = _normalize_header_text(ws.cell(r, c).value)
                    if not token:
                        continue
                    if local_device_col is None and any(t in token for t in device_tokens):
                        local_device_col = c
                    if local_date_col is None and any(t in token for t in date_tokens):
                        local_date_col = c
                if local_device_col and local_date_col:
                    header_row = r
                    device_col = local_device_col
                    date_col = local_date_col
                    break

            if header_row is None:
                continue
            if device_col is None or date_col is None:
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                raw_device = ws.cell(r, device_col).value
                raw_date = ws.cell(r, date_col).value
                device_id = _parse_device_id(raw_device)
                switch_date = _as_date(raw_date)
                if device_id is None or switch_date is None:
                    continue
                result[int(device_id)] = switch_date
    finally:
        wb.close()

    print(f"Wczytano daty przejscia IT CARD dla {len(result)} automatow")
    return result


def _normalize_excel_commission_header(value):
    """
    Normalizuje nagłówek kolumny arkusza prowizji do porównań.
    """
    if value is None:
        return ''
    text = str(value).strip().upper()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r'[^A-Z0-9]+', '', text)


def _find_header_col(row_map, predicate):
    """
    Zwraca indeks kolumny dla pierwszego nagłówka spełniającego warunek.
    """
    for key, col in row_map.items():
        if predicate(key):
            return col
    return None


def _candidate_commission_sheet_names(wb, month_str):
    """
    Zwraca arkusze prowizji w kolejności od najbardziej pasujących do miesiąca.
    """
    year, month = month_str.split('-')
    month_tokens = {
        _normalize_header_text(month_str),
        _normalize_header_text(year),
        _normalize_header_text(f"{year}-{month}"),
        _normalize_header_text(f"{month}.{year}"),
        _normalize_header_text(f"{month}/{year}"),
        _normalize_header_text(f"{year}_{month}"),
        _normalize_header_text(f"{month}_{year}"),
        _normalize_header_text(f"{month}{year}"),
        _normalize_header_text(f"{year}{month}"),
    }

    ordered = []
    fallback = []

    for sheet_name in wb.sheetnames:
        normalized_sheet_name = _normalize_header_text(sheet_name)
        if not normalized_sheet_name:
            fallback.append(sheet_name)
            continue

        score = 0
        if normalized_sheet_name == _normalize_header_text(month_str):
            score += 50
        if normalized_sheet_name == _normalize_header_text(year):
            score += 40
        if any(token and token in normalized_sheet_name for token in month_tokens):
            score += 20
        if year in normalized_sheet_name:
            score += 10

        if score > 0:
            ordered.append((score, sheet_name))
        else:
            fallback.append(sheet_name)

    ordered.sort(key=lambda item: (-item[0], wb.sheetnames.index(item[1])))
    return [sheet_name for _, sheet_name in ordered] + fallback


def _pick_commission_rules_for_month(rules, month_start, month_end):
    """
    Wybiera najtrafniejsze rekordy prowizji dla miesiąca.

    Najpierw filtruje reguły nachodzące na miesiąc raportowy,
    a następnie bierze komplet rekordów z "najlepszego" okna:
    najpóźniejsze valid_from i (w tym zbiorze) najpóźniejsze valid_to.

    Dzięki temu jeżeli ryczałt i procent są zapisane w osobnych wierszach
    dla tego samego okresu, oba składniki są uwzględnione.
    """
    if not rules:
        return []

    matched = []
    for rule in rules:
        valid_from = rule.get('valid_from')
        valid_to = rule.get('valid_to')
        if valid_from is None and valid_to is None:
            continue
        if valid_from is None:
            valid_from = date.min
        if valid_to is None:
            valid_to = date.max
        if valid_from <= month_end and valid_to >= month_start:
            matched.append(rule)

    if not matched:
        return []

    best_valid_from = max((r.get('valid_from') or date.min) for r in matched)
    best_from_rules = [
        r for r in matched
        if (r.get('valid_from') or date.min) == best_valid_from
    ]
    best_valid_to = max((r.get('valid_to') or date.min) for r in best_from_rules)

    return [
        r for r in best_from_rules
        if (r.get('valid_to') or date.min) == best_valid_to
    ]


def _commission_amount_from_rule(rule, net_amount):
    """
    Wylicza prowizję dla jednej reguły i kwoty netto.
    """
    if not rule:
        return 0.0

    fixed_amount = float(rule.get('fixed_amount', 0.0) or 0.0)
    percent = float(rule.get('percent', 0.0) or 0.0)
    percent_amount = 0.0
    if percent > 0:
        percent_amount = float(net_amount or 0.0) * percent
    return max(fixed_amount, 0.0) + percent_amount


def _commission_amount_from_rules(rules, net_amount):
    """
    Sumuje prowizję ze wszystkich przekazanych reguł.
    """
    if not rules:
        return 0.0
    return sum(_commission_amount_from_rule(rule, net_amount) for rule in rules)

def build_monthly_commission_data_from_rules(revenue_data, commission_rules, month_str):
    """
    Buduje agregat prowizji per automat na podstawie reguł z pliku prowizji
    i obrotu per przewoźnik.

    Zwraca format zgodny z compare_commission_with_reference:
      {device_id: {'prowizja_zl': float, 'liczba_rekordow': int}}
    """
    month_start, month_end = _month_date_bounds(month_str)
    commission_data = {}
    missing_rules = 0
    missing_samples = []

    for device_id, rev_entry in (revenue_data or {}).items():
        if not isinstance(rev_entry, dict):
            continue
        by_carrier = rev_entry.get('by_carrier', {})
        if not isinstance(by_carrier, dict):
            continue

        device_id_int = int(device_id)
        total_device_commission = 0.0
        matched_records = 0

        for carrier_code, carrier_entry in by_carrier.items():
            normalized_carrier_code = _normalize_carrier_code(carrier_code)
            net_amount = 0.0
            by_payment_method = {}
            if isinstance(carrier_entry, dict):
                by_payment_method = carrier_entry.get('by_payment_method', {}) or {}
                cashless_gross_amount = (
                    float(by_payment_method.get('karta', 0.0) or 0.0)
                    + float(by_payment_method.get('blik', 0.0) or 0.0)
                )
                net_amount = cashless_gross_amount / 1.08

            rules_for_key = commission_rules.get((device_id_int, normalized_carrier_code), [])
            picked_rules = _pick_commission_rules_for_month(rules_for_key, month_start, month_end)
            if picked_rules:
                matched_records += 1
            else:
                missing_rules += 1
                if len(missing_samples) < 8:
                    missing_samples.append((device_id_int, normalized_carrier_code))

            calc_commission = _commission_amount_from_rules(picked_rules, net_amount)
            if normalized_carrier_code == 'KS' and (
                by_payment_method.get('karta', 0.0) > 0
                or by_payment_method.get('blik', 0.0) > 0
                or net_amount > 0
            ):
                calc_commission = 1400.0
                # Ukrywamy komunikat o braku w Excelu, bo stawka KS jest przypisana na sztywno:
                if not picked_rules:
                    matched_records += 1
                    missing_rules -= 1
                    if (device_id_int, normalized_carrier_code) in missing_samples:
                        missing_samples.remove((device_id_int, normalized_carrier_code))
            total_device_commission += calc_commission

        commission_data[device_id_int] = {
            'prowizja_zl': float(total_device_commission),
            'liczba_rekordow': int(matched_records),
        }

    if missing_rules:
        print(f"⚠ Brak dopasowanej reguły prowizji dla {missing_rules} par (automat, przewoźnik)")
        if missing_samples:
            print(f"  Przykłady braków: {missing_samples}")

    return commission_data


def load_commission_rules_and_locations_from_xlsx(month_str, commission_file=DEFAULT_PROWIZJE_FILE):
    """
    Wczytuje reguły prowizji i lokalizacje z pliku Prowizje_AB.xlsx.
    Zwraca:
      - commission_rules: {(device_id, carrier_code): [rule, ...]}
      - location_by_device: {device_id: location}
      - contract_summary_by_device: {device_id: {'first_valid_from_by_carrier': {...}, ...}}
    """
    commission_path = Path(commission_file)
    if not commission_path.exists():
        print(f"⚠ Brak pliku prowizji: {commission_path}")
        return {}, {}, {}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(commission_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku prowizji: {e}")
        return {}, {}, {}

    if not wb.sheetnames:
        return {}, {}, {}

    header_row = None
    header_map = {}

    ws = None
    for sheet_name in _candidate_commission_sheet_names(wb, month_str):
        candidate_ws = wb[sheet_name]
        scan_to = min(candidate_ws.max_row, 40)
        candidate_header_row = None
        candidate_header_map = {}

        for r in range(1, scan_to + 1):
            row_map = {}
            for c in range(1, candidate_ws.max_column + 1):
                token = _normalize_excel_commission_header(candidate_ws.cell(r, c).value)
                if not token:
                    continue
                row_map[token] = c

            if not row_map:
                continue

            carrier_col = _find_header_col(row_map, lambda k: 'PRZEWOZNIK' in k)
            device_col = _find_header_col(row_map, lambda k: 'NUMERAUTOMATU' in k)
            valid_from_col = _find_header_col(
                row_map,
                lambda k: (
                    'WAZNOSCOD' in k
                    or 'WANOOD' in k
                    or (k.startswith('WA') and k.endswith('OD'))
                ),
            )
            valid_to_col = _find_header_col(
                row_map,
                lambda k: (
                    'WAZNOSCDO' in k
                    or 'WANODO' in k
                    or (k.startswith('WA') and k.endswith('DO'))
                ),
            )
            fixed_col = _find_header_col(
                row_map,
                lambda k: (
                    'RYCZALT' in k
                    or 'RYCZAT' in k
                    or 'STALA' in k
                    or 'KWOTOWA' in k
                ),
            )
            percent_col = _find_header_col(row_map, lambda k: 'PROCENTOWA' in k or 'PROCENT' in k)
            location_col = _find_header_col(row_map, lambda k: 'TVMLOCATION1' in k)

            if carrier_col and device_col and valid_from_col and valid_to_col and (fixed_col or percent_col):
                candidate_header_row = r
                candidate_header_map = {
                    'carrier_col': carrier_col,
                    'device_col': device_col,
                    'valid_from_col': valid_from_col,
                    'valid_to_col': valid_to_col,
                    'fixed_col': fixed_col,
                    'percent_col': percent_col,
                    'location_col': location_col,
                }
                break

        if candidate_header_row is not None:
            ws = candidate_ws
            header_row = candidate_header_row
            header_map = candidate_header_map
            break

    if header_row is None or ws is None:
        print("⚠ Nie znaleziono wymaganych kolumn w pliku prowizji")
        return {}, {}, {}

    commission_rules = {}
    location_by_device = {}
    contract_start_rows = []
    month_start, month_end = _month_date_bounds(month_str)
    invalid_date_ranges = 0
    skipped_missing_required = 0
    skipped_zero_after_parse = 0
    fixed_only_rules = 0
    percent_only_rules = 0
    mixed_rules = 0
    parse_to_zero_fixed = 0
    parse_to_zero_percent = 0
    parse_to_zero_samples = []

    for r in range(header_row + 1, ws.max_row + 1):
        raw_carrier = ws.cell(r, header_map['carrier_col']).value
        raw_device = ws.cell(r, header_map['device_col']).value
        raw_from = ws.cell(r, header_map['valid_from_col']).value
        raw_to = ws.cell(r, header_map['valid_to_col']).value

        if raw_carrier is None and raw_device is None and raw_from is None and raw_to is None:
            continue

        carrier_code = _normalize_carrier_code(raw_carrier)
        device_id = _parse_device_id(raw_device)
        valid_from = _as_date(raw_from)
        valid_to = _as_date(raw_to)

        if carrier_code and device_id is not None and valid_from is not None:
            contract_start_rows.append({
                'device_id': int(device_id),
                'carrier': carrier_code,
                'valid_from': valid_from,
            })

        if not carrier_code or device_id is None or valid_from is None or valid_to is None:
            skipped_missing_required += 1
            continue

        if valid_from > valid_to:
            invalid_date_ranges += 1
            continue

        fixed_amount = 0.0
        percent = 0.0
        raw_fixed = None
        raw_percent = None
        if header_map.get('fixed_col'):
            raw_fixed = ws.cell(r, header_map['fixed_col']).value
            fixed_amount = _as_float(raw_fixed)
        if header_map.get('percent_col'):
            raw_percent = ws.cell(r, header_map['percent_col']).value
            percent = _as_float(raw_percent, as_percent=True)

        if _has_nonempty_value(raw_fixed) and fixed_amount == 0.0 and re.search(r'[1-9]', str(raw_fixed)):
            parse_to_zero_fixed += 1
            if len(parse_to_zero_samples) < 8:
                parse_to_zero_samples.append((r, 'fixed', str(raw_fixed)))

        if _has_nonempty_value(raw_percent) and percent == 0.0 and re.search(r'[1-9]', str(raw_percent)):
            parse_to_zero_percent += 1
            if len(parse_to_zero_samples) < 8:
                parse_to_zero_samples.append((r, 'percent', str(raw_percent)))

        if fixed_amount <= 0.0 and percent <= 0.0:
            skipped_zero_after_parse += 1
            continue

        if fixed_amount > 0.0 and percent > 0.0:
            mixed_rules += 1
        elif fixed_amount > 0.0:
            fixed_only_rules += 1
        else:
            percent_only_rules += 1

        key = (int(device_id), carrier_code)
        commission_rules.setdefault(key, []).append({
            'valid_from': valid_from,
            'valid_to': valid_to,
            'fixed_amount': fixed_amount,
            'percent': percent,
        })

        raw_location = ws.cell(r, header_map['location_col']).value if header_map.get('location_col') else None
        normalized_location = _normalize_tvm_location_text(raw_location)
        if (
            normalized_location
            and int(device_id) not in location_by_device
            and valid_from <= month_end
            and valid_to >= month_start
        ):
            location_by_device[int(device_id)] = normalized_location

    if invalid_date_ranges:
        print(
            f"⚠ Pominięto {invalid_date_ranges} wierszy prowizji z nieprawidłowym zakresem dat (od > do)"
        )
    if skipped_missing_required:
        print(
            f"⚠ Pominięto {skipped_missing_required} wierszy prowizji z brakiem wymaganych danych "
            "(przewoźnik/automat/daty)"
        )
    if skipped_zero_after_parse:
        print(
            f"⚠ Pominięto {skipped_zero_after_parse} wierszy prowizji, bo ryczałt i procent wyszły jako 0"
        )

    parse_to_zero_total = parse_to_zero_fixed + parse_to_zero_percent
    if parse_to_zero_total:
        print(
            "⚠ Podejrzane parse-to-zero w prowizjach: "
            f"fixed={parse_to_zero_fixed}, percent={parse_to_zero_percent}"
        )
        if parse_to_zero_samples:
            print(f"  Przykłady parse-to-zero: {parse_to_zero_samples}")

    print(
        "✓ Typy reguł prowizji: "
        f"ryczałt={fixed_only_rules}, "
        f"procent={percent_only_rules}, "
        f"mieszane={mixed_rules}"
    )
    loaded_carriers = sorted({carrier for _, carrier in commission_rules.keys()})
    if loaded_carriers:
        print(f"✓ Przewoźnicy w regułach prowizji: {loaded_carriers}")

    contract_summary_by_device = _summarize_commission_contract_starts(contract_start_rows, month_end)
    print(f"✓ Agregacja dat startu umów: {len(contract_summary_by_device)} automatów")

    return commission_rules, location_by_device, contract_summary_by_device


def _roman_month_label(month_str):
    """
    Zwraca dokładną etykietę miesiąca rzymskiego I..XII.
    """
    month_no = int(month_str.split('-')[1])
    labels = {
        1: 'I',
        2: 'II',
        3: 'III',
        4: 'IV',
        5: 'V',
        6: 'VI',
        7: 'VII',
        8: 'VIII',
        9: 'IX',
        10: 'X',
        11: 'XI',
        12: 'XII',
    }
    return labels.get(month_no)


def _resolve_amortyzacja_file_for_month(month_str):
    """
    Wybiera plik amortyzacji na podstawie roku miesiąca raportowego.
    """
    year = int(month_str.split('-')[0])
    if year == 2025:
        return DEFAULT_AMORTYZACJA_2025_FILE
    if year == 2026:
        return DEFAULT_AMORTYZACJA_2026_FILE
    return None


def _normalize_amortyzacja_header_text(value):
    """
    Normalizuje tekst nagłówka amortyzacji (bez diakrytyków i znaków specjalnych).
    """
    if value is None:
        return ''
    text = str(value).strip().upper()
    text = unicodedata.normalize('NFKD', text)
    text = ''.join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r'[^A-Z0-9]+', '', text)


def _is_amortyzacja_header_cell(value):
    """
    Sprawdza czy komórka wygląda jak nagłówek bloku amortyzacji miesięcznej.
    """
    normalized = _normalize_amortyzacja_header_text(value)
    if not normalized:
        return False
    has_prefix = 'WARTOSCMIESIECZNEGO' in normalized
    has_umorzenie = ('UMORZEMIA' in normalized) or ('UMORZENIA' in normalized)
    has_amort = 'AMORTYZACJI' in normalized
    return has_prefix and has_umorzenie and has_amort


def _find_nr_automatu_header(ws):
    """
    Zwraca pozycję nagłówka kolumny Nr automatu (row, col).
    """
    scan_to = min(ws.max_row, 100)
    for r in range(1, scan_to + 1):
        for c in range(1, ws.max_column + 1):
            normalized = _normalize_header_text(ws.cell(r, c).value)
            if normalized == 'NRAUTOMATU':
                return r, c
    return None, None


def _find_work_header_col(ws):
    """
    Zwraca indeks kolumny WORK (lub UWAGA) w obszarze nagłówka amortyzacji.
    """
    tokens = {
        'WORK',
        'UWAGA',
        'UWAGI',
    }
    scan_to = min(ws.max_row, 120)
    for r in range(1, scan_to + 1):
        for c in range(1, ws.max_column + 1):
            normalized = _normalize_amortyzacja_header_text(ws.cell(r, c).value)
            if normalized in tokens:
                return c
    return None


def _find_amortyzacja_month_col(ws, roman_month_label):
    """
    Zwraca (month_header_row, month_col) dla kolumny miesiąca rzymskiego.
    """
    merged_cells = getattr(ws, 'merged_cells', None)
    merged_ranges = getattr(merged_cells, 'ranges', None)
    if merged_ranges:
        for merged_range in merged_ranges:
            top_row = merged_range.min_row
            top_col = merged_range.min_col
            if not _is_amortyzacja_header_cell(ws.cell(top_row, top_col).value):
                continue

            month_header_row = top_row + 1
            if month_header_row > ws.max_row:
                continue

            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell_value = ws.cell(month_header_row, col).value
                if str(cell_value or '').strip().upper() == roman_month_label:
                    return month_header_row, col

    scan_to = min(ws.max_row, 120)
    for r in range(1, scan_to + 1):
        for c in range(1, ws.max_column + 1):
            if not _is_amortyzacja_header_cell(ws.cell(r, c).value):
                continue
            month_header_row = r + 1
            if month_header_row > ws.max_row:
                continue
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(month_header_row, col).value
                if str(cell_value or '').strip().upper() == roman_month_label:
                    return month_header_row, col

    return None, None


def _load_amortyzacja_by_device_for_month(month_str):
    """
    Wczytuje amortyzację per automat z plików amortyzacja_25/amortyzacja_26.
    """
    result = {}
    warehouse_device_ids = set()
    conflicts = set()
    stats = {
        'rows_read': 0,
        'rows_with_device': 0,
        'rows_without_device': 0,
        'conflicts': 0,
    }

    roman_month_label = _roman_month_label(month_str)
    if not roman_month_label:
        return result, warehouse_device_ids

    amort_path = _resolve_amortyzacja_file_for_month(month_str)
    if amort_path is None:
        print(f"WARN Brak mapowania pliku amortyzacji dla miesiaca: {month_str}")
        return result, warehouse_device_ids
    if not amort_path.exists():
        print(f"WARN Brak pliku amortyzacji: {amort_path}")
        return result, warehouse_device_ids

    try:
        from openpyxl import load_workbook
        # Potrzebujemy merged_cells do znalezienia naglowka miesiecy rzymskich.
        wb = load_workbook(amort_path, data_only=True, read_only=False)
    except Exception as e:
        print(f"WARN Nie udalo sie odczytac pliku amortyzacji: {e}")
        return result, warehouse_device_ids

    try:
        target_sheet_groups = [
            ('bb8', ('bb8',)),
            ('ST_40FA2000', ('ST_40FA2000', 'ST40_FAA2000')),
        ]
        resolved_sheet_names = []
        available_by_lower = {str(sheet_name).strip().lower(): sheet_name for sheet_name in wb.sheetnames}

        for canonical_name, aliases in target_sheet_groups:
            resolved_name = None
            for alias in aliases:
                resolved_name = available_by_lower.get(alias.lower())
                if resolved_name is not None:
                    break

            if resolved_name is None:
                print(f"WARN Brak arkusza amortyzacji: {canonical_name} w {amort_path.name}")
                continue
            resolved_sheet_names.append(resolved_name)

        for sheet_name in resolved_sheet_names:

            ws = wb[sheet_name]
            nr_header_row, nr_col = _find_nr_automatu_header(ws)
            month_header_row, month_col = _find_amortyzacja_month_col(ws, roman_month_label)
            work_col = _find_work_header_col(ws)

            if nr_col is None or month_col is None:
                print(
                    "WARN Nie znaleziono wymaganych naglowkow amortyzacji "
                    f"(arkusz {sheet_name}, plik {amort_path.name})"
                )
                continue

            data_start_row = max(int(nr_header_row or 1), int(month_header_row or 1)) + 1

            empty_streak = 0
            for r in range(data_start_row, ws.max_row + 1):
                raw_device = ws.cell(r, nr_col).value
                raw_amount = ws.cell(r, month_col).value

                if raw_device is None and raw_amount is None:
                    empty_streak += 1
                    if empty_streak >= 200:
                        break
                    continue

                empty_streak = 0
                stats['rows_read'] += 1

                device_id = _parse_device_id(raw_device)
                if device_id is None:
                    stats['rows_without_device'] += 1
                    continue

                stats['rows_with_device'] += 1
                if device_id in conflicts:
                    continue

                if device_id in result:
                    conflicts.add(device_id)
                    stats['conflicts'] += 1
                    del result[device_id]
                    continue

                result[device_id] = _as_float(raw_amount)

                warehouse_detected = False
                if work_col is not None:
                    raw_work = ws.cell(r, work_col).value
                    if 'MAGAZYN' in str(raw_work or '').strip().upper():
                        warehouse_detected = True

                if not warehouse_detected:
                    for c in range(1, ws.max_column + 1):
                        raw_value = ws.cell(r, c).value
                        if isinstance(raw_value, str) and 'MAGAZYN' in raw_value.strip().upper():
                            warehouse_detected = True
                            break

                if warehouse_detected:
                    warehouse_device_ids.add(device_id)
    finally:
        wb.close()

    if stats['conflicts']:
        sample = sorted(conflicts)[:10]
        print(
            "WARN Konflikt amortyzacji (ten sam automat w wielu arkuszach): "
            f"{stats['conflicts']}; pominiete automaty (przyklad): {sample}"
        )

    print(
        "Wczytano amortyzacje per automat: "
        f"{len(result)} (plik {amort_path.name}, miesiac {roman_month_label})"
    )
    return result, warehouse_device_ids


def _resolve_sheet_name_by_month(wb, month_str, preferred_sheet=None):
    """
    Wybiera arkusz najlepiej pasujący do miesiąca.
    """
    if not wb.sheetnames:
        return None

    if preferred_sheet and preferred_sheet in wb.sheetnames:
        return preferred_sheet

    year = month_str.split('-')[0]
    if year in wb.sheetnames:
        return year

    normalized_year = _normalize_header_text(year)
    best_sheet = None
    best_score = 0

    for sheet_name in wb.sheetnames:
        normalized_sheet = _normalize_header_text(sheet_name)
        if not normalized_sheet:
            continue

        score = 0
        if normalized_sheet == normalized_year:
            score += 50
        if normalized_year and normalized_year in normalized_sheet:
            score += 20

        if score > best_score:
            best_score = score
            best_sheet = sheet_name

    if best_sheet is not None and best_score > 0:
        return best_sheet

    return wb.sheetnames[0]




def _load_rent_costs_by_device(month_str, rent_file=DEFAULT_RENT_FILE, rent_sheet=None):
    """
    Wczytuje czynsz i prąd per automat z pliku najmu.
    Oczekuje kolumn: Nr.automatu, Czynsz MM/YYYY, Energia MM/YYYY.
    """
    if rent_file is None:
        return {}

    rent_path = Path(rent_file)
    if not rent_path.exists():
        print(f"⚠ Brak pliku najmu: {rent_path}")
        return {}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(rent_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku najmu: {e}")
        return {}

    sheet_name = _resolve_sheet_name_by_month(wb, month_str, preferred_sheet=rent_sheet)
    if sheet_name is None:
        print(f"⚠ Nie znaleziono arkusza najmu dla miesiąca {month_str} w pliku: {rent_path.name}")
        return {}

    try:
        ws = wb[sheet_name]
        month_label_slash, month_label_compact = _build_month_year_labels(month_str)

        header_row = None
        device_col = None
        czynsz_col = None
        prad_col = None

        czynsz_targets = {
            _normalize_header_text(f"Czynsz {month_label_slash}"),
            _normalize_header_text(f"Czynsz{month_label_slash}"),
            _normalize_header_text(f"Czynsz {month_label_compact}"),
            _normalize_header_text(f"Czynsz{month_label_compact}"),
        }
        prad_targets = {
            _normalize_header_text(f"Energia {month_label_slash}"),
            _normalize_header_text(f"Energia{month_label_slash}"),
            _normalize_header_text(f"Energia {month_label_compact}"),
            _normalize_header_text(f"Energia{month_label_compact}"),
            _normalize_header_text(f"Prad {month_label_slash}"),
            _normalize_header_text(f"Prad{month_label_slash}"),
            _normalize_header_text(f"Prad {month_label_compact}"),
            _normalize_header_text(f"Prad{month_label_compact}"),
        }

        scan_to = min(ws.max_row, 60)
        for r, row_values in enumerate(
            ws.iter_rows(min_row=1, max_row=scan_to, values_only=True),
            start=1,
        ):
            row_device_col = None
            row_czynsz_col = None
            row_prad_col = None

            for c, raw_header in enumerate(row_values, start=1):
                normalized_header = _normalize_header_text(raw_header)
                if not normalized_header:
                    continue

                if normalized_header == 'NRAUTOMATU':
                    row_device_col = c
                if normalized_header in czynsz_targets:
                    row_czynsz_col = c
                if normalized_header in prad_targets:
                    row_prad_col = c

            if row_device_col and row_czynsz_col and row_prad_col:
                header_row = r
                device_col = row_device_col
                czynsz_col = row_czynsz_col
                prad_col = row_prad_col
                break

        if header_row is None:
            print(
                "⚠ Nie znaleziono wymaganych kolumn najmu: "
                f"Nr.automatu, Czynsz {month_label_slash}, Energia {month_label_slash}"
            )
            return {}

        if device_col is None or czynsz_col is None or prad_col is None:
            print("⚠ Niekompletna konfiguracja kolumn najmu")
            return {}

        min_col = min(device_col, czynsz_col, prad_col)
        max_col = max(device_col, czynsz_col, prad_col)
        device_idx = device_col - min_col
        czynsz_idx = czynsz_col - min_col
        prad_idx = prad_col - min_col

        rent_map = {}
        skipped_rows = 0
        empty_streak = 0

        for row_values in ws.iter_rows(
            min_row=header_row + 1,
            max_row=ws.max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            raw_device = row_values[device_idx]
            raw_czynsz = row_values[czynsz_idx]
            raw_prad = row_values[prad_idx]

            if raw_device is None and raw_czynsz is None and raw_prad is None:
                empty_streak += 1
                if empty_streak >= 200:
                    break
                continue

            empty_streak = 0

            device_id = _parse_device_id(raw_device)
            if device_id is None:
                skipped_rows += 1
                continue

            rent_map[device_id] = {
                'czynsz': _as_float(raw_czynsz),
                'prad': _as_float(raw_prad),
            }

        if skipped_rows:
            print(f"⚠ Pominięto {skipped_rows} wierszy najmu z nieprawidłowym Nr.automatu")

        return rent_map
    finally:
        wb.close()


def _load_elavon_total_from_costs_file(month_str, elavon_costs_file):
    """
    Wczytuje łączną kwotę ELAVON dla miesiąca z pliku kosztów.
    Plik: P&L_koszty_ROP.xlsx.
    Zwraca łączną kwotę ELAVON (float).
    """
    if elavon_costs_file is None:
        return 0.0

    elavon_path = Path(elavon_costs_file)
    if not elavon_path.exists():
        print(f"⚠ Brak pliku ELAVON: {elavon_path}")
        return 0.0

    try:
        year, month = map(int, month_str.split('-'))
    except (ValueError, IndexError):
        print(f"⚠ Nieprawidłowy format miesiąca dla ELAVON: {month_str}")
        return 0.0

    try:
        from openpyxl import load_workbook
        wb = load_workbook(elavon_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku ELAVON: {e}")
        return 0.0

    elavon_amount = 0.0

    def _matches_month_cell(cell_value):
        return _matches_month(cell_value, month_str)

    try:
        for ws in wb.worksheets:
            section_row = None
            for r in range(1, min(ws.max_row, 160) + 1):
                cell_value = ws.cell(r, 2).value
                if isinstance(cell_value, str) and cell_value.strip().upper() == 'ELAVON':
                    section_row = r
                    break

            if section_row is None:
                continue

            section_end_row = _get_section_end_row(ws, section_row, section_label_col=2)
            for r in range(section_row + 1, section_end_row + 1):
                if not _matches_month_cell(ws.cell(r, 2).value):
                    continue

                value = ws.cell(r, 3).value
                if isinstance(value, (int, float)) and value > 0:
                    elavon_amount = float(value)
                    break

            if elavon_amount > 0:
                break
    finally:
        wb.close()

    if elavon_amount <= 0:
        print(f"ELAVON dla {month_str}: kwota = 0 lub nie znaleziona")
        return 0.0

    print(f"✓ ELAVON {month_str}: kwota łączna={elavon_amount:.2f}")
    return float(elavon_amount)


def _load_service_overrides_from_xlsx(service_file=None, sheet_index=DEFAULT_SERWIS_SHEET_INDEX, month_str=None):
    """
    Wczytuje koszt serwisu per automat z pliku serwis_2026.xlsx.
    Obowiązuje od 2026-01.
    Oczekiwane kolumny: NR TVM, SERWIS KOSZT (lokalizacja opcjonalna).
    """
    location_by_device = {}
    service_cost_by_device = {}
    
    # Serwis obowiązuje od 2026-01
    if month_str:
        try:
            year, month = map(int, month_str.split('-'))
            if (year, month) < (2026, 1):
                return location_by_device, service_cost_by_device
            if service_file is None:
                service_file = DEFAULT_SERWIS_FILE
        except (ValueError, IndexError):
            pass
    
    if service_file is None:
        return location_by_device, service_cost_by_device

    service_path = Path(service_file)
    if not service_path.exists():
        print(f"⚠ Brak pliku serwisu: {service_path}")
        return location_by_device, service_cost_by_device

    try:
        from openpyxl import load_workbook
        wb = load_workbook(service_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku serwisu: {e}")
        return location_by_device, service_cost_by_device

    try:
        worksheets = [wb[sheet_name] for sheet_name in wb.sheetnames]
        skipped_rows = 0

        for ws in worksheets:
            header_row = None
            col_map = {}
            scan_to = min(ws.max_row, 60)
            for r in range(1, scan_to + 1):
                row_map = {}
                for c in range(1, ws.max_column + 1):
                    token = _normalize_header_text(ws.cell(r, c).value)
                    if token:
                        row_map[token] = c

                if not row_map:
                    continue

                nr_col = row_map.get('NRTVM') or row_map.get('NRAUTOMATU')
                loc_col = row_map.get('LOKALIZACJA')
                serwis_col = row_map.get('SERWISKOSZT')
                if nr_col and serwis_col:
                    header_row = r
                    col_map = {
                        'nr_col': nr_col,
                        'loc_col': loc_col,
                        'serwis_col': serwis_col,
                    }
                    break

            if header_row is None:
                continue

            for r in range(header_row + 1, ws.max_row + 1):
                raw_device = ws.cell(r, col_map['nr_col']).value
                raw_serwis = ws.cell(r, col_map['serwis_col']).value
                if raw_device is None and raw_serwis is None:
                    continue

                device_id = _parse_device_id(raw_device)
                if device_id is None:
                    if _has_nonempty_value(raw_device):
                        skipped_rows += 1
                    continue

                if col_map.get('loc_col'):
                    raw_location = ws.cell(r, col_map['loc_col']).value
                    normalized_location = _normalize_tvm_location_text(raw_location)
                    if normalized_location:
                        location_by_device[device_id] = normalized_location

                if _has_nonempty_value(raw_serwis):
                    service_cost_by_device[device_id] = _as_float(raw_serwis)

        if skipped_rows:
            print(f"⚠ Pominięto {skipped_rows} wierszy serwisu z nieprawidłowym NR TVM")

        print(
            "✓ Wczytano serwis 2026: "
            f"lokalizacje={len(location_by_device)}, "
            f"koszt_serwisu={len(service_cost_by_device)}"
        )
        return location_by_device, service_cost_by_device
    finally:
        wb.close()


def _get_monthly_costs_per_device(
    month_str,
    device_ids,
    costs_file=DEFAULT_COSTS_FILE,
    rent_file=DEFAULT_RENT_FILE,
    rent_sheet=None,
    service_cost_by_device=None,
    it_card_switch_dates=None,
    cash_total_by_device=None,
):
    """
    Wczytuje koszty z pliku Excel i rozdziela je na automaty.
    Dla kosztów globalnych stosuje równy podział na wszystkie automaty AB.
    """
    normalized_ids = sorted({int(device_id) for device_id in device_ids})
    if not normalized_ids:
        return {}

    default_entry = {
        **{key: 0.0 for key in TVM_COST_KEYS},
        **{key: 0.0 for key in OTHER_COST_KEYS},
    }

    if costs_file is None:
        return {device_id: dict(default_entry) for device_id in normalized_ids}

    costs_path = Path(costs_file)
    if not costs_path.exists():
        fallback_costs_path = Path(__file__).parent / 'Koszty' / costs_path.name
        if fallback_costs_path.exists():
            costs_path = fallback_costs_path
        else:
            print(f"⚠ Brak pliku kosztów: {costs_path}")
            return {device_id: dict(default_entry) for device_id in normalized_ids}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(costs_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"⚠ Nie udało się odczytać pliku kosztów: {e}")
        return {device_id: dict(default_entry) for device_id in normalized_ids}

    sheet_name = _resolve_sheet_name_by_month(wb, month_str)
    if sheet_name is None:
        print(f"⚠ Nie znaleziono arkusza kosztów dla miesiąca {month_str} w pliku: {costs_path.name}")
        return {device_id: dict(default_entry) for device_id in normalized_ids}

    ws = wb[sheet_name]
    global_costs = dict(default_entry)
    rent_by_device = _load_rent_costs_by_device(
        month_str,
        rent_file=rent_file,
        rent_sheet=rent_sheet,
    )
    amortyzacja_by_device, _ = _load_amortyzacja_by_device_for_month(month_str)

    # Koszty TVM (sekcje w kolumnie B)
    global_costs['elavon'] = 0.0
    
    global_costs['poczta_polska'] = _extract_monthly_value_in_section(ws, 'POCZTA POLSKA', month_str)
    # Amortyzacja jest odczytywana per automat z plików amortyzacja_25/amortyzacja_26.
    global_costs['amortyzacja'] = 0.0
    papier_year_total, papier_year, papier_mode = _resolve_annual_section_cost_for_month(wb, month_str, 'PAPIER')
    global_costs['papier'] = float(papier_year_total or 0.0) / 12.0
    if papier_mode == 'previous':
        print(f"INFO {month_str} PAPIER: użyto wartości rocznej z {papier_year} (fallback); miesięcznie {global_costs['papier']:.2f}")
    else:
        print(f"INFO {month_str} PAPIER: roczna suma {papier_year_total or 0.0:.2f} -> miesięcznie {global_costs['papier']:.2f}")
    global_costs['transmisja_danych'] = _extract_monthly_value_in_section(ws, 'TELEFONY/INTERNET', month_str)
    global_costs['utrzymanie_oprogramowania'] = _extract_monthly_value_in_section(ws, 'NORDPLUS', month_str)
    ubezpieczenie_year_total, ubezpieczenie_year, ubezpieczenie_mode = _resolve_annual_section_cost_for_month(wb, month_str, 'UBEZPIECZENIE AUTOMATÓW')
    global_costs['ubezpieczenie'] = float(ubezpieczenie_year_total or 0.0) / 12.0
    if ubezpieczenie_mode == 'previous':
        print(f"INFO {month_str} UBEZPIECZENIE: użyto wartości rocznej z {ubezpieczenie_year} (fallback); miesięcznie {global_costs['ubezpieczenie']:.2f}")
    else:
        print(f"INFO {month_str} UBEZPIECZENIE: roczna suma {ubezpieczenie_year_total or 0.0:.2f} -> miesięcznie {global_costs['ubezpieczenie']:.2f}")

    # Serwis z pliku serwis_2026.xlsx; tu brak wartości globalnej.
    global_costs['serwis'] = 0.0

    # Czynsz i prąd są pobierane per automat z pliku najmu.
    global_costs['czynsz'] = 0.0
    global_costs['prad'] = 0.0
    global_costs['it_card'] = 0.0

    # Other project costs
    global_costs['non_tvm'] = _extract_monthly_value_from_header_table(
        ws,
        month_str,
        header_label='Non TVMs',
        month_col=9,
        value_col=10,
    )
    global_costs['project_variable_costs'] = _extract_monthly_value_from_header_table(
        ws,
        month_str,
        header_label='Project',
        month_col=9,
        value_col=11,
    )
    # OH będzie liczony per-device poniżej

    device_count = len(normalized_ids)
    costs_by_device = {}
    service_cost_by_device = service_cost_by_device or {}
    cash_total_by_device = {
        int(device_id): float(value or 0.0)
        for device_id, value in (cash_total_by_device or {}).items()
    }
    poczta_polska_month_total = float(global_costs.get('poczta_polska', 0.0) or 0.0)
    poczta_polska_eligible_ids = {
        device_id
        for device_id in normalized_ids
        if cash_total_by_device.get(device_id, 0.0) > 0.0
    }
    if poczta_polska_month_total and not poczta_polska_eligible_ids:
        print(
            f"WARN POCZTA POLSKA {month_str}: brak automatow z Gotowka Suma > 0; "
            f"koszt {poczta_polska_month_total:,.2f} nie zostal rozdzielony"
        )

    it_card_switch_dates = it_card_switch_dates or {}
    elavon_month_total = _load_elavon_total_from_costs_file(month_str, costs_path)

    # ELAVON dzielony proporcjonalnie do liczby dni pracy na ELAVON w miesiącu.
    elavon_days_by_device = {}
    total_elavon_days = 0.0
    for device_id in normalized_ids:
        it_card_ratio = _it_card_active_ratio_for_month(
            it_card_switch_dates.get(device_id),
            month_str,
        )
        elavon_ratio = 1.0 - it_card_ratio
        days = elavon_ratio * _days_in_month(month_str)
        elavon_days_by_device[device_id] = days
        total_elavon_days += days

    for device_id in normalized_ids:
        per_device = {}
        for key, value in global_costs.items():
            if key == 'elavon':
                if elavon_month_total > 0 and total_elavon_days > 0:
                    per_device[key] = float(elavon_month_total) * (elavon_days_by_device[device_id] / total_elavon_days)
                else:
                    per_device[key] = 0.0
            elif key == 'oh':
                continue
            elif key == 'poczta_polska':
                # Poczta Polska dzieli się wyłącznie na automaty z Gotówka Suma > 0.
                if poczta_polska_eligible_ids and device_id in poczta_polska_eligible_ids:
                    per_device[key] = poczta_polska_month_total / len(poczta_polska_eligible_ids)
                else:
                    per_device[key] = 0.0
            else:
                per_device[key] = float(value or 0.0) / device_count

        rent_entry = rent_by_device.get(device_id, {})
        per_device['czynsz'] = float(rent_entry.get('czynsz', 0.0) or 0.0)
        per_device['prad'] = float(rent_entry.get('prad', 0.0) or 0.0)
        per_device['utrzymanie_oprogramowania'] = 20.0
        per_device['transmisja_danych'] = 5.0
        per_device['amortyzacja'] = float(amortyzacja_by_device.get(device_id, 0.0) or 0.0)
        if device_id in service_cost_by_device:
            per_device['serwis'] = float(service_cost_by_device.get(device_id, 0.0) or 0.0)
        
        # Calculate OH per device: (NON_TVM + Project_variable_costs + TVM_sum) * 0.2
        tvm_sum = sum(float(per_device.get(key, 0.0) or 0.0) for key in TVM_COST_KEYS)
        per_device['oh'] = (
            float(per_device.get('non_tvm', 0.0) or 0.0) +
            float(per_device.get('project_variable_costs', 0.0) or 0.0)
        ) * 0.2
        
        costs_by_device[device_id] = per_device

    return costs_by_device


def _format_worksheet_columns(ws):
    """
    Dostosowuje szerokości kolumn arkusza.
    """
    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        column = get_column_letter(col_idx)
        ws.column_dimensions[column].width = min(int(max_length * 1.2), 50)


def _month_label_pl(month_str):
    """
    Zwraca nazwę miesiąca po polsku do nagłówków summary.
    """
    month_no = int(month_str.split('-')[1])
    labels = {
        1: 'styczeń',
        2: 'luty',
        3: 'marzec',
        4: 'kwiecień',
        5: 'maj',
        6: 'czerwiec',
        7: 'lipiec',
        8: 'sierpień',
        9: 'wrzesień',
        10: 'październik',
        11: 'listopad',
        12: 'grudzień',
    }
    return labels.get(month_no, month_str)


def _create_profit_loss_summary_sheet(wb, month_to_profit_loss):
    """
    Tworzy 3. arkusz zbiorczy Profit/loss per automat z podziałem i sumowaniem rocznym.
    Wszystkie wartości pozostają liczbowe, także dla urządzeń magazynowych.
    """
    ws = wb.create_sheet(title='ProfitLoss_Summary')
    ordered_months = sorted(month_to_profit_loss.keys())
    
    # 1. Group months by year
    years_to_months = {}
    for m in ordered_months:
        year = int(m.split('-')[0])
        years_to_months.setdefault(year, []).append(m)
        
    year_to_cols = {}
    headers = ['Nr aut.']
    current_col = 2
    
    # 2. Build headers dynamically
    for year in sorted(years_to_months.keys()):
        start_col = current_col
        for m in years_to_months[year]:
            headers.append(f"Profit/loss {_month_label_pl(m)}")
            current_col += 1
        headers.append("Zestawienie roczne (+/-)")
        end_col = current_col
        year_to_cols[year] = (start_col, end_col)
        current_col += 1

    # 3. Create the Year Group Header (Row 1)
    ws.cell(row=1, column=1).value = ''
    ws.cell(row=1, column=1).fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')

    for year, (start_col, end_col) in year_to_cols.items():
        if start_col == end_col:
            year_cell = ws.cell(row=1, column=start_col)
        else:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            year_cell = ws.cell(row=1, column=start_col)
            
        year_cell.value = f"Rok {year}"
        year_cell.font = Font(bold=True, color='FFFFFF', size=12)
        year_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        year_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4. Create the Column Header (Row 2)
    header_row = 2
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color='FFFFFF')
        
        if col_num == 1:
            cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 5. Populate the Data
    all_device_ids = sorted({
        int(device_id)
        for month_map in month_to_profit_loss.values()
        for device_id in month_map.keys()
    })

    row_num = 3
    for device_id in all_device_ids:
        ws.cell(row=row_num, column=1).value = device_id
        
        col_idx = 2
        for year in sorted(years_to_months.keys()):
            first_month_col = get_column_letter(col_idx)
            for m in years_to_months[year]:
                value = float(month_to_profit_loss.get(m, {}).get(device_id, 0.0) or 0.0)
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.number_format = '#,##0.00'
                col_idx += 1

            last_month_col = get_column_letter(col_idx - 1)
            # Add yearly sum formula
            sum_cell = ws.cell(row=row_num, column=col_idx)
            sum_cell.value = f"=SUM({first_month_col}{row_num}:{last_month_col}{row_num})"
            sum_cell.number_format = '#,##0.00'
            col_idx += 1
            
        row_num += 1

    last_data_row = row_num - 1

    # Podsumowanie dla każdego miesiąca (Suma)
    sum_row = row_num
    ws.cell(row=sum_row, column=1).value = "Suma"
    ws.cell(row=sum_row, column=1).font = Font(bold=True)
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=sum_row, column=col)
        cell.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)
    
    row_num += 1

    if last_data_row >= 2:
        first_value_col_letter = get_column_letter(2)
        last_value_col_letter = get_column_letter(len(headers))
        summary_range = f"{first_value_col_letter}3:{last_value_col_letter}{last_data_row}"

        ws.conditional_formatting.add(
            summary_range,
            CellIsRule(
                operator='lessThan',
                formula=['-100'],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            ),
        )
        ws.conditional_formatting.add(
            summary_range,
            CellIsRule(
                operator='between',
                formula=['-100', '100'],
                fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            ),
        )
        ws.conditional_formatting.add(
            summary_range,
            CellIsRule(
                operator='greaterThan',
                formula=['100'],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            ),
        )

    # Add black borders to data cells
    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in range(2, row_num):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Tabela sezonowości
    curr_row = row_num + 2
    for year in sorted(years_to_months.keys()):
        months_in_year = years_to_months[year]
        if not months_in_year:
            continue
            
        start_col, end_col = year_to_cols[year]
        
        ws.cell(row=curr_row, column=1).value = f"Wskaźnik sezonowości Rok {year}"
        ws.cell(row=curr_row, column=1).font = Font(bold=True)
        ws.cell(row=curr_row+1, column=1).value = "Miesiąc"
        ws.cell(row=curr_row+2, column=1).value = "Wynik miesięczny"
        ws.cell(row=curr_row+3, column=1).value = "Średnia"
        ws.cell(row=curr_row+4, column=1).value = "Odchylenie (+/-)"
        ws.cell(row=curr_row+5, column=1).value = "Wskaźnik (%)"
        
        for r_offset in range(1, 6):
            ws.cell(row=curr_row+r_offset, column=1).font = Font(bold=True)
            
        # Wypisanie miesięcy 
        col_idx = 2
        first_month_col_letter = get_column_letter(col_idx)
        last_month_col_letter = get_column_letter(col_idx + len(months_in_year) - 1)
        avg_formula = f"=AVERAGE({first_month_col_letter}{curr_row+2}:{last_month_col_letter}{curr_row+2})"
        
        for i, m in enumerate(months_in_year):
            c_current = col_idx + i
            c_letter = get_column_letter(c_current)
            orig_col_letter = get_column_letter(start_col + i)
            
            # Miesiąc
            c_month = ws.cell(row=curr_row+1, column=c_current)
            c_month.value = _month_label_pl(m)
            c_month.alignment = Alignment(horizontal='center')
            
            # Wynik miesięczny
            c_wynik = ws.cell(row=curr_row+2, column=c_current)
            c_wynik.value = f"={orig_col_letter}{sum_row}"
            c_wynik.number_format = '#,##0.00'
            
            # Średnia
            c_srednia = ws.cell(row=curr_row+3, column=c_current)
            c_srednia.value = avg_formula
            c_srednia.number_format = '#,##0.00'
            
            # Odchylenie
            c_odchylenie = ws.cell(row=curr_row+4, column=c_current)
            c_odchylenie.value = f"={c_letter}{curr_row+2}-{c_letter}{curr_row+3}"
            c_odchylenie.number_format = '#,##0.00'
            
            # Wskaźnik
            c_wskaznik = ws.cell(row=curr_row+5, column=c_current)
            c_wskaznik.value = f"=IFERROR({c_letter}{curr_row+2}/{c_letter}{curr_row+3}, 0)"
            c_wskaznik.number_format = '0.00%'
            
        curr_row += 7

    # Remove gridlines
    ws.sheet_view.showGridLines = False

    _format_worksheet_columns(ws)


def get_monthly_revenue(
    conn,
    month_str,
    schema=SCHEMA,
    actioncodes=None,
    mctype=None,
    use_device_range_filter=True,
):
    """
    Pobiera obrót brutto per automat za dany miesiąc z tabeli <schema>.moneystats.
    Opcjonalnie filtruje po:
    - actioncodes: lista kodów ms_actioncode
    - mctype: wartość ms_mctype
    - use_device_range_filter: zakresy DEVICE_ID_RANGES
    month_str: 'YYYY-MM'
    Zwraca: dict {device_id: {'obrot_brutto_zl': float, 'liczba_transakcji': int}}
    """
    start_date, end_date = get_month_range(month_str)

    cursor = conn.cursor()

    params = [start_date, end_date]

    action_filter_sql = sql.SQL('')
    if actioncodes:
        placeholders = sql.SQL(', ').join([sql.Placeholder()] * len(actioncodes))
        action_filter_sql = sql.SQL(" AND ms_actioncode IN ({vals})").format(vals=placeholders)
        params.extend(actioncodes)

    mctype_filter_sql = sql.SQL('')
    if mctype is not None:
        mctype_filter_sql = sql.SQL(" AND ms_mctype = %s")
        params.append(mctype)

    device_filter_sql = sql.SQL('')
    if use_device_range_filter:
        device_filter_sql = sql.SQL(" AND {device_filter}").format(
            device_filter=_build_device_filter_sql(sql.Identifier('ms_deviceid'))
        )

    query = sql.SQL(
        """
        SELECT
            ms_deviceid AS device_id,
            SUM(COALESCE(ms_vtotal, 0)) AS obrot_brutto_zl,
            COUNT(*) AS liczba_transakcji
        FROM {schema}.moneystats
        WHERE ms_creadate >= %s
          AND ms_creadate < %s
          {action_filter}
          {mctype_filter}
          {device_filter}
        GROUP BY ms_deviceid
        ORDER BY device_id
        """
    ).format(
        schema=sql.Identifier(schema),
        action_filter=action_filter_sql,
        mctype_filter=mctype_filter_sql,
        device_filter=device_filter_sql,
    )

    cursor.execute(query, tuple(params))
    rows = cursor.fetchall()
    cursor.close()
    
    # Konwertuj na dict
    revenue_data = {}
    for row in rows:
        device_id, obrot, liczba = row
        revenue_data[device_id] = {
            'obrot_brutto_zl': float(obrot),
            'liczba_transakcji': int(liczba)
        }
    
    return revenue_data


def _normalize_carrier_code(carrier):
    """
    Normalizuje kod przewoźnika do postaci technicznej używanej w kluczach.
    Usuwa diakrytyki, np. KMŁ -> KML.
    """
    normalized = str(carrier or '').strip().upper()
    normalized = unicodedata.normalize('NFKD', normalized)
    normalized = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return normalized


def _build_amount_expr_sql(amount_col_name):
    """
    Buduje wyrażenie SQL kwoty obrotu.
    Dla pól finansowych monitora przechowywanych w groszach zwraca wartość w złotych.
    """
    amount_col_sql = sql.Identifier(amount_col_name)
    if amount_col_name.lower() in {'fin_nalezn', 'fin_ptu_kwota'}:
        return sql.SQL("(COALESCE({amount_col}, 0)::numeric / 100.0)").format(
            amount_col=amount_col_sql
        )
    return sql.SQL("COALESCE({amount_col}, 0)").format(amount_col=amount_col_sql)


def _build_ticket_count_expr_sql(ticket_count_col_name):
    """
    Buduje wyrazenie SQL liczby sprzedanych biletow.
    """
    if not ticket_count_col_name:
        return sql.SQL("0")
    return sql.SQL("COALESCE({ticket_count_col}, 0)").format(
        ticket_count_col=sql.Identifier(ticket_count_col_name)
    )


def _display_carrier_label(carrier_code):
    """
    Zwraca etykietę przewoźnika do nagłówka raportu.
    """
    normalized = _normalize_carrier_code(carrier_code)
    if normalized == 'KML':
        return 'KMŁ'
    return normalized


def _resolve_transactions_table_for_carrier(carrier_code, default_table):
    """
    Zwraca nazwę tabeli transakcji dla przewoźnika.
    Pozwala nadpisać domyślną tabelę dla wybranych schematów.
    """
    normalized = _normalize_carrier_code(carrier_code)
    return TRANSACTIONS_TABLE_OVERRIDES.get(normalized, default_table)


def _resolve_transactions_device_col_for_carrier(carrier_code, default_device_col):
    """
    Zwraca kolumnę urządzenia dla przewoźnika (string lub lista fallbacków).
    """
    normalized = _normalize_carrier_code(carrier_code)
    return TRANSACTIONS_DEVICE_COL_OVERRIDES.get(normalized, default_device_col)


def _resolve_transactions_amount_col_for_carrier(carrier_code, default_amount_col):
    """
    Zwraca kolumnę kwoty dla przewoźnika (string lub lista fallbacków).
    """
    normalized = _normalize_carrier_code(carrier_code)
    return TRANSACTIONS_AMOUNT_COL_OVERRIDES.get(normalized, default_amount_col)


def _resolve_transactions_payment_method_col_for_carrier(carrier_code, default_payment_method_col):
    """
    Zwraca kolumnę metody płatności dla przewoźnika (string lub lista fallbacków).
    """
    normalized = _normalize_carrier_code(carrier_code)
    return TRANSACTIONS_PAYMENT_METHOD_COL_OVERRIDES.get(normalized, default_payment_method_col)


def _resolve_transactions_ticket_count_col_for_carrier(carrier_code, default_ticket_count_col):
    """
    Zwraca kolumne liczby biletow dla przewoznika (string lub lista fallbackow).
    """
    normalized = _normalize_carrier_code(carrier_code)
    return TRANSACTIONS_TICKET_COUNT_COL_OVERRIDES.get(normalized, default_ticket_count_col)


def _preview_column_name(column_ref):
    """
    Zwraca nazwę kolumny do podglądu SQL (dla list fallbacków używa pierwszej pozycji).
    """
    if isinstance(column_ref, (list, tuple)):
        return column_ref[0] if column_ref else ''
    return column_ref


def _get_table_columns(conn, schema_name, table_name):
    """
    Zwraca listę kolumn tabeli lub pustą listę, gdy tabela nie istnieje.
    """
    cursor = conn.cursor()
    cursor.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s
          AND table_name = %s
        ORDER BY ordinal_position
        """,
        (schema_name, table_name),
    )
    rows = cursor.fetchall()
    cursor.close()
    return [r[0] for r in rows]


def _resolve_transactions_source(
    conn,
    schema_name,
    table_name,
    amount_col,
    date_col,
    device_col=None,
    payment_method_col=None,
    ticket_count_col=None,
):
    """
    Rozwiązuje źródło transakcji z jednego schematu.
    Wykrywa kolumny urządzenia i przewoźnika lub używa jawnych nazw z CLI.
    Zwraca dict lub None, gdy brak poprawnego źródła.
    """
    columns = _get_table_columns(conn, schema_name, table_name)
    if not columns:
        return None

    lowered_map = {c.lower(): c for c in columns}

    if isinstance(amount_col, (list, tuple)):
        picked_amount_col = _pick_first_matching_column(columns, list(amount_col))
    else:
        picked_amount_col = lowered_map.get(str(amount_col).lower())

    if isinstance(date_col, (list, tuple)):
        picked_date_col = _pick_first_matching_column(columns, list(date_col))
    else:
        picked_date_col = lowered_map.get(str(date_col).lower())

    if picked_amount_col is None or picked_date_col is None:
        return None

    if device_col:
        if isinstance(device_col, (list, tuple)):
            picked_device_col = _pick_first_matching_column(columns, list(device_col))
        else:
            picked_device_col = lowered_map.get(str(device_col).lower())
    else:
        picked_device_col = _pick_first_matching_column(columns, TRANSACTIONS_DEVICE_COL_CANDIDATES)

    if picked_device_col is None:
        return None

    if payment_method_col:
        if isinstance(payment_method_col, (list, tuple)):
            picked_payment_method_col = _pick_first_matching_column(columns, list(payment_method_col))
        else:
            picked_payment_method_col = lowered_map.get(str(payment_method_col).lower())
    else:
        picked_payment_method_col = _pick_first_matching_column(
            columns,
            TRANSACTIONS_PAYMENT_METHOD_COL_CANDIDATES,
        )

    if picked_payment_method_col is None:
        return None

    picked_ticket_count_col = None
    if ticket_count_col:
        if isinstance(ticket_count_col, (list, tuple)):
            picked_ticket_count_col = _pick_first_matching_column(columns, list(ticket_count_col))
        else:
            picked_ticket_count_col = lowered_map.get(str(ticket_count_col).lower())

    return {
        'schema': schema_name,
        'table': table_name,
        'device_col': picked_device_col,
        'amount_col': picked_amount_col,
        'date_col': picked_date_col,
        'payment_method_col': picked_payment_method_col,
        'ticket_count_col': picked_ticket_count_col,
    }


def build_transactions_debug_query(
    schema_name,
    month_str,
    table_name=TRANSACTIONS_SOURCE_CONFIG['table'],
    amount_col=TRANSACTIONS_SOURCE_CONFIG['amount_col'],
    date_col=TRANSACTIONS_SOURCE_CONFIG['date_col'],
    device_col=TRANSACTIONS_SOURCE_CONFIG['device_col'],
    payment_method_col=TRANSACTIONS_SOURCE_CONFIG['payment_method_col'],
):
    """
    Buduje gotowe zapytanie SQL diagnostyczne dla jednego schematu przewoźnika.
    """
    amount_col = _preview_column_name(amount_col)
    date_col = _preview_column_name(date_col)
    device_col = _preview_column_name(device_col)
    payment_method_col = _preview_column_name(payment_method_col)
    start_date, end_date = get_month_range_closed(month_str)
    amount_expr = f"COALESCE({amount_col}, 0)"
    if amount_col.lower() in {'fin_nalezn', 'fin_ptu_kwota'}:
        amount_expr = f"(COALESCE({amount_col}, 0)::numeric / 100.0)"
    return f"""
SELECT
    {device_col} AS device_id,
    {payment_method_col} AS fin_spos_opl,
    SUM({amount_expr}) AS obrot_metoda_zl,
    COUNT(*) AS liczba_transakcji
FROM {schema_name}.{table_name}
WHERE {device_col}::text ~ '^[0-9]+$'
  AND {_build_device_filter_preview_expr(device_col)}
  AND {date_col} >= TIMESTAMP '{start_date.strftime('%Y-%m-%d %H:%M:%S')}'
  AND {date_col} <= TIMESTAMP '{end_date.strftime('%Y-%m-%d %H:%M:%S')}'
    AND {payment_method_col}::text IN ('1', '2', '6')
GROUP BY {device_col}, {payment_method_col}
ORDER BY device_id, {payment_method_col};
""".strip()


def get_monthly_revenue_by_carrier(
    conn,
    month_str,
    carriers=None,
    table_name=TRANSACTIONS_SOURCE_CONFIG['table'],
    amount_col=TRANSACTIONS_SOURCE_CONFIG['amount_col'],
    date_col=TRANSACTIONS_SOURCE_CONFIG['date_col'],
    device_col=TRANSACTIONS_SOURCE_CONFIG['device_col'],
    payment_method_col=TRANSACTIONS_SOURCE_CONFIG['payment_method_col'],
    ticket_count_col=TRANSACTIONS_TICKET_COUNT_COL,
    use_device_range_filter=True,
):
    """
    Pobiera obrót brutto per automat i per przewoźnik z tabeli <carrier_schema>.transactions.
    Zwraca: dict {
        device_id: {
            'by_carrier': {
                carrier: {
                    'obrot_brutto_zl': float,
                    'liczba_transakcji': int,
                    'liczba_biletow': int,
                    'by_payment_method': {
                        'gotowka': float,
                        'karta': float,
                        'blik': float,
                    }
                }
            }
        }
    }
    """
    start_date, end_date = get_month_range_closed(month_str)
    carriers = carriers or CARRIERS
    expected_carriers = sorted({_normalize_carrier_code(carrier) for carrier in carriers})

    revenue_data = {}

    for carrier in carriers:
        carrier_code = _normalize_carrier_code(carrier)

        # Reguła dla KS: tylko miesiące '2025-01', '2025-02', '2025-03', '2025-04'
        if carrier_code == 'KS' and month_str not in ('2025-01', '2025-02', '2025-03', '2025-04'):
            continue

        carrier_table_name = _resolve_transactions_table_for_carrier(carrier_code, table_name)
        carrier_device_col = _resolve_transactions_device_col_for_carrier(carrier_code, device_col)
        carrier_amount_col = _resolve_transactions_amount_col_for_carrier(carrier_code, amount_col)
        carrier_payment_method_col = _resolve_transactions_payment_method_col_for_carrier(
            carrier_code,
            payment_method_col,
        )
        carrier_ticket_count_col = _resolve_transactions_ticket_count_col_for_carrier(
            carrier_code,
            ticket_count_col,
        )
        source = _resolve_transactions_source(
            conn,
            schema_name=carrier_code,
            table_name=carrier_table_name,
            amount_col=carrier_amount_col,
            date_col=date_col,
            device_col=carrier_device_col,
            payment_method_col=carrier_payment_method_col,
            ticket_count_col=carrier_ticket_count_col,
        )
        if source is None:
            print(
                f"⚠ Pomijam przewoźnika {carrier} (schemat {carrier_code}): "
                f"brak źródła {carrier_code}.{carrier_table_name} lub wymaganych kolumn "
                f"({carrier_device_col}, {date_col}, {carrier_amount_col}, {carrier_payment_method_col})"
            )
            continue

        if not source.get('ticket_count_col'):
            print(
                f"WARN {carrier_code}: brak kolumny {_preview_column_name(carrier_ticket_count_col)} "
                f"w {source['schema']}.{source['table']} - liczba biletow bedzie 0"
            )

        params = [start_date, end_date]
        device_filter_sql = sql.SQL('')
        if use_device_range_filter:
            device_expr = sql.SQL("{device_col}::bigint").format(
                device_col=sql.Identifier(source['device_col'])
            )
            device_filter_sql = sql.SQL(" AND {device_filter}").format(
                device_filter=_build_device_filter_sql(device_expr)
            )
        numeric_device_sql = sql.SQL(" AND {device_col}::text ~ '^[0-9]+$'").format(
            device_col=sql.Identifier(source['device_col'])
        )

        query = sql.SQL(
            """
            SELECT
                {device_col} AS device_id,
                {payment_method_col} AS fin_spos_opl,
                SUM({amount_expr}) AS obrot_metoda_zl,
                COUNT(*) AS liczba_transakcji,
                SUM({ticket_count_expr}) AS liczba_biletow
            FROM {schema}.{table}
            WHERE {date_col} >= %s
                            AND {date_col} <= %s
                            {numeric_device_filter}
              AND {payment_method_col}::text IN ('1', '2', '6')
              {device_filter}
            GROUP BY {device_col}, {payment_method_col}
            ORDER BY device_id, {payment_method_col}
            """
        ).format(
            schema=sql.Identifier(source['schema']),
            table=sql.Identifier(source['table']),
            device_col=sql.Identifier(source['device_col']),
            amount_expr=_build_amount_expr_sql(source['amount_col']),
            ticket_count_expr=_build_ticket_count_expr_sql(source.get('ticket_count_col')),
            date_col=sql.Identifier(source['date_col']),
            payment_method_col=sql.Identifier(source['payment_method_col']),
            numeric_device_filter=numeric_device_sql,
            device_filter=device_filter_sql,
         )

        cursor = conn.cursor()
        try:
            cursor.execute(query, tuple(params))
            rows = cursor.fetchall()
            print(
                f"✓ Obrót {carrier_code}: {source['schema']}.{source['table']} "
                f"({source['device_col']}, {source['amount_col']}, {source['date_col']}, {source['payment_method_col']}) -> {len(rows)} rekordów"
            )
        except psycopg2.Error as e:
            print(f"⚠ Błąd pobierania obrotu dla {carrier_code}: {e}")
            rows = []
        finally:
            cursor.close()

        for row in rows:
            device_id, payment_method_value, obrot, liczba, liczba_biletow = row
            try:
                device_id_int = int(device_id)
            except (TypeError, ValueError):
                continue

            method_key = PAYMENT_METHOD_CODE_TO_KEY.get(str(payment_method_value).strip())
            if method_key is None:
                continue

            revenue_data.setdefault(device_id_int, {'by_carrier': {}})
            carrier_entry = revenue_data[device_id_int]['by_carrier'].setdefault(
                carrier_code,
                {
                    'obrot_brutto_zl': 0.0,
                    'liczba_transakcji': 0,
                    'liczba_biletow': 0,
                    'by_payment_method': {key: 0.0 for key in PAYMENT_METHOD_KEYS},
                },
            )

            method_amount = float(obrot or 0.0)
            method_count = int(liczba or 0)
            ticket_count = int(liczba_biletow or 0)
            carrier_entry['by_payment_method'][method_key] += method_amount
            carrier_entry['obrot_brutto_zl'] += method_amount
            carrier_entry['liczba_transakcji'] += method_count
            carrier_entry['liczba_biletow'] += ticket_count

        carrier_total = sum(
            rev_entry.get('obrot_brutto_zl', 0.0)
            for rev_device in revenue_data.values()
            for code, rev_entry in rev_device.get('by_carrier', {}).items()
            if code == carrier_code
        )
        carrier_ticket_total = sum(
            int(rev_entry.get('liczba_biletow', 0) or 0)
            for rev_device in revenue_data.values()
            for code, rev_entry in rev_device.get('by_carrier', {}).items()
            if code == carrier_code
        )
        print(f"  Suma {carrier_code}: {carrier_total:,.2f} zł")

        print(f"  Bilety {carrier_code}: {carrier_ticket_total:,}")

    observed_carriers = sorted({
        code
        for rev_device in revenue_data.values()
        for code in rev_device.get('by_carrier', {}).keys()
    })
    print(f"✓ Przewoźnicy w obrocie: {observed_carriers}")
    missing_expected = sorted(set(expected_carriers) - set(observed_carriers))
    if missing_expected:
        print(f"⚠ Brak danych obrotu dla przewoźników: {missing_expected}")

    return revenue_data

 
# === EKSPORT DO EXCEL (KOLUMNY 1-2, RESZTA PLACEHOLDER) ===

def export_to_excel_PL(
    dictionary_comparison,
    revenue_data,
    month_str,
    filename=None,
    commission_rules=None,
    carriers=None,
    location_by_device=None,
    automat_type_by_device=None,
    costs_file=DEFAULT_COSTS_FILE,
    rent_file=DEFAULT_RENT_FILE,
    rent_sheet=None,
    amortyzacja_file=DEFAULT_AMORTYZACJA_FILE,
    amortyzacja_sheet=DEFAULT_AMORTYZACJA_SHEET,
    workbook=None,
    sheet_title=None,
    save_workbook=True,
    service_cost_by_device=None,
    it_card_switch_dates=None,
    relocation_active_device_ids=None,
    warehouse_device_ids=None,
    included_device_ids=None,
):
    """
    Eksportuje raport P&L do Excela.
    Układ danych: jeden wiersz per automat, kolumny brutto i metody płatności per przewoźnik.
    """
    if commission_rules is None:
        commission_rules = {}

    carrier_order = []
    for carrier in (carriers or CARRIERS):
        code = _normalize_carrier_code(carrier)
        if code not in carrier_order:
            carrier_order.append(code)

    for rev_entry in revenue_data.values():
        by_carrier = rev_entry.get('by_carrier') if isinstance(rev_entry, dict) else None
        if isinstance(by_carrier, dict):
            for code in by_carrier.keys():
                if code not in carrier_order:
                    carrier_order.append(code)

    filepath = None
    if save_workbook:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        if filename is None:
            filename = build_output_filename(month_str, naming_mode='default')
        filepath = OUTPUT_DIR / filename

    if workbook is None:
        wb = Workbook()
        ws = wb.active
    else:
        wb = workbook
        active_ws = wb.active
        is_placeholder_active = (
            active_ws is not None
            and len(wb.worksheets) == 1
            and active_ws.max_row == 1
            and active_ws.max_column == 1
            and active_ws['A1'].value is None
        )
        if is_placeholder_active:
            ws = active_ws
        else:
            ws = wb.create_sheet()

    assert ws is not None

    ws.title = sheet_title or month_str

    title_row = 1
    header_row = 2
    data_start_row = 3

    title_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    warehouse_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    number_block_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    summary_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, color='FFFFFF')
    base_font = Font(size=11)

    thin_side = Side(style='thin', color='000000')
    medium_side = Side(style='medium', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    def _set_cell_style(r, c, *, fill=None, font=None, alignment=None, border=None, number_format=None):
        cell = ws.cell(row=r, column=c)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        if number_format is not None:
            cell.number_format = number_format
        return cell

    def _carrier_header_label(code):
        if str(code).upper() == 'KML':
            return 'KMŁ'
        return _display_carrier_label(code)

    detected_carriers = []
    detected_set = set()
    for rev_entry in revenue_data.values():
        by_carrier = rev_entry.get('by_carrier') if isinstance(rev_entry, dict) else None
        if isinstance(by_carrier, dict):
            for code in by_carrier.keys():
                normalized_code = _normalize_carrier_code(code)
                if normalized_code not in detected_set:
                    detected_set.add(normalized_code)
                    detected_carriers.append(normalized_code)

    if not detected_carriers:
        detected_carriers = list(carrier_order)

    preferred_order = ['ARP', 'IC', 'KD', 'KML', 'KW', 'LKA', 'PR', 'SKM']
    carrier_order = [code for code in preferred_order if code in detected_carriers]
    carrier_order += [code for code in detected_carriers if code not in carrier_order]

    # === KONFIGURACJA KOLUMN (NOWA KOLEJNOŚĆ) ===

    # 1. INFO
    num1_col = 1
    info_nr_col = 2
    info_loc_col = 3
    info_type_col = 4
    info_start_col = info_nr_col
    info_end_col = info_type_col

    # 2. PODSUMOWANIE (Przeniesione na 2. miejsce)
    num2_col = info_end_col + 1
    podsum_start_col = num2_col + 1
    podsum_cols = {
        'koszty': podsum_start_col,
        'przychody': podsum_start_col + 1,
        'brutto_suma': podsum_start_col + 2,
        'netto_suma': podsum_start_col + 3,
        'karta_suma': podsum_start_col + 4,
        'blik_suma': podsum_start_col + 5,
        'gotowka_suma': podsum_start_col + 6,
        'bilety_suma': podsum_start_col + 7,
        'profit_loss': podsum_start_col + 8,
    }
    podsum_end_col = podsum_start_col + 8

    # 3. DANE PRZEWOŹNIKÓW (Przesunięte na 3. miejsce)
    num3_col = podsum_end_col + 1
    carrier_start_col = num3_col + 1
    carrier_columns = {}
    col_cursor = carrier_start_col
    for code in carrier_order:
        carrier_columns[code] = {
            'start': col_cursor,
            'brutto': col_cursor,
            'prowizja': col_cursor + 1,
            'netto': col_cursor + 2,
            'gotowka': col_cursor + 3,
            'karta': col_cursor + 4,
            'blik': col_cursor + 5,
            'end': col_cursor + 5,
        }
        col_cursor += 6
    carrier_end_col = col_cursor - 1

    # 4. LICZBA BILETÓW (Przesunięte na 4. miejsce)
    num4_col = carrier_end_col + 1
    tickets_start_col = num4_col + 1
    ticket_columns = {}
    col_cursor = tickets_start_col
    for code in carrier_order:
        ticket_columns[code] = col_cursor
        col_cursor += 1
    tickets_end_col = col_cursor - 1

    # 5. PRZYCHODY (Przesunięte na 5. miejsce)
    num5_col = tickets_end_col + 1
    przychody_start_col = num5_col + 1
    przychody_cols = {
        'prowizja_suma': przychody_start_col,
        'interchange': przychody_start_col + 1,
        'dodatkowe_zyski': przychody_start_col + 2,
    }
    przychody_end_col = przychody_start_col + 2

    # 6. KOSZTY (Przesunięte na 6. miejsce)
    num6_col = przychody_end_col + 1
    koszty_start_col = num6_col + 1
    tvm_cost_cols = {}
    col_cursor = koszty_start_col
    for key in TVM_COST_KEYS:
        tvm_cost_cols[key] = col_cursor
        col_cursor += 1
    koszty_tvm_col = col_cursor
    col_cursor += 1
    other_cost_cols = {}
    for key in OTHER_COST_KEYS:
        other_cost_cols[key] = col_cursor
        col_cursor += 1
    dodatkowe_koszty_col = col_cursor
    koszty_end_col = dodatkowe_koszty_col

    # 7. UWAGI
    num7_col = koszty_end_col + 1
    uwagi_col_idx = num7_col + 1
    result_col_idx = podsum_cols['profit_loss']

    # --- RENDEROWANIE NAGŁÓWKÓW (TITLE ROW) ---

    if info_start_col <= info_end_col:
        ws.merge_cells(start_row=title_row, start_column=info_start_col, end_row=title_row, end_column=info_end_col)
    ws.cell(row=title_row, column=info_start_col, value='INFO')
    _set_cell_style(title_row, info_start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    ws.merge_cells(start_row=title_row, start_column=podsum_start_col, end_row=title_row, end_column=podsum_end_col)
    ws.cell(row=title_row, column=podsum_start_col, value='Podsumowanie')
    _set_cell_style(title_row, podsum_start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    if carrier_order:
        for code in carrier_order:
            start_col = carrier_columns[code]['start']
            end_col = carrier_columns[code]['end']
            ws.merge_cells(start_row=title_row, start_column=start_col, end_row=title_row, end_column=end_col)
            ws.cell(row=title_row, column=start_col, value=_carrier_header_label(code))
            _set_cell_style(title_row, start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    if ticket_columns:
        if tickets_start_col < tickets_end_col:
            ws.merge_cells(start_row=title_row, start_column=tickets_start_col, end_row=title_row, end_column=tickets_end_col)
        ws.cell(row=title_row, column=tickets_start_col, value='Liczba biletów')
        _set_cell_style(title_row, tickets_start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    ws.merge_cells(start_row=title_row, start_column=przychody_start_col, end_row=title_row, end_column=przychody_end_col)
    ws.cell(row=title_row, column=przychody_start_col, value='Przychody')
    _set_cell_style(title_row, przychody_start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    ws.merge_cells(start_row=title_row, start_column=koszty_start_col, end_row=title_row, end_column=koszty_end_col)
    ws.cell(row=title_row, column=koszty_start_col, value='KOSZTY')
    _set_cell_style(title_row, koszty_start_col, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    ws.cell(row=title_row, column=uwagi_col_idx, value='Uwagi')
    _set_cell_style(title_row, uwagi_col_idx, fill=title_fill, font=title_font, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

    header_values = {
        info_nr_col: 'Nr TVM',
        info_loc_col: 'Lokalizacja automatu',
        info_type_col: 'Rodzaj automatu',
        przychody_cols['prowizja_suma']: 'Prowizja Suma',
        przychody_cols['interchange']: 'Interchange',
        przychody_cols['dodatkowe_zyski']: 'Dodatkowe zyski',
        koszty_tvm_col: 'Koszty TVM',
        dodatkowe_koszty_col: 'Dodatkowe koszty',
        podsum_cols['koszty']: 'Koszty',
        podsum_cols['przychody']: 'Przychody',
        podsum_cols['brutto_suma']: 'Brutto Suma',
        podsum_cols['netto_suma']: 'Netto Suma',
        podsum_cols['karta_suma']: 'Karta Suma',
        podsum_cols['blik_suma']: 'BLIK Suma',
        podsum_cols['gotowka_suma']: 'Gotówka Suma',
        podsum_cols['bilety_suma']: 'Suma biletów',
        podsum_cols['profit_loss']: 'Profit/loss',
        uwagi_col_idx: 'Uwagi',
    }

    for code in carrier_order:
        label = _carrier_header_label(code)
        cols = carrier_columns[code]
        header_values[cols['brutto']] = f"Brutto {label}"
        header_values[cols['prowizja']] = f"Prowizja {label}"
        header_values[cols['netto']] = f"Netto {label}"
        header_values[cols['gotowka']] = f"{PAYMENT_METHOD_LABELS['gotowka']} {label}"
        header_values[cols['karta']] = f"{PAYMENT_METHOD_LABELS['karta']} {label}"
        header_values[cols['blik']] = f"{PAYMENT_METHOD_LABELS['blik']} {label}"
        if code in ticket_columns:
            header_values[ticket_columns[code]] = label

    for key in TVM_COST_KEYS:
        header_values[tvm_cost_cols[key]] = TVM_COST_LABELS[key]
    for key in OTHER_COST_KEYS:
        header_values[other_cost_cols[key]] = OTHER_COST_LABELS[key]

    for col_num in range(info_nr_col, uwagi_col_idx + 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header_values.get(col_num, '')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # === DANE ===
    row_num = data_start_row
    location_by_device = location_by_device or {}
    automat_type_by_device = automat_type_by_device or {}
    relocation_active_device_ids = {int(device_id) for device_id in (relocation_active_device_ids or set())}
    warehouse_device_ids = {int(device_id) for device_id in (warehouse_device_ids or set())}
    row_to_device_id = {}
    # Upewnij się, że wszystkie klucze są integerami
    if included_device_ids is None:
        all_device_ids = sorted(set(
            [int(k) for k in dictionary_comparison.keys()] +
            [int(k) for k in revenue_data.keys()]
        ))
    else:
        all_device_ids = sorted({int(device_id) for device_id in (included_device_ids or [])})
    rop_eligible_ids = _build_rop_eligible_device_ids(all_device_ids, commission_rules, month_str)
    cash_total_by_device = _build_cash_total_by_device(
        revenue_data,
        all_device_ids,
        warehouse_device_ids=warehouse_device_ids,
    )
    costs_by_device = _get_monthly_costs_per_device(
        month_str,
        all_device_ids,
        costs_file=costs_file,
        rent_file=rent_file,
        rent_sheet=rent_sheet,
        service_cost_by_device=service_cost_by_device,
        it_card_switch_dates=it_card_switch_dates,
        cash_total_by_device=cash_total_by_device,
    )
    profit_loss_by_device = {}
    tickets_by_device = {}
    it_card_switch_dates = it_card_switch_dates or {}
    
    for device_id in all_device_ids:
        # 1-INFO (najpierw mapa lokalizacji zebrana po przewoźnikach)
        lokalizacja = location_by_device.get(device_id, '')
        if not lokalizacja and device_id in dictionary_comparison:
            info = dictionary_comparison[device_id]
            data = info['data']
            lokalizacja = extract_city_name(data.get('description', ''))
        if device_id in warehouse_device_ids:
            lokalizacja = 'magazyn'
        
        rev_by_carrier = {}
        if device_id in revenue_data:
            rev_entry = revenue_data[device_id]
            if isinstance(rev_entry, dict) and isinstance(rev_entry.get('by_carrier'), dict):
                rev_by_carrier = rev_entry['by_carrier']
            elif isinstance(rev_entry, dict):
                if carrier_order:
                    rev_by_carrier = {carrier_order[0]: rev_entry}

        ws.cell(row=row_num, column=info_nr_col, value=device_id)
        ws.cell(row=row_num, column=info_loc_col, value=lokalizacja)
        ws.cell(row=row_num, column=info_type_col, value=automat_type_by_device.get(device_id, ''))
        row_to_device_id[row_num] = device_id

        row_total = 0.0
        row_commission_total = 0.0
        row_interchange_total = 0.0
        row_netto_total = 0.0
        cash_total = 0.0
        card_total = 0.0
        blik_total = 0.0
        month_start, month_end = _month_date_bounds(month_str)
        is_warehouse_device = device_id in warehouse_device_ids
        row_ticket_total = 0
        for carrier_code in carrier_order:
            rev = rev_by_carrier.get(carrier_code)
            amount = 0.0
            netto_amount = 0.0
            ticket_count = 0
            by_payment_method = {key: 0.0 for key in PAYMENT_METHOD_KEYS}
            if isinstance(rev, dict):
                amount = float(rev.get('obrot_brutto_zl', 0.0) or 0.0)
                ticket_count = int(rev.get('liczba_biletow', 0) or 0)
                method_values = rev.get('by_payment_method')
                if isinstance(method_values, dict):
                    for method_key in PAYMENT_METHOD_KEYS:
                        by_payment_method[method_key] = float(method_values.get(method_key, 0.0) or 0.0)
            netto_amount = amount / 1.08

            if is_warehouse_device:
                amount = 0.0
                netto_amount = 0.0
                ticket_count = 0
                by_payment_method = {key: 0.0 for key in PAYMENT_METHOD_KEYS}

            cash_total += by_payment_method['gotowka']
            card_total += by_payment_method['karta']
            blik_total += by_payment_method['blik']
            cashless_for_carrier = by_payment_method['karta'] + by_payment_method['blik']

            rules_for_key = commission_rules.get((int(device_id), carrier_code), [])
            picked_rules = _pick_commission_rules_for_month(rules_for_key, month_start, month_end)
            commission_amount = _commission_amount_from_rules(picked_rules, netto_amount)

            cols = carrier_columns[carrier_code]
            ws.cell(row=row_num, column=cols['brutto'], value=amount)
            ws.cell(row=row_num, column=cols['brutto']).number_format = '#,##0.00'
            ws.cell(row=row_num, column=cols['prowizja'], value=commission_amount)
            ws.cell(row=row_num, column=cols['prowizja']).number_format = '#,##0.00'
            ws.cell(row=row_num, column=cols['netto'], value=netto_amount)
            ws.cell(row=row_num, column=cols['netto']).number_format = '#,##0.00'
            ws.cell(row=row_num, column=cols['gotowka'], value=by_payment_method['gotowka'])
            ws.cell(row=row_num, column=cols['gotowka']).number_format = '#,##0.00'
            ws.cell(row=row_num, column=cols['karta'], value=by_payment_method['karta'])
            ws.cell(row=row_num, column=cols['karta']).number_format = '#,##0.00'
            ws.cell(row=row_num, column=cols['blik'], value=by_payment_method['blik'])
            ws.cell(row=row_num, column=cols['blik']).number_format = '#,##0.00'
            if carrier_code in ticket_columns:
                ws.cell(row=row_num, column=ticket_columns[carrier_code], value=ticket_count)
                ws.cell(row=row_num, column=ticket_columns[carrier_code]).number_format = '#,##0'
            row_total += amount
            row_commission_total += commission_amount
            row_ticket_total += ticket_count
            interchange_rate = INTERCHANGE_RATE_BY_CARRIER.get(carrier_code, 0.0)
            row_interchange_total += cashless_for_carrier * interchange_rate
            row_netto_total += netto_amount

        additional_profit_value = _get_additional_profit_value(month_str, device_id)
        if is_warehouse_device:
            additional_profit_value = 0.0
        ws.cell(row=row_num, column=przychody_cols['prowizja_suma'], value=row_commission_total)
        ws.cell(row=row_num, column=przychody_cols['prowizja_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=przychody_cols['interchange'], value=row_interchange_total)
        ws.cell(row=row_num, column=przychody_cols['interchange']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=przychody_cols['dodatkowe_zyski'], value=additional_profit_value)
        ws.cell(row=row_num, column=przychody_cols['dodatkowe_zyski']).number_format = '#,##0.00'

        cost_entry = dict(costs_by_device.get(device_id, {}))
        # IT CARD: 1.34% z (karta+blik), proporcjonalnie do dni po dacie przejścia.
        cashless_total = card_total + blik_total
        it_card_ratio = _it_card_active_ratio_for_month(it_card_switch_dates.get(device_id), month_str)
        it_card_value = cashless_total * 0.0134 * it_card_ratio
        cost_entry['it_card'] = it_card_value

        tvm_cost_sum = 0.0
        for key in TVM_COST_KEYS:
            val = float(cost_entry.get(key, 0.0) or 0.0)
            tvm_cost_sum += val
            ws.cell(row=row_num, column=tvm_cost_cols[key], value=val)
            ws.cell(row=row_num, column=tvm_cost_cols[key]).number_format = '#,##0.00'

        ws.cell(row=row_num, column=koszty_tvm_col, value=tvm_cost_sum)
        ws.cell(row=row_num, column=koszty_tvm_col).number_format = '#,##0.00'

        other_total = 0.0
        for key in OTHER_COST_KEYS:
            val = float(cost_entry.get(key, 0.0) or 0.0)
            other_total += val
            ws.cell(row=row_num, column=other_cost_cols[key], value=val)
            ws.cell(row=row_num, column=other_cost_cols[key]).number_format = '#,##0.00'

        all_costs_total = tvm_cost_sum + other_total
        ws.cell(row=row_num, column=dodatkowe_koszty_col, value=other_total)
        ws.cell(row=row_num, column=dodatkowe_koszty_col).number_format = '#,##0.00'

        row_przychody_total = row_commission_total + row_interchange_total + additional_profit_value
        ws.cell(row=row_num, column=podsum_cols['koszty'], value=all_costs_total)
        ws.cell(row=row_num, column=podsum_cols['koszty']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['przychody'], value=row_przychody_total)
        ws.cell(row=row_num, column=podsum_cols['przychody']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['brutto_suma'], value=row_total)
        ws.cell(row=row_num, column=podsum_cols['brutto_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['netto_suma'], value=row_netto_total)
        ws.cell(row=row_num, column=podsum_cols['netto_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['karta_suma'], value=card_total)
        ws.cell(row=row_num, column=podsum_cols['karta_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['blik_suma'], value=blik_total)
        ws.cell(row=row_num, column=podsum_cols['blik_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['gotowka_suma'], value=cash_total)
        ws.cell(row=row_num, column=podsum_cols['gotowka_suma']).number_format = '#,##0.00'
        ws.cell(row=row_num, column=podsum_cols['bilety_suma'], value=row_ticket_total)
        ws.cell(row=row_num, column=podsum_cols['bilety_suma']).number_format = '#,##0'
        row_profit_loss = row_przychody_total - all_costs_total
        ws.cell(row=row_num, column=podsum_cols['profit_loss'], value=row_profit_loss)
        ws.cell(row=row_num, column=podsum_cols['profit_loss']).number_format = '#,##0.00'
        profit_loss_by_device[device_id] = float(row_profit_loss)
        tickets_by_device[device_id] = row_ticket_total
        ws.cell(row=row_num, column=uwagi_col_idx, value=None)
        row_num += 1

    last_data_row = row_num - 1

    if last_data_row >= 2:
        summary_row = row_num
        ws.cell(row=summary_row, column=info_loc_col, value='PODSUMOWANIE')
        ws.cell(row=summary_row, column=info_nr_col, value=None)
        ws.cell(row=summary_row, column=uwagi_col_idx, value=None)

        numeric_columns = []
        for code in carrier_order:
            cols = carrier_columns[code]
            numeric_columns.extend([cols['brutto'], cols['prowizja'], cols['netto'], cols['gotowka'], cols['karta'], cols['blik']])
            if code in ticket_columns:
                numeric_columns.append(ticket_columns[code])
        numeric_columns.extend([
            przychody_cols['prowizja_suma'],
            przychody_cols['interchange'],
            przychody_cols['dodatkowe_zyski'],
        ])
        numeric_columns.extend(list(tvm_cost_cols.values()))
        numeric_columns.append(koszty_tvm_col)
        numeric_columns.extend(list(other_cost_cols.values()))
        numeric_columns.append(dodatkowe_koszty_col)
        numeric_columns.extend([
            podsum_cols['koszty'],
            podsum_cols['przychody'],
            podsum_cols['brutto_suma'],
            podsum_cols['netto_suma'],
            podsum_cols['karta_suma'],
            podsum_cols['blik_suma'],
            podsum_cols['gotowka_suma'],
            podsum_cols['bilety_suma'],
            podsum_cols['profit_loss'],
        ])

        for col_num in numeric_columns:
            col_letter = get_column_letter(col_num)
            summary_cell = ws.cell(
                row=summary_row,
                column=col_num,
                value=f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})",
            )
            if col_num in ticket_columns.values() or col_num == podsum_cols['bilety_suma']:
                summary_cell.number_format = '#,##0'
            else:
                summary_cell.number_format = '#,##0.00'

        for col_num in range(1, uwagi_col_idx + 1):
            cell = ws.cell(row=summary_row, column=col_num)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        row_num += 1

    if last_data_row >= 2:
        result_col_letter = get_column_letter(result_col_idx)
        result_range = f"{result_col_letter}{data_start_row}:{result_col_letter}{last_data_row}"
        ws.conditional_formatting.add(
            result_range,
            CellIsRule(
                operator='greaterThanOrEqual',
                formula=['0'],
                fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            ),
        )
        ws.conditional_formatting.add(
            result_range,
            CellIsRule(
                operator='lessThan',
                formula=['0'],
                fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            ),
        )

    for r in range(data_start_row, last_data_row + 1):
        device_id = row_to_device_id.get(r)
        is_warehouse_row = device_id in warehouse_device_ids
        for c in range(info_nr_col, uwagi_col_idx + 1):
            row_fill = data_fill
            if is_warehouse_row and c not in {num1_col, num2_col, num3_col, num4_col, num5_col, num6_col, num7_col}:
                row_fill = warehouse_fill
            _set_cell_style(
                r,
                c,
                fill=row_fill,
                font=base_font,
                alignment=Alignment(horizontal='center' if c != info_loc_col and c != uwagi_col_idx else 'left', vertical='center'),
                border=thin_border,
            )

    block_end_row = last_data_row if last_data_row >= data_start_row else header_row
    number_blocks = [
        (num1_col, '1'),
        (num2_col, '2'),
        (num3_col, '3'),
        (num4_col, '4'),
        (num5_col, '5'),
        (num6_col, '6'),
        (num7_col, '7'),
    ]
    for num_col, label in number_blocks:
        ws.cell(row=title_row, column=num_col, value=label)
        _set_cell_style(
            title_row,
            num_col,
            fill=number_block_fill,
            font=Font(bold=True, size=18),
            alignment=Alignment(horizontal='center', vertical='center'),
            border=Border(left=medium_side, right=medium_side, top=medium_side, bottom=medium_side),
        )
        ws.merge_cells(start_row=title_row, start_column=num_col, end_row=block_end_row, end_column=num_col)

    ws.row_dimensions[title_row].height = 32
    ws.row_dimensions[header_row].height = 22
    for r in range(data_start_row, last_data_row + 1):
        ws.row_dimensions[r].height = 19

    # Dynamic column sizing (+ 10% margin)
    for col_num in range(1, uwagi_col_idx + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=col_num).value
            if cell_val is not None:
                lines = str(cell_val).split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
        if max_len > 0:
            # Special case to give a bit more for small columns and limit max width
            new_width = max_len * 1.1
            ws.column_dimensions[get_column_letter(col_num)].width = max(new_width, 4)
    
    # Set fixed width for Uwagi column
    ws.column_dimensions[get_column_letter(uwagi_col_idx)].width = 18

    # Set Nr TVM column width to header length + 10 characters margin
    try:
        nr_header = header_values.get(info_nr_col, 'Nr TVM')
        ws.column_dimensions[get_column_letter(info_nr_col)].width = max(len(str(nr_header)) + 10, 6)
    except Exception:
        pass

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.freeze_panes = f"{get_column_letter(info_type_col)}{data_start_row}"
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    if save_workbook and filepath is not None:
        _create_wyszukiwarka_sheet(wb)
        wb.save(filepath)
        print(f"\n✓ Raport P&L eksportowany: {filepath}")
    print(f"✓ Liczba wierszy: {len(all_device_ids)}")
    return {
        'profit_loss_by_device': profit_loss_by_device,
        'tickets_by_device': tickets_by_device,
        'all_device_ids': all_device_ids,
        'sheet_title': ws.title,
        'workbook': wb,
        'filepath': filepath,
    }


def _create_tickets_summary_sheet(wb, month_to_tickets):
    """
    Tworzy arkusz zbiorczy wg. biletów z podziałem na automat / miesiąc.
    Kolorowanie (ColorScale = Red>Yellow>Green) odbywa się na poziomie pojedynczego wiersza.
    """
    ws = wb.create_sheet(title='Tickets_Summary')
    ordered_months = sorted(month_to_tickets.keys())

    # Group months by year
    years_to_months = {}
    for m in ordered_months:
        year = int(m.split('-')[0])
        years_to_months.setdefault(year, []).append(m)

    # Build headers with per-year grouped months and yearly SUM, plus final Average
    headers = ['Nr aut.']
    month_col_positions = {}
    year_to_cols = {}
    current_col = 2
    for year in sorted(years_to_months.keys()):
        start_col = current_col
        for m in years_to_months[year]:
            headers.append(f"Bilety {_month_label_pl(m)}")
            month_col_positions[m] = current_col
            current_col += 1
        headers.append('Zestawienie roczne (+/-)')
        end_col = current_col
        year_to_cols[year] = (start_col, end_col)
        current_col += 1

    headers.append('Średnia')

    # Create year header row with merged cells
    ws.cell(row=1, column=1).value = ''
    ws.cell(row=1, column=1).fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    for year, (start_col, end_col) in year_to_cols.items():
        if start_col == end_col:
            year_cell = ws.cell(row=1, column=start_col)
        else:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            year_cell = ws.cell(row=1, column=start_col)
        year_cell.value = f"Rok {year}"
        year_cell.font = Font(bold=True, color='FFFFFF', size=12)
        year_cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        year_cell.alignment = Alignment(horizontal='center', vertical='center')

    sum_fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    sum_font = Font(bold=True, color='000000')
    sum_data_alignment = Alignment(horizontal='right', vertical='center')

    # Column headers (row 2)
    header_row = 2
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        if col_num == 1:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        elif header in ('Zestawienie roczne (+/-)', 'Średnia'):
            cell.font = sum_font
            cell.fill = sum_fill
        else:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    all_device_ids = sorted({
        int(device_id)
        for month_map in month_to_tickets.values()
        for device_id in month_map.keys()
    })

    row_num = 3
    total_months = len(ordered_months)
    for device_id in all_device_ids:
        ws.cell(row=row_num, column=1).value = device_id
        col_idx = 2
        # Fill monthly values grouped by year and insert yearly SUM formula
        for year in sorted(years_to_months.keys()):
            first_month_col = get_column_letter(col_idx)
            for m in years_to_months[year]:
                value = int(month_to_tickets.get(m, {}).get(device_id, 0) or 0)
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.number_format = '#,##0'
                col_idx += 1

            last_month_col = get_column_letter(col_idx - 1)
            sum_cell = ws.cell(row=row_num, column=col_idx)
            sum_cell.value = f"=SUM({first_month_col}{row_num}:{last_month_col}{row_num})"
            sum_cell.number_format = '#,##0'
            sum_cell.fill = sum_fill
            sum_cell.font = sum_font
            sum_cell.alignment = sum_data_alignment
            col_idx += 1

        # Average across all month columns (exclude the yearly SUM columns)
        if total_months > 0:
            first_value_col_letter = get_column_letter(2)
            last_month_col_letter = get_column_letter(2 + total_months - 1)
            avg_cell = ws.cell(row=row_num, column=col_idx)
            avg_cell.value = f'=AVERAGEIF({first_value_col_letter}{row_num}:{last_month_col_letter}{row_num}, "<>0")'
            avg_cell.number_format = '#,##0.00'
            avg_cell.fill = sum_fill
            avg_cell.font = sum_font
            avg_cell.alignment = sum_data_alignment

        # Color scale for monthly values only (exclude yearly and average columns)
        if total_months > 1:
            for year in sorted(years_to_months.keys()):
                month_cols = [month_col_positions[m] for m in years_to_months[year] if m in month_col_positions]
                if len(month_cols) <= 1:
                    continue
                year_month_range = f"{get_column_letter(month_cols[0])}{row_num}:{get_column_letter(month_cols[-1])}{row_num}"
                scale_rule = ColorScaleRule(start_type='min', start_color='FFC7CE',
                                            mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                                            end_type='max', end_color='C6EFCE')
                ws.conditional_formatting.add(year_month_range, scale_rule)

        row_num += 1

    last_data_row = row_num - 1

    # Podsumowanie dla każdego miesiąca (Suma)
    sum_row = row_num
    ws.cell(row=sum_row, column=1).value = "Suma biletów"
    ws.cell(row=sum_row, column=1).font = sum_font
    ws.cell(row=sum_row, column=1).fill = sum_fill
    for col in range(2, len(headers) + 1):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=sum_row, column=col)
        cell.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
        cell.number_format = '#,##0'
        cell.font = sum_font
        cell.fill = sum_fill
        if ws.cell(row=2, column=col).value in ('Zestawienie roczne (+/-)', 'Średnia'):
            cell.alignment = sum_data_alignment

    row_num += 1

    thin_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in range(2, row_num):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border

    # Additional April conditional formatting (column-wise, per each April present)
    april_months = [m for m in ordered_months if m.endswith('-04')]
    if row_num > 3:
        for april_month in april_months:
            april_col = month_col_positions.get(april_month)
            if april_col is None:
                continue
            april_range = f"{get_column_letter(april_col)}3:{get_column_letter(april_col)}{row_num - 1}"
            april_scale = ColorScaleRule(start_type='min', start_color='FFC7CE',
                                         mid_type='percentile', mid_value=50, mid_color='FFEB9C',
                                         end_type='max', end_color='C6EFCE')
            ws.conditional_formatting.add(april_range, april_scale)

    # Adjust column widths (10% extra margin)
    for col in range(1, len(headers) + 1):
        max_len = 0
        for row in range(2, row_num):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        if max_len > 0:
            ws.column_dimensions[get_column_letter(col)].width = max_len * 1.1

    # Tabela sezonowości
    curr_row = row_num + 2
    for year in sorted(years_to_months.keys()):
        months_in_year = years_to_months[year]
        if not months_in_year:
            continue
            
        start_col, end_col = year_to_cols[year]
        
        ws.cell(row=curr_row, column=1).value = f"Wskaźnik sezonowości Rok {year}"
        ws.cell(row=curr_row, column=1).font = Font(bold=True)
        ws.cell(row=curr_row+1, column=1).value = "Miesiąc"
        ws.cell(row=curr_row+2, column=1).value = "Suma biletów"
        ws.cell(row=curr_row+3, column=1).value = "Średnia miesięczna z przypisanych"
        ws.cell(row=curr_row+4, column=1).value = "Odchylenie (+/-)"
        ws.cell(row=curr_row+5, column=1).value = "Wskaźnik (%)"
        
        for r_offset in range(1, 6):
            ws.cell(row=curr_row+r_offset, column=1).font = Font(bold=True)
            
        col_idx = 2
        first_month_col_letter = get_column_letter(col_idx)
        last_month_col_letter = get_column_letter(col_idx + len(months_in_year) - 1)
        avg_formula = f'=AVERAGEIF({first_month_col_letter}{curr_row+2}:{last_month_col_letter}{curr_row+2}, "<>0")'
        
        for i, m in enumerate(months_in_year):
            c_current = col_idx + i
            c_letter = get_column_letter(c_current)
            orig_col_letter = get_column_letter(start_col + i)
            
            # Miesiąc
            c_month = ws.cell(row=curr_row+1, column=c_current)
            c_month.value = _month_label_pl(m)
            c_month.alignment = Alignment(horizontal='center')
            
            # Suma biletów
            c_wynik = ws.cell(row=curr_row+2, column=c_current)
            c_wynik.value = f"={orig_col_letter}{sum_row}"
            c_wynik.number_format = '#,##0'
            
            # Średnia
            c_srednia = ws.cell(row=curr_row+3, column=c_current)
            c_srednia.value = avg_formula
            c_srednia.number_format = '#,##0.00'
            
            # Odchylenie
            c_odchylenie = ws.cell(row=curr_row+4, column=c_current)
            c_odchylenie.value = f"={c_letter}{curr_row+2}-{c_letter}{curr_row+3}"
            c_odchylenie.number_format = '#,##0.00'
            
            # Wskaźnik
            c_wskaznik = ws.cell(row=curr_row+5, column=c_current)
            c_wskaznik.value = f"=IFERROR({c_letter}{curr_row+2}/{c_letter}{curr_row+3}, 0)"
            c_wskaznik.number_format = '0.00%'
            
        curr_row += 7

    ws.sheet_view.showGridLines = False


def _create_wyszukiwarka_sheet(wb):
    """
    Tworzy arkusz 'Wyszukiwarka' (sam layout wizualny) dla mechanizmu VBA.
    """
    ws = wb.create_sheet(title='Wyszukiwarka')
    
    # 1. Styl nagłówka ogólnego
    ws.merge_cells('A1:T2')
    title_cell = ws['A1']
    title_cell.value = 'WYSZUKIWARKA TVM - PORÓWNANIE DWÓCH OKRESÓW'
    title_cell.font = Font(bold=True, color='FFFFFF', size=16)
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Styl dla ramek dookoła
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

    # Kolumny A-D (wspólne)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # 2. Panel wejściowy (Nr automatu, Okres 1, Okres 2)
    ws['A4'].value = 'NR AUTOMATU:'
    ws['A4'].font = Font(bold=True)
    ws['A4'].border = thin_border
    
    ws['B4'].value = '1132'  # Przykładowa/domyślna wartość
    ws['B4'].alignment = Alignment(horizontal='center')
    ws['B4'].border = thin_border

    # Tabela z wyborem roków/miesięcy
    ws['B6'].value = 'ROK'
    ws['B6'].border = thin_border
    ws['B6'].alignment = Alignment(horizontal='center')
    
    ws['C6'].value = 'Miesiąc'
    ws['C6'].border = thin_border
    ws['C6'].alignment = Alignment(horizontal='center')

    ws['A7'].value = 'OKRES 1:'
    ws['A7'].font = Font(bold=True)
    ws['A7'].border = thin_border

    ws['B7'].value = '2026'
    ws['B7'].alignment = Alignment(horizontal='center')
    ws['B7'].border = thin_border
    
    ws['C7'].value = 'Luty'
    ws['C7'].alignment = Alignment(horizontal='center')
    ws['C7'].border = thin_border

    ws['A8'].value = 'OKRES 2:'
    ws['A8'].font = Font(bold=True)
    ws['A8'].border = thin_border

    ws['B8'].value = '2025'
    ws['B8'].alignment = Alignment(horizontal='center')
    ws['B8'].border = thin_border

    ws['C8'].value = 'Sierpień'
    ws['C8'].alignment = Alignment(horizontal='center')
    ws['C8'].border = thin_border

    # 3. Tabela wynikowa
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    headers = ['Brutto', 'Netto', 'Przychody', 'Koszty']
    for idx, text in enumerate(headers, start=2):  # Kolumny B, C, D, E (index 2-5)
        cell = ws.cell(row=12, column=idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Komórki z etykietami okresów w tabeli i obramowaniami
    ws['A13'].value = 'Luty 2026' # Do podmiany przez VBA
    ws['A13'].font = header_font
    ws['A13'].fill = header_fill
    ws['A13'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A13'].border = thin_border

    ws['A14'].value = 'Sierpień 2025'   # Do podmiany przez VBA
    ws['A14'].font = header_font
    ws['A14'].fill = header_fill
    ws['A14'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A14'].border = thin_border

    for row in [13, 14]:
        for col in range(2, 6): # B do E
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.number_format = '#,##0.00'

    ws.sheet_view.showGridLines = False


def export_multi_month_PL(
    dictionary_comparison,
    month_payloads,
    filename,
    carriers=None,
    costs_file=DEFAULT_COSTS_FILE,
    rent_file=DEFAULT_RENT_FILE,
    rent_sheet=None,
    amortyzacja_file=DEFAULT_AMORTYZACJA_FILE,
    amortyzacja_sheet=DEFAULT_AMORTYZACJA_SHEET,
):
    """
    Eksportuje wiele miesięcy do jednego skoroszytu.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    filepath = OUTPUT_DIR / filename
    wb = Workbook()

    for month_str, payload in month_payloads:
        warehouse_ids = payload.get('warehouse_device_ids') or set()

        export_to_excel_PL(
            dictionary_comparison,
            payload['revenue_data'],
            month_str,
            filename=None,
            commission_rules=payload.get('commission_rules') or {},
            carriers=carriers,
            location_by_device=payload.get('location_by_device') or {},
            automat_type_by_device=payload.get('automat_type_by_device') or {},
            costs_file=costs_file,
            rent_file=rent_file,
            rent_sheet=rent_sheet,
            amortyzacja_file=amortyzacja_file,
            amortyzacja_sheet=amortyzacja_sheet,
            workbook=wb,
            sheet_title=f"P&L {month_str}",
            save_workbook=False,
            service_cost_by_device=payload.get('service_cost_by_device') or {},
            it_card_switch_dates=payload.get('it_card_switch_dates') or {},
            relocation_active_device_ids=payload.get('relocation_active_device_ids') or set(),
            warehouse_device_ids=warehouse_ids,
            included_device_ids=payload.get('included_device_ids') or [],
        )
    _create_wyszukiwarka_sheet(wb)
    
    # Zapisz jako .xlsx najpierw
    temp_xlsx_path = filepath.with_suffix('.xlsx')
    wb.save(temp_xlsx_path)
    
    # Ładuj kody VBA z plików
    vba_modules_dict = _get_vba_modules_dict()
    
    # Konwertuj .xlsx → .xlsm z VBA
    _convert_xlsx_to_xlsm_with_vba(temp_xlsx_path, filepath, vba_modules_dict)
    print(f"\n✓ Raport P&L eksportowany (wieloarkuszowy): {filepath}")


def _build_month_export_payload(conn, month_str, args, dictionary_comparison, revenue_data_override=None):
    """
    Buduje dane wejściowe do eksportu dla wskazanego miesiąca.
    """
    if revenue_data_override is not None:
        revenue_data = revenue_data_override
    else:
        revenue_month_str = month_str

        if args.obrot_source_mode == 'carrier-transactions':
            revenue_data = get_monthly_revenue_by_carrier(
                conn,
                revenue_month_str,
                carriers=args.obrot_carriers,
                table_name=args.obrot_transactions_table,
                amount_col=args.obrot_transactions_amount_col,
                date_col=args.obrot_transactions_date_col,
                device_col=args.obrot_transactions_device_col,
                payment_method_col=args.obrot_transactions_payment_method_col,
                use_device_range_filter=not args.obrot_no_device_range_filter,
            )
        else:
            revenue_data = get_monthly_revenue(
                conn,
                revenue_month_str,
                schema=args.schema,
                actioncodes=args.obrot_actioncodes,
                mctype=args.obrot_mctype,
                use_device_range_filter=not args.obrot_no_device_range_filter,
            )

    commission_rules, xlsx_location_by_device, contract_summary_by_device = load_commission_rules_and_locations_from_xlsx(
        month_str,
        commission_file=args.prowizja_file,
    )

    candidate_device_ids = sorted(set(
        [int(k) for k in dictionary_comparison.keys()] +
        [int(k) for k in revenue_data.keys()]
    ))
    included_device_ids, excluded_device_details = _filter_device_ids_for_report_month(
        candidate_device_ids,
        revenue_data,
        contract_summary_by_device,
        month_str,
    )
    _log_device_filter_summary(month_str, included_device_ids, excluded_device_details)
    included_device_ids_set = set(included_device_ids)
    revenue_data = {
        int(device_id): entry
        for device_id, entry in (revenue_data or {}).items()
        if int(device_id) in included_device_ids_set
    }

    relocation_location_by_device, relocation_active_device_ids = _load_relocations_for_month(
        month_str,
        relokacje_file=args.relokacje_file,
        device_ids=included_device_ids,
    )

    # 1) Prefer locations from "lista automatow" (highest priority)
    resolved_lista_file = _resolve_lista_automatow_file_for_month(
        month_str,
        getattr(args, 'lista_automatow_file', DEFAULT_LISTA_AUTOMATOW_FILE),
    )
    lista_location_map = _load_lista_automatow_locations_from_xlsx(
        lista_file=resolved_lista_file,
        month_str=month_str,
    )

    # Initialize from lista (priority), then fill from xlsx (prowizje) if missing
    location_by_device = {}
    for device_id, location in lista_location_map.items():
        if device_id in included_device_ids_set and location:
            location_by_device[device_id] = location

    for device_id, location in xlsx_location_by_device.items():
        if device_id in included_device_ids_set and device_id not in location_by_device and location:
            location_by_device[device_id] = location

    # Apply relocations (overrides lista/xlsx)
    for device_id, location in relocation_location_by_device.items():
        if device_id in included_device_ids_set and location:
            location_by_device[device_id] = location

    # Service overrides (as before) — may override previous values
    year = int(month_str.split('-')[0])
    service_cost_by_device = {}
    if year >= 2026:
        resolved_service_file = _resolve_service_file_for_month(month_str, args.serwis_file)
        service_location_map, service_cost_map = _load_service_overrides_from_xlsx(
            service_file=resolved_service_file,
            sheet_index=args.serwis_sheet_index,
            month_str=month_str,
        )
        for device_id, location in service_location_map.items():
            if device_id in included_device_ids_set and location:
                location_by_device[device_id] = location
        service_cost_by_device = {
            int(device_id): float(value or 0.0)
            for device_id, value in service_cost_map.items()
            if int(device_id) in included_device_ids_set
        }

    missing_location_ids = [device_id for device_id in included_device_ids if device_id not in location_by_device]
    if missing_location_ids:
        fallback_location_by_device = get_locations_by_carrier(
            conn,
            device_ids=missing_location_ids,
            carriers=args.obrot_carriers,
        )
        for device_id, location in fallback_location_by_device.items():
            if device_id not in location_by_device and location:
                location_by_device[device_id] = location

    automat_type_by_device = _load_automat_type_map_from_lista(
        month_str,
        getattr(args, 'lista_automatow_file', DEFAULT_LISTA_AUTOMATOW_FILE),
        included_device_ids,
    )

    _, warehouse_device_ids = _load_amortyzacja_by_device_for_month(month_str)
    it_card_switch_dates = {
        int(device_id): switch_date
        for device_id, switch_date in _load_it_card_switch_dates(args.it_card_file).items()
        if int(device_id) in included_device_ids_set
    }

    return {
        'revenue_data': revenue_data,
        'commission_rules': commission_rules,
        'location_by_device': location_by_device,
        'automat_type_by_device': automat_type_by_device,
        'service_cost_by_device': service_cost_by_device,
        'it_card_switch_dates': it_card_switch_dates,
        'relocation_active_device_ids': relocation_active_device_ids,
        'warehouse_device_ids': warehouse_device_ids,
        'included_device_ids': included_device_ids,
    }


def _get_connection_dbname(conn):
    """
    Bezpiecznie odczytuje nazwe bazy z aktywnego lub ostatnio uzywanego polaczenia.
    """
    if conn is None:
        return None
    try:
        return conn.info.dbname
    except Exception:
        return None


def _is_connection_closed(conn):
    """
    Zwraca True, gdy obiekt polaczenia jest pusty albo PostgreSQL juz go zamknal.
    """
    if conn is None:
        return True
    try:
        return bool(conn.closed)
    except Exception:
        return True


def _reconnect_db(conn, fallback_database_name=None, month_str=None, max_retries=5, retry_delay=8):
    """
    Odtwarza polaczenie DB, preferujac te sama baze co poprzednie polaczenie.
    """
    database_name = _get_connection_dbname(conn) or fallback_database_name
    try:
        if conn is not None:
            conn.close()
    except Exception:
        pass

    if month_str:
        print(f"WARN DB reconnect ({month_str}): lacze ponownie z baza {database_name or '[auto]'}")

    return connect_to_db(
        database_name=database_name,
        max_retries=max_retries,
        retry_delay=retry_delay,
    )


def _build_month_export_payload_with_retry(
    conn,
    month_str,
    args,
    dictionary_comparison,
    revenue_data_override=None,
    max_attempts=4,
):
    """
    Buduje payload miesiaca z prostym retry/reconnect, gdy polaczenie DB zostanie zerwane.
    Zwraca tuple: (payload, conn)
    """
    attempt = 1
    fallback_database_name = _get_connection_dbname(conn)
    while attempt <= max_attempts:
        try:
            if _is_connection_closed(conn):
                conn = _reconnect_db(
                    conn,
                    fallback_database_name=fallback_database_name,
                    month_str=month_str,
                )
                if conn is None:
                    raise psycopg2.OperationalError("Nie udalo sie odtworzyc polaczenia DB")
                fallback_database_name = _get_connection_dbname(conn) or fallback_database_name

            payload = _build_month_export_payload(
                conn,
                month_str,
                args,
                dictionary_comparison,
                revenue_data_override=revenue_data_override,
            )
            return payload, conn
        except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
            if attempt >= max_attempts:
                raise

            print(
                f"WARN DB reconnect ({month_str}): proba {attempt}/{max_attempts} nieudana: {e}"
            )
            try:
                conn = _reconnect_db(
                    conn,
                    fallback_database_name=fallback_database_name,
                    month_str=month_str,
                )
            except psycopg2.Error as reconnect_error:
                print(
                    f"WARN DB reconnect ({month_str}): blad ponownego laczenia "
                    f"(proba {attempt}/{max_attempts}): {reconnect_error}"
                )
                attempt += 1
                time.sleep(5)
                continue

            if conn is None:
                attempt += 1
                time.sleep(5)
                continue

            fallback_database_name = _get_connection_dbname(conn) or fallback_database_name
            attempt += 1

    raise psycopg2.OperationalError(f"Nie udalo sie zbudowac payloadu dla {month_str} po {max_attempts} probach")


# === MAIN ===

def main():
    parser = argparse.ArgumentParser(description='Miesięczne zestawienie TVM P&L')
    parser.add_argument(
        '--schema',
        type=str,
        default=SCHEMA,
        help='Schemat źródłowy dla dictionary i moneystats (domyślnie: AV)'
    )
    parser.add_argument(
        '--obrot-actioncodes',
        nargs='*',
        type=int,
        default=None,
        help='Filtr obrotu: lista ms_actioncode (np. --obrot-actioncodes 1 2)'
    )
    parser.add_argument(
        '--obrot-mctype',
        type=int,
        default=None,
        help='Filtr obrotu: ms_mctype (np. --obrot-mctype 4)'
    )
    parser.add_argument(
        '--obrot-no-device-range-filter',
        action='store_true',
        help='Wyłącza filtr zakresów urządzeń dla obrotu (1101-1141, 1201-1275, 1276-1278, 1286-1299)'
    )
    parser.add_argument(
        '--obrot-source-mode',
        choices=['carrier-transactions', 'moneystats'],
        default='carrier-transactions',
        help='Źródło obrotu: per schemat przewoźnika z <carrier>.transactions lub klasycznie z <schema>.moneystats'
    )
    parser.add_argument(
        '--obrot-carriers',
        nargs='*',
        default=CARRIERS,
        help='Lista przewoźników/schematów do pobrania (np. ARP IC KD KML KW LKA PR SKM)'
    )
    parser.add_argument(
        '--obrot-transactions-table',
        type=str,
        default=TRANSACTIONS_SOURCE_CONFIG['table'],
        help='Nazwa tabeli transakcji per przewoźnik (domyślnie: transactions)'
    )
    parser.add_argument(
        '--obrot-transactions-amount-col',
        type=str,
        default=TRANSACTIONS_SOURCE_CONFIG['amount_col'],
        help='Kolumna kwoty w tabeli transakcji (domyślnie: fin_nalezn)'
    )
    parser.add_argument(
        '--obrot-transactions-date-col',
        type=str,
        default=TRANSACTIONS_SOURCE_CONFIG['date_col'],
        help='Kolumna daty w tabeli transakcji (domyślnie: fin_data_sp)'
    )
    parser.add_argument(
        '--obrot-transactions-device-col',
        type=str,
        default=TRANSACTIONS_SOURCE_CONFIG['device_col'],
        help='Kolumna numeru automatu (domyślnie: tvm_tvm_id)'
    )
    parser.add_argument(
        '--obrot-transactions-payment-method-col',
        type=str,
        default=TRANSACTIONS_SOURCE_CONFIG['payment_method_col'],
        help='Kolumna metody płatności (domyślnie: fin_spos_opl)'
    )
    parser.add_argument(
        '--obrot-print-sql',
        action='store_true',
        help='Wypisz zapytania SQL per przewoźnik/schemat dla wskazanego miesiąca'
    )
    parser.add_argument(
        '--miesiac',
        type=str,
        help='Miesiąc raportu w formacie YYYY-MM (domyślnie: poprzedni miesiąc)',
        default=None
    )
    parser.add_argument('--prowizja-schema', type=str, default=COMMISSION_SOURCE_CONFIG['schema'], help='Schema tabeli prowizji (domyślnie: public)')
    parser.add_argument('--prowizja-table', type=str, default=COMMISSION_SOURCE_CONFIG['table'], help='Nazwa tabeli prowizji (domyślnie: provision)')
    parser.add_argument('--prowizja-device-col', type=str, default=COMMISSION_SOURCE_CONFIG['device_col'], help='Kolumna z numerem automatu (np. TVM_ID)')
    parser.add_argument('--prowizja-amount-col', type=str, default=COMMISSION_SOURCE_CONFIG['amount_col'], help='Kolumna z kwotą prowizji (np. provision_amount)')
    parser.add_argument('--prowizja-date-col', type=str, default=COMMISSION_SOURCE_CONFIG['date_col'], help='Kolumna daty początkowej')
    parser.add_argument('--prowizja-date-to-col', type=str, default=COMMISSION_SOURCE_CONFIG['date_to_col'], help='Kolumna daty końca obowiązywania (opcjonalnie)')
    parser.add_argument('--prowizja-provider-col', type=str, default=None, help='Kolumna z operatorem płatności (np. ELAVON)')
    parser.add_argument(
        '--prowizja-provider-values',
        nargs='*',
        default=None,
        help='Dozwolone wartości operatora płatności (np. ELAVON INTERCHANGE)'
    )
    parser.add_argument(
        '--strict-prowizja-month-window',
        action='store_true',
        help='Ścisły zakres dla prowizji: date_from >= 1 dzień miesiąca i date_to <= ostatni dzień miesiąca'
    )
    parser.add_argument(
        '--no-strict-prowizja-month-window',
        action='store_false',
        dest='strict_prowizja_month_window',
        help='Tryb nakładających się okresów (date_from < koniec_miesiąca i date_to >= początek_miesiąca)'
    )
    parser.add_argument(
        '--reference-provision-xls',
        type=str,
        default=None,
        help='Ścieżka do pliku .xls z raportu PROVISION (single sheet) do porównania'
    )
    parser.add_argument(
        '--prowizja-file',
        type=str,
        default=str(DEFAULT_PROWIZJE_FILE),
        help='Ścieżka do pliku Excel z prowizjami i lokalizacjami (domyślnie: Prowizje_AB.xlsx)'
    )
    parser.add_argument(
        '--lista-automatow-file',
        type=str,
        default=str(DEFAULT_LISTA_AUTOMATOW_FILE),
        help='Ścieżka do pliku Excel z listą automatów i nazwami stacji (domyślnie: lista automatów.xlsx)'
    )
    parser.add_argument(
        '--relokacje-file',
        type=str,
        default=str(DEFAULT_RELOKACJE_FILE),
        help='Ścieżka do pliku relokacji (domyślnie: Relokacje AB.txt)'
    )
    parser.add_argument(
        '--koszty-file',
        type=str,
        default=str(DEFAULT_COSTS_FILE),
        help='Ścieżka do pliku Excel z kosztami projektu'
    )
    parser.add_argument(
        '--najem-file',
        type=str,
        default=str(DEFAULT_RENT_FILE),
        help='Ścieżka do pliku Excel z czynszem i energią per automat'
    )
    parser.add_argument(
        '--najem-sheet',
        type=str,
        default=None,
        help='Nazwa arkusza z najmem (domyślnie: rok z --miesiac)'
    )
    parser.add_argument(
        '--amortyzacja-file',
        type=str,
        default=str(DEFAULT_AMORTYZACJA_FILE),
        help='Parametr legacy (nieużywany): amortyzacja jest czytana z Koszty/amortyzacja_25.xlsx lub Koszty/amortyzacja_26.xlsx'
    )
    parser.add_argument(
        '--amortyzacja-sheet',
        type=str,
        default=DEFAULT_AMORTYZACJA_SHEET,
        help='Parametr legacy (nieużywany): nowy loader czyta arkusze 1 i 2'
    )
    parser.add_argument(
        '--serwis-file',
        type=str,
        default=str(DEFAULT_SERWIS_FILE),
        help='Ścieżka do pliku Excel serwisu (kolumny: NR TVM, SERWIS KOSZT).'
    )
    parser.add_argument(
        '--serwis-sheet-index',
        type=int,
        default=DEFAULT_SERWIS_SHEET_INDEX,
        help='Numer arkusza (1-based) w pliku serwisu (nieużywany, pozostawiony dla kompatybilności)'
    )
    parser.add_argument(
        '--it-card-file',
        type=str,
        default=str(DEFAULT_IT_CARD_SWITCH_FILE),
        help='Ścieżka do pliku IT CARD z datą przejścia ELAVON -> IT CARD per automat'
    )
    parser.add_argument(
        '--output-naming',
        choices=['default', 'monitor-style'],
        default='default',
        help='Tryb nazewnictwa pliku wyjściowego'
    )
    parser.add_argument(
        '--month-payload-source',
        choices=['auto', 'fresh', 'cache-only', 'no-cache'],
        default='fresh',
        help='Źródło payloadu miesięcznego: auto (cache z fallback), fresh (pełne pobranie + zapis cache), cache-only (tylko cache), no-cache (pełne pobranie bez odczytu i bez zapisu cache)'
    )
    parser.add_argument(
        '--validate-provision-only',
        action='store_true',
        help='Uruchamia tylko walidację prowizji SQL vs plik PROVISION .xls (bez obrotu i bez eksportu)'
    )
    parser.set_defaults(strict_prowizja_month_window=True)
    args = parser.parse_args()

    normalized_carriers = []
    for carrier in args.obrot_carriers or []:
        code = _normalize_carrier_code(carrier)
        if not code:
            continue
        if code not in normalized_carriers:
            normalized_carriers.append(code)

    disallowed = sorted(code for code in normalized_carriers if code in DISALLOWED_CARRIERS)
    if disallowed:
        parser.error(
            f"Przewoźnik(i) niedozwoleni: {', '.join(disallowed)}. "
            f"Dozwolone wartości: {', '.join(CARRIERS)}"
        )
    args.obrot_carriers = normalized_carriers
    
    # Określ miesiąc raportowy
    if args.miesiac:
        month_str = args.miesiac
        # Walidacja formatu
        try:
            datetime.strptime(month_str, '%Y-%m')
        except ValueError:
            print(f"❌ Nieprawidłowy format miesiąca: {month_str}. Użyj formatu YYYY-MM")
            return
    else:
        # Domyślnie: poprzedni miesiąc
        today = datetime.now()
        prev_month = today - relativedelta(months=1)
        month_str = prev_month.strftime('%Y-%m')
    
    commission_source = {
        'schema': args.prowizja_schema,
        'table': args.prowizja_table,
        'device_col': args.prowizja_device_col,
        'amount_col': args.prowizja_amount_col,
        'date_col': args.prowizja_date_col,
        'date_to_col': args.prowizja_date_to_col,
        'provider_col': args.prowizja_provider_col,
        'provider_values': args.prowizja_provider_values
    }

    source_schema = args.schema

    print(f"\n{'='*60}")
    print(f"  Miesięczne zestawienie TVM P&L - {month_str}")
    print(f"  Schemat danych: {source_schema}")
    print(f"  Tryb payloadu miesięcznego: {args.month_payload_source}")
    print(f"{'='*60}\n")

    use_month_payload_cache_read = args.month_payload_source in ('auto', 'cache-only')
    use_month_payload_cache_write = args.month_payload_source in ('auto', 'fresh')
    
    # Połączenie z bazą (auto-detect bazy danych)
    try:
        conn = connect_to_db(database_name=None, max_retries=3, retry_delay=5)
        if conn is None:
            print("❌ Nie można nawiązać połączenia z bazą danych")
            raise SystemExit(1)
    except psycopg2.Error as e:
        print(f"❌ Błąd połączenia z bazą: {e}")
        raise SystemExit(1)

    if args.validate_provision_only:
        if not args.reference_provision_xls:
            print("❌ Dla trybu --validate-provision-only podaj --reference-provision-xls")
            conn.close()
            return

        print("\n[VALIDATION] SQL vs PROVISION XLS...")
        _, validation_exit_code = validate_provision_sql_vs_xls(
            conn,
            month_str,
            commission_source,
            args.reference_provision_xls,
            strict_month_window=args.strict_prowizja_month_window,
        )
        conn.close()
        if validation_exit_code == 0:
            print("\n✓ Walidacja zakończona: pełna zgodność SQL vs XLS")
        elif validation_exit_code == 1:
            print("\n⚠ Walidacja zakończona: wykryto różnice SQL vs XLS")
        return
    
    # === KROK 1: Słownik automatów (ARP.dictionary) ===
    print(f"\n[1/6] Pobieranie słownika automatów ({source_schema}.dictionary)...")
    current_snapshot = get_dictionary_snapshot(conn, schema=source_schema)
    print(f"✓ Pobrano {len(current_snapshot)} automatów ze słownika")
    
    # Porównaj z poprzednim miesiącem
    print("\n[2/6] Porównanie z poprzednim miesiącem...")
    previous_snapshot = load_previous_snapshot(month_str)
    
    if previous_snapshot is None:
        print("ℹ  Brak poprzedniego snapshot — pierwszy raport dla tego miesiąca")
        dictionary_comparison = {
            device_id: {'status': 'new', 'data': data}
            for device_id, data in current_snapshot.items()
        }
    else:
        dictionary_comparison = compare_snapshots(current_snapshot, previous_snapshot)
        new_count = sum(1 for v in dictionary_comparison.values() if v['status'] == 'new')
        changed_count = sum(1 for v in dictionary_comparison.values() if v['status'] == 'changed')
        unchanged_count = sum(1 for v in dictionary_comparison.values() if v['status'] == 'unchanged')
        deleted_count = sum(1 for v in dictionary_comparison.values() if v['status'] == 'deleted')
        
        print(f"✓ Nowe: {new_count}, Zmienione: {changed_count}, "
              f"Bez zmian: {unchanged_count}, Usunięte: {deleted_count}")
    
    # Zapisz snapshot dla bieżącego miesiąca
    save_snapshot(current_snapshot, month_str)
    
    # === KROK 3-5: Payload miesiąca (cache/fresh) ===
    print(f"\n[3/6] Przygotowanie payloadu miesiąca ({month_str})...")
    current_month_payload = None

    if use_month_payload_cache_read:
        current_month_payload = load_month_payload_cache(month_str, args, dictionary_comparison)
        if current_month_payload is None:
            print(f"CACHE MISS {month_str}")
            if args.month_payload_source == 'cache-only':
                print(f"❌ Brak cache payloadu dla {month_str} w trybie cache-only")
                conn.close()
                return

    if current_month_payload is None:
        if args.month_payload_source in ('fresh', 'no-cache'):
            print(f"CACHE BYPASS {month_str}: tryb {args.month_payload_source}")

        current_month_payload, conn = _build_month_export_payload_with_retry(
            conn,
            month_str,
            args,
            dictionary_comparison,
            revenue_data_override=None,
        )

        if use_month_payload_cache_write:
            save_month_payload_cache(month_str, args, dictionary_comparison, current_month_payload)

    revenue_data = current_month_payload.get('revenue_data') or {}
    commission_rules = current_month_payload.get('commission_rules') or {}
    location_by_device = current_month_payload.get('location_by_device') or {}
    automat_type_by_device = current_month_payload.get('automat_type_by_device') or {}
    included_device_ids = current_month_payload.get('included_device_ids') or sorted(
        int(device_id) for device_id in revenue_data.keys()
    )

    if args.obrot_source_mode == 'carrier-transactions':
        total_revenue = sum(
            carrier_values.get('obrot_brutto_zl', 0.0)
            for device_values in revenue_data.values()
            for carrier_values in device_values.get('by_carrier', {}).values()
        )
        total_transactions = sum(
            int(carrier_values.get('liczba_transakcji', 0) or 0)
            for device_values in revenue_data.values()
            for carrier_values in device_values.get('by_carrier', {}).values()
        )
        total_tickets = sum(
            int(carrier_values.get('liczba_biletow', 0) or 0)
            for device_values in revenue_data.values()
            for carrier_values in device_values.get('by_carrier', {}).values()
        )
    else:
        total_revenue = sum(v.get('obrot_brutto_zl', 0.0) for v in revenue_data.values())
        total_transactions = sum(int(v.get('liczba_transakcji', 0) or 0) for v in revenue_data.values())
        total_tickets = None

    print(f"✓ Obrót dla {len(revenue_data)} automatów")
    print(f"  Obrót brutto: {total_revenue:,.2f} zł")
    print(f"  Liczba transakcji: {total_transactions:,}")

    if total_tickets is not None:
        print(f"  Liczba bilet\u00f3w: {total_tickets:,}")

    print(f"\n[4/6] Przeliczanie prowizji dla {month_str}...")
    print(f"✓ Wczytano reguły prowizji: {len(commission_rules)} par (automat, przewoźnik)")

    commission_data = build_monthly_commission_data_from_rules(
        revenue_data,
        commission_rules,
        month_str,
    )
    total_commission = sum(v['prowizja_zl'] for v in commission_data.values())
    print(f"  Prowizja łączna (z pliku): {total_commission:,.2f} zł")

    if args.reference_provision_xls:
        print("  🔎 Porównanie danych prowizji vs plik referencyjny .xls...")
        reference_map = load_reference_provision_from_xls(args.reference_provision_xls)
        comparison = compare_commission_with_reference(commission_data, reference_map)
        if comparison is None:
            print("  ⚠ Pominięto porównanie (nie udało się odczytać pliku .xls)")
        else:
            print(f"  Brakujące w SQL: {len(comparison['missing_in_sql'])}")
            print(f"  Nadmiarowe w SQL: {len(comparison['extra_in_sql'])}")
            print(f"  Różnice kwot: {len(comparison['mismatches'])}")

    print(f"\n[5/6] Podsumowanie metadanych lokalizacji/typów...")
    all_device_ids = sorted({int(device_id) for device_id in included_device_ids})
    print(
        f"✓ Lokalizacje: {len(location_by_device)} / {len(all_device_ids)}, "
        f"Typy automatów: {len(automat_type_by_device)} / {len(all_device_ids)}"
    )

    # === KROK 5: Eksport do Excel ===
    print(f"\n[6/6] Eksport do Excel (wieloarkuszowy)...")
    report_months = _build_commission_report_months()
    month_payloads = []
    for report_month in report_months:
        if report_month == month_str:
            payload = current_month_payload
            month_payloads.append((report_month, payload))
            continue

        payload = None
        if use_month_payload_cache_read:
            payload = load_month_payload_cache(report_month, args, dictionary_comparison)
            if payload is None:
                print(f"CACHE MISS {report_month}")
                if args.month_payload_source == 'cache-only':
                    print(f"❌ Brak cache payloadu dla {report_month} w trybie cache-only")
                    conn.close()
                    return

        if payload is None:
            if args.month_payload_source in ('fresh', 'no-cache'):
                print(f"CACHE BYPASS {report_month}: tryb {args.month_payload_source}")
            payload, conn = _build_month_export_payload_with_retry(
                conn,
                report_month,
                args,
                dictionary_comparison,
                revenue_data_override=None,
            )
            if use_month_payload_cache_write:
                save_month_payload_cache(report_month, args, dictionary_comparison, payload)

        month_payloads.append((report_month, payload))

    filename = build_output_filename(month_str, naming_mode=args.output_naming)
    if month_str == '2026-03' and str(filename).lower().startswith('raport_tymczasowy'):
        filename = build_output_filename(month_str, naming_mode='default')

    if month_str == '2026-03':
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        for legacy_temp in OUTPUT_DIR.glob('Raport_tymczasowy*2026-03*.xlsx'):
            try:
                legacy_temp.unlink()
            except OSError:
                pass

    export_multi_month_PL(
        dictionary_comparison,
        month_payloads,
        filename,
        carriers=args.obrot_carriers,
        costs_file=args.koszty_file,
        rent_file=args.najem_file,
        rent_sheet=args.najem_sheet,
        amortyzacja_file=args.amortyzacja_file,
        amortyzacja_sheet=args.amortyzacja_sheet,
    )

    conn.close()
    


if __name__ == "__main__":
    main()
