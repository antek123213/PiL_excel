#!/usr/bin/env python3
"""
Script to update existing Excel files with new formatting:
1. Remove gridlines from all sheets
2. Change gray backgrounds to white
3. Replace numeric values with 'M' for warehouse devices
4. Add/ensure black borders on data cells
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent / 'output'

def update_excel_file(filepath):
    """Update an Excel file with new formatting."""
    print(f"\n📝 Updating {filepath.name}...")
    
    wb = load_workbook(filepath)
    
    for sheet_idx, ws in enumerate(wb.sheetnames):
        sheet = wb[ws]
        print(f"  Sheet {sheet_idx + 1}: '{sheet.title}'")
        
        # Remove gridlines
        sheet.sheet_view.showGridLines = False
        
        # Define styles
        thin_side = Side(style='thin', color='000000')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        
        # Find location column (usually column C with header "Lokalizacja automatu")
        location_col = None
        device_col = None
        info_end_col = None
        
        # Check header row (row 2 for P&L sheets, row 1 for summary)
        header_row_candidates = [1, 2]
        
        for header_row in header_row_candidates:
            for col in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=header_row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if 'Lokalizacja' in cell_value:
                        location_col = col
                    if 'Nr TVM' in cell_value or 'Nr aut.' in cell_value:
                        device_col = col
            
            if location_col and device_col:
                info_end_col = location_col
                break
        
        if location_col is None:
            print(f"    ⚠️  Location column not found, skipping warehouse processing")
        else:
            print(f"    Found location column: {get_column_letter(location_col)}")
            print(f"    Found device column: {get_column_letter(device_col) if device_col else 'N/A'}")
        
        # Process data rows
        data_start_row = 3 if location_col and location_col > 1 else 2
        
        # Identify warehouse rows
        warehouse_rows = set()
        if location_col:
            for row in range(data_start_row, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=location_col)
                if cell.value and isinstance(cell.value, str):
                    if 'magazyn' in cell.value.lower():
                        warehouse_rows.add(row)
                        if device_col:
                            device_id = sheet.cell(row=row, column=device_col).value
                            print(f"    Found warehouse: Device {device_id}, Row {row}")
        
        # Update all cells
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                # Skip header rows
                if cell.row <= 2:
                    continue
                
                # Add black borders
                if cell.border is None or not cell.border.left:
                    cell.border = thin_border
                
                # Change gray fills to white (for non-header rows)
                if cell.fill and cell.fill.start_color:
                    fill_color = str(cell.fill.start_color.rgb) if hasattr(cell.fill.start_color, 'rgb') else str(cell.fill.start_color.index)
                    # Match gray colors (D9D9D9, EDEDED, E6E6E6, D0D0D0)
                    if any(x in fill_color.upper() for x in ['D9D9D9', 'EDEDED', 'E6E6E6', 'D0D0D0']):
                        cell.fill = white_fill
                
                # Replace warehouse values with 'M'
                if cell.row in warehouse_rows and cell.column not in [device_col, location_col] if device_col and location_col else True:
                    # Skip label columns (first few columns like device ID, location, notes)
                    if info_end_col and cell.column > info_end_col:
                        # Check if it's a numeric cell (not header, not text label)
                        if cell.value is not None and not isinstance(cell.value, str):
                            cell.value = 'M'
                            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"    ✓ Updated {sheet.title}")
    
    wb.save(filepath)
    print(f"✓ File saved: {filepath.name}")

def main():
    print("🔄 Updating Excel files with new formatting...")
    
    if not OUTPUT_DIR.exists():
        print(f"❌ Output directory not found: {OUTPUT_DIR}")
        return
    
    excel_files = list(OUTPUT_DIR.glob('*.xlsx'))
    
    if not excel_files:
        print("No .xlsx files found in output folder")
        return
    
    for filepath in excel_files:
        try:
            update_excel_file(filepath)
        except Exception as e:
            print(f"❌ Error updating {filepath.name}: {e}")
            import traceback
            traceback.print_exc()
    
    print("\n✅ Excel formatting update complete!")

if __name__ == '__main__':
    main()
